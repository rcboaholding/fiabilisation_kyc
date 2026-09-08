from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import close_old_connections, models
from django.db.models import Q
from urllib.parse import urlencode

from .models import TauxEvolution, Devise
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.hashers import make_password
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth.views import PasswordChangeView
from django.contrib.messages.views import SuccessMessageMixin

from django.core.mail import send_mail, BadHeaderError

from .models import TauxEvolution_filiale
import json
import csv
import re


import openpyxl
from django.template.loader import get_template
try:
    from xhtml2pdf import pisa
except ImportError:
    pisa = None
from kyc.models import (
    Notation, Historique, Kyc_pm, Kyc_pp, TauxEvolution, DATEREV,
    DataQualityRule, DataQualityCondition, DataQualityRuleAudit, KycDocumentExtraction, KycExpiredDocumentScanMatch,
    KycDocumentMatchJob, KycDocumentMatchSettings, DOCUMENT_EXTRACTION_TYPE_CHOICES,
    KycFieldVisibilityConfig, KycDocumentType, Filiales, CLIENT_TYPE_CHOICES,
    DATA_QUALITY_FIELD_CHOICES, EmailReminderConfig, KycDocumentOcrJob,
    KycMatchValidatorRole, KycMatchDecision, KycScreeningAccess, TauxQualite,
)
from django.utils import timezone
from django.utils.decorators import method_decorator

from django.http import (HttpResponse, HttpResponseRedirect, HttpResponseForbidden,
                         JsonResponse, FileResponse, Http404)
from pathlib import Path
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils.encoding import force_bytes
from django.utils.http import urlsafe_base64_encode
from django.views.decorators.csrf import csrf_exempt
from oauthlib.oauth2.rfc6749.endpoints import token

from django.db.models import Max, Avg, Count, F, Q, Exists, OuterRef, Case, When, IntegerField, CharField
from django.db.models.functions import TruncDate, Length

                                                                               
CharField.register_lookup(Length)

from django.http import JsonResponse
from .models import Kyc_pp

from accounts.models import AuditEvent, Organe, ProfileV, UserLoginHistory
from accounts.audit import log_audit
from django.core.cache import cache
from kyc import forms
from kyc.forms import CustomUserCreationForm, LoginForm, ResetPasswordForm, UserEditForm, VoyageurProfileForm,\
    CambProfileForm,\
    ProfileModify, NotationForm, ProfileForm
from django.contrib.sessions.models import Session
from django.utils.timezone import now
from django.utils.dateparse import parse_date, parse_datetime

from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from .models import Kyc_pp
from django.conf import settings
from django.core.files.base import ContentFile
import os
import re
import sys
import subprocess
import hashlib
import math
import threading
import uuid
import zipfile
from .document_extraction import SUPPORTED_EXTENSIONS, extract_document_data, extract_pdf_grouped_documents, learn_document_keywords

class CachedPaginator(Paginator):
    @property
    def count(self):
        query_str = str(self.object_list.query)
        query_hash = hashlib.md5(query_str.encode('utf-8')).hexdigest()
        cache_key = f"paginator_count_{query_hash}"
        cached_count = cache.get(cache_key)
        if cached_count is not None:
            return cached_count
        actual_count = super().count
        cache.set(cache_key, actual_count, 300)                            
        return actual_count

def floor_one_decimal(value):
    return math.floor(value * 10) / 10


def _rate_floor_1dec(good, total):
    """Taux = good / total en %, **tronqué** à 0,1 % (jamais arrondi).

    Règle unique complétude / qualité pour éviter un « 100 % » trompeur :
    - 100,0 % uniquement si good == total (0 défaut, décidé sur les compteurs) ;
    - sinon plafonné à 99,9 % (protège d'un arrondi flottant à 100,0).
    Renvoie None si total nul."""
    if not total:
        return None
    if good >= total:
        return 100.0
    return min(floor_one_decimal(good / total * 100), 99.9)


def completeness_rate_r(empty_cells, total_cells):
    """Taux de complétude selon la méthodologie du script R `calcul_de_taux.r` :
    floor(100 * (1 - cellules_vides / total_cellules)). Renvoie None si total nul.
    (Identique à global_rate/champ_rate du script — troncature à l'entier.)"""
    if not total_cells:
        return None
    return float(math.floor(100 * (1 - empty_cells / total_cells)))


from django.db.models.functions import Trim as _Trim
from django.db.models import CharField as _CharField
_CharField.register_lookup(_Trim, "trim")


def empty_field_q(field_name):
    """Q() identifiant une valeur « vide » au sens du script R : NULL, chaîne vide
    ou composée uniquement d'espaces (équivalent de trimws(x) == "").
    Utilise le lookup __trim (LTRIM/RTRIM en SQL, portable SQLite/MSSQL) plutôt
    qu'un __regex : SQL Server n'a pas de REGEXP_LIKE natif."""
    from django.db.models import Q
    return (Q(**{f"{field_name}__isnull": True})
            | Q(**{f"{field_name}__trim": ""}))

def compliance_rate_floor(ok_count, total, fail_count=0):
    """Taux de conformité qualité, tronqué à 0,1 %. 100 % seulement si aucune
    anomalie (ok_count == total) ; sinon plafonné à 99,9 %. `fail_count` n'est
    plus nécessaire (conservé pour compat d'appel)."""
    return _rate_floor_1dec(ok_count, total)

def _pdf_link_callback(uri, rel):
    """Resolve static/media URIs for xhtml2pdf on local filesystem."""
    if uri.startswith(settings.MEDIA_URL):
        path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))
    elif uri.startswith(settings.STATIC_URL):
        path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))
        if not os.path.isfile(path):
            path = os.path.join(settings.BASE_DIR, "static", uri.replace(settings.STATIC_URL, ""))
    else:
        return uri

    if not os.path.isfile(path):
        raise Exception(f"Media URI must start with {settings.MEDIA_URL} or {settings.STATIC_URL}: {uri}")
    return path


def format_date_for_export(value, output_format="%d/%m/%Y", empty_value="-"):
    """Format date/datetime/string values safely for exports."""
    if value in (None, ""):
        return empty_value

    if hasattr(value, "strftime"):
        try:
            return value.strftime(output_format)
        except Exception:
            pass

    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return empty_value

        parsed_dt = parse_datetime(cleaned)
        if parsed_dt:
            return parsed_dt.strftime(output_format)

        parsed_d = parse_date(cleaned)
        if parsed_d:
            return parsed_d.strftime(output_format)

        for input_format in ("%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.strptime(cleaned, input_format).strftime(output_format)
            except ValueError:
                continue

        return cleaned

    return str(value)

def get_data_quality_field_options():
    return {
        'PP': [
            ('FILIALE', 'FILIALE'),
            ('AGENCE', 'AGENCE'),
            ('LIB_AGENCE', 'LIB_AGENCE'),
            ('EXPL', 'EXPL'),
            ('CLIENT', 'CLIENT'),
            ('CODAPE', 'CODAPE'),
            ('IDP', 'IDP'),
            ('PAYNAIS', 'PAYNAIS'),
            ('PROFESSION', 'PROFESSION'),
            ('ADRESSE', 'ADRESSE'),
            ('PAYS_RESID', 'PAYS_RESID'),
            ('NUMID', 'NUMID'),
            ('SALAIRE', 'SALAIRE'),
            ('ORIGINE_REV', 'ORIGINE_REV'),
            ('DATVALID', 'DATVALID'),
            ('DATNAIS', 'DATNAIS'),
            ('TEL', 'TEL'),
            ('DATOUV', 'DATOUV'),
            ('PPE', 'PPE'),
            ('DEVISE', 'DEVISE'),
            ('RESID', 'RESID'),
            ('DATEREV', 'DATEREV'),
            ('RISQUE', 'RISQUE'),
            ('BOITE_POSTALE', 'BOITE_POSTALE'),
            ('CONSENT_BIC', 'CONSENT_BIC'),
            ('EMPLOYEUR', 'EMPLOYEUR'),
            ('INTITULE_COMPTE', 'INTITULE_COMPTE'),
            ('LIEU_DELIVRANCE_CIN', 'LIEU_DELIVRANCE_CIN'),
        ],
        'PM': [
            ('FILIALE', 'FILIALE'),
            ('AGENCE', 'AGENCE'),
            ('LIB_AGENCE', 'LIB_AGENCE'),
            ('EXPL', 'EXPL'),
            ('CLIENT', 'CLIENT'),
            ('AGEC', 'AGEC'),
            ('CODAPE', 'CODAPE'),
            ('IDM', 'IDM'),
            ('RCSNO', 'RCSNO'),
            ('CAPITAL', 'CAPITAL'),
            ('CA', 'CA'),
            ('RESULTAT', 'RESULTAT'),
            ('ORIGINE_REV', 'ORIGINE_REV'),
            ('DATOUV', 'DATOUV'),
            ('TEL', 'TEL'),
            ('DEVISE', 'DEVISE'),
            ('RESID', 'RESID'),
            ('DATEREV', 'DATEREV'),
            ('PPE', 'PPE'),
            ('RISQUE', 'RISQUE'),
            ('ACTIONNAIRE', 'ACTIONNAIRE'),
            ('ADRESSE_SOCIALE', 'ADRESSE_SOCIALE'),
            ('BOITE_POSTALE', 'BOITE_POSTALE'),
            ('CONSENT_BIC', 'CONSENT_BIC'),
            ('INTITULE_COMPTE', 'INTITULE_COMPTE'),
            ('MANDATAIRE', 'MANDATAIRE'),
            ('NUMERO_FISCAL', 'NUMERO_FISCAL'),
            ('PAYS_JUR', 'PAYS_JUR'),
        ],
    }


def evaluate_data_quality_scope(user):
    """Détermine le périmètre de calcul qualité selon l'organe utilisateur."""
    organe = (getattr(user, 'organe', '') or '').strip()
    filiale = (getattr(user, 'filiale', '') or '').strip()
    agence = (getattr(user, 'agence', '') or '').strip()
    code_expl = (getattr(user, 'code_expl', '') or '').strip()

    if organe == 'Chargé Client':
        return {
            'filiale': filiale or None,
            'agence': agence or None,
            'expl': code_expl or None,
            'label': 'Mon portefeuille',
        }
    if organe == 'Directeur Agence':
        return {
            'filiale': filiale or None,
            'agence': agence or None,
            'expl': None,
            'label': f"Agence {agence}" if agence else 'Mon agence',
        }
    if organe == 'PASS' or 'Groupe' in organe:
        return {
            'filiale': None,
            'agence': None,
            'expl': None,
            'label': 'GROUPE (toutes filiales)',
        }

    return {
        'filiale': filiale or None,
        'agence': None,
        'expl': None,
        'label': filiale or 'Ma filiale',
    }


def _quality_cache_version():
    return cache.get('quality_control_rules_version', 1)

def _rule_eval_filiale(rule, user_filiale):
    from kyc.forms import DataQualityRuleForm
    parsed = DataQualityRuleForm._parse_filiales(rule.filiale)
    if not parsed:
        return user_filiale
    if user_filiale:
        return user_filiale if user_filiale in parsed else None
    if len(parsed) == 1:
        return parsed[0]
    return None

DQ_LOGIC_AND = 'AND'
DQ_LOGIC_OR = 'OR'


def _dq_eval_condition(op, raw_val, raw_target, today, parse_date, calc_age):
    """Évalue UNE condition. Retourne True si elle 'matche' (contribue à l'anomalie)."""
    val = str(raw_val or '').strip()
    target = str(raw_target or '').strip()
    if op == '=': return val == target
    if op == '!=': return val != target
    if op == '>':
        try: return float(val.replace(',', '.')) > float(target.replace(',', '.'))
        except Exception: return False
    if op == '<':
        try: return float(val.replace(',', '.')) < float(target.replace(',', '.'))
        except Exception: return False
    if op == '>=':
        try: return float(val.replace(',', '.')) >= float(target.replace(',', '.'))
        except Exception: return False
    if op == '<=':
        try: return float(val.replace(',', '.')) <= float(target.replace(',', '.'))
        except Exception: return False
    if op == 'contains': return target.lower() in val.lower()
    if op == 'not_contains': return target.lower() not in val.lower()
    if op == 'word_contains':
        return bool(target and re.search(rf'(?<!\w){re.escape(target)}(?!\w)', val, re.IGNORECASE))
    if op == 'word_not_contains':
        return not bool(target and re.search(rf'(?<!\w){re.escape(target)}(?!\w)', val, re.IGNORECASE))
    if op == 'contains_alpha': return any(c.isalpha() for c in val)
    if op == 'contains_digit': return any(c.isdigit() for c in val)
    if op == 'is_empty': return not val
    if op == 'is_not_empty': return bool(val)
    if op == 'expired':
        p = parse_date(val)
        return bool(p and p < today)
    if op == 'age_gt':
        age = calc_age(val)
        try: return age is not None and age > int(target)
        except Exception: return False
    if op == 'age_lt':
        age = calc_age(val)
        try: return age is not None and age < int(target)
        except Exception: return False
    if op == 'min_length':
        try: return len(val) < int(target)
        except Exception: return False
    if op == 'max_length':
        try: return len(val) > int(target)
        except Exception: return False
    if op == 'regex':
        try:
            return re.search(target, val) is not None
        except re.error:
            return False
    return False


def _dq_safe_parse_date(value):
    if not value: return None
    if hasattr(value, 'date'): return value.date()
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y%m%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _dq_calculate_age(birth_date_str):
    parsed = _dq_safe_parse_date(birth_date_str)
    if not parsed: return None
    today_date = datetime.today().date()
    return today_date.year - parsed.year - ((today_date.month, today_date.day) < (parsed.month, parsed.day))


def _dq_conditions_flag(conditions, get_value, today, parse_date, calc_age):
    """Combine les conditions en forme normale disjonctive (DNF) :
      - le connecteur 'OU' (logic=OR) démarre un nouveau groupe,
      - au sein d'un groupe les conditions sont reliées en 'ET',
      - anomalie détectée si AU MOINS un groupe est entièrement vrai.
    Le connecteur porté par la 1re condition est ignoré."""
    result = False
    current = True
    started = False
    for cond in conditions:
        m = _dq_eval_condition(cond.operator, get_value(cond.field_name),
                               cond.value, today, parse_date, calc_age)
        logic = (getattr(cond, 'logic', DQ_LOGIC_AND) or DQ_LOGIC_AND)
        if not started:
            current = m
            started = True
        elif logic == DQ_LOGIC_OR:
            result = result or current
            current = m
        else:
            current = current and m
    return (result or current) if started else False


def _evaluate_data_quality_rule_scoped(rule, filiale=None, agence=None, expl=None):
    return evaluate_data_quality_rule(rule, filiale=filiale, agence=agence, expl=expl)

def get_incomplete_clients_queryset(queryset, client_type):
    from kyc.models import KycFieldVisibilityConfig
    from django.db.models import Q
    
    filiales = list(queryset.values_list('FILIALE', flat=True).distinct())
    if not filiales:
        return queryset.none()
        
    configs = list(KycFieldVisibilityConfig.objects.filter(client_type=client_type))
    
    combined_q = Q()
    for filiale in filiales:
        config = None
        if filiale:
            config = next((c for c in configs if filiale in (c.filiales or [])), None)
        if not config:
            config = next((c for c in configs if not c.filiales), None)
            
        if config and config.empty_check_fields:
            fields = config.empty_check_fields
        else:
            if client_type == 'pp':
                fields = ["NUMID", "DATNAIS", "ADRESSE", "TEL"]
            else:
                fields = ["NUMERO_FISCAL", "RCSNO", "ADRESSE_SOCIALE", "TEL"]
                
        field_q = Q()
        for f in fields:
            if f in ["CLIENT", "EXPL", "FILIALE", "AGENCE", "LIB_AGENCE"]:
                continue
            field_q |= Q(**{f"{f}__isnull": True}) | Q(**{f"{f}": ""})
            
        if filiale:
            combined_q |= Q(FILIALE=filiale) & field_q
        else:
            combined_q |= (Q(FILIALE__isnull=True) | Q(FILIALE="")) & field_q
            
    return queryset.filter(combined_q)

def flux_datouv_window(reference_date=None):
    """Bornes ISO (début, fin) incluses de la fenêtre « flux » configurée.

    Lit QualityFluxConfig (admin Django) : 'veille' = DATOUV d'hier uniquement,
    'mois' = mois calendaire précédent. DATOUV est stocké en ISO YYYY-MM-DD,
    la comparaison lexicale équivaut donc à la comparaison chronologique
    (même principe que DATEREV).
    """
    from kyc.models import QualityFluxConfig
    ref = reference_date or timezone.localdate()
    config = QualityFluxConfig.objects.filter(active=True).order_by('-updated_at').first()
    window = config.flux_window if config else 'veille'
    if window == 'mois':
        first_of_month = ref.replace(day=1)
        end = first_of_month - timedelta(days=1)                                      
        start = end.replace(day=1)                                                 
    else:
        start = end = ref - timedelta(days=1)                     
    return start.isoformat(), end.isoformat()


def apply_datouv_period_filter(queryset, request):
    """Filtre période sur DATOUV via les paramètres GET datouv_start / datouv_end.

    DATOUV est stocké en ISO YYYY-MM-DD (comme DATEREV) : la comparaison
    lexicale équivaut à la comparaison chronologique. Quand une borne de fin
    est fournie, on exclut les DATOUV vides ('' <= borne serait toujours vrai).
    """
    start = (request.GET.get('datouv_start') or '').strip()
    end = (request.GET.get('datouv_end') or '').strip()
    if start:
        queryset = queryset.filter(DATOUV__gte=start)
    if end:
        queryset = queryset.exclude(DATOUV='').filter(DATOUV__lte=end)
    return queryset


def evaluate_data_quality_rule(rule, filiale=None, agence=None, expl=None,
                               datouv_start=None, datouv_end=None):
    model = Kyc_pp if rule.applicability == 'PP' else Kyc_pm
    field_names = [f.name for f in model._meta.get_fields() if not f.many_to_many and not f.one_to_many]
    if rule.control_type != 'composite' and rule.field_name not in field_names and rule.control_type not in ['expired_document', 'codape_agec_match']:
        return {'total': 0, 'fail_count': 0, 'ok_count': 0, 'clients': [], 'message': 'Champ de contrôle invalide'}

    from kyc.forms import DataQualityRuleForm
    parsed_filiales = DataQualityRuleForm._parse_filiales(rule.filiale)

    queryset = model.objects.all()
    if filiale and filiale != 'GROUPE':
        if parsed_filiales and filiale not in parsed_filiales:
            return {'total': 0, 'fail_count': 0, 'ok_count': 0, 'clients': [], 'message': 'Non applicable à cette filiale'}
        queryset = queryset.filter(FILIALE=filiale)
    elif parsed_filiales:
        queryset = queryset.filter(FILIALE__in=parsed_filiales)

    if agence:
        queryset = queryset.filter(AGENCE=agence)
    if expl:
        queryset = queryset.filter(EXPL=expl)

                                                                              
                                                                                       
    if datouv_start and datouv_end:
        queryset = queryset.exclude(DATOUV='').filter(
            DATOUV__gte=datouv_start, DATOUV__lte=datouv_end,
        )

    total = queryset.count()
    if total == 0:
        return {'total': 0, 'fail_count': 0, 'ok_count': 0, 'clients': [], 'message': 'Aucune donnée disponible pour ce segment'}

    failures = []
    today = datetime.today().date()
    client_fields = ['CLIENT', 'EXPL', 'FILIALE', 'AGENCE']

    safe_parse_date = _dq_safe_parse_date
    calculate_age = _dq_calculate_age

    def build_clients_from_values(rows, value_key):
        return [{
            'client': row.get('CLIENT', ''),
            'expl': row.get('EXPL', ''),
            'filiale': row.get('FILIALE', ''),
            'agence': row.get('AGENCE', ''),
            'field_value': str(row.get(value_key, '') or ''),
        } for row in rows]

    if rule.control_type == 'simple':
        param = (rule.parameter or '').strip().lower()
        
                                                             
        if not param or param == 'existence':
                                         
            failures = queryset.filter(Q(**{f"{rule.field_name}__isnull": True}) | Q(**{f"{rule.field_name}": ""}))
            fail_count = failures.count()
            clients = build_clients_from_values(list(failures.values(*client_fields, rule.field_name)[:15]), rule.field_name)
        
        elif param.isdigit() or (param.startswith('len') or param.startswith('long')):
                          
            try:
                                                                    
                import re
                match = re.search(r'\d+', param)
                target_len = int(match.group()) if match else int(param)
                
                fail_count = 0
                clients = []
                for row in queryset.values(*client_fields, rule.field_name).iterator(chunk_size=2000):
                    val = str(row.get(rule.field_name) or '')
                    if len(val) != target_len:
                        fail_count += 1
                        if len(clients) < 15:
                            clients.append({
                                'client': row.get('CLIENT', ''),
                                'expl': row.get('EXPL', ''),
                                'filiale': row.get('FILIALE', ''),
                                'agence': row.get('AGENCE', ''),
                                'field_value': val,
                            })
            except:
                return {'total': total, 'fail_count': 0, 'ok_count': total, 'clients': [], 'message': 'Paramètre de longueur invalide'}
        
        else:
                               
            failures = queryset.exclude(**{f"{rule.field_name}": rule.parameter})
            fail_count = failures.count()
            clients = build_clients_from_values(list(failures.values(*client_fields, rule.field_name)[:15]), rule.field_name)

        ok_count = total - fail_count
        return {
            'total': total,
            'fail_count': fail_count,
            'ok_count': ok_count,
            'clients': clients,
            'message': '',
        }

    elif rule.control_type == 'composite':
        conditions = rule.conditions.all()
        if not conditions.exists():
            return {'total': total, 'fail_count': 0, 'ok_count': total, 'clients': [], 'message': 'Pas de conditions'}
        
        fail_count = 0
        clients = []
        needed_fields = set(client_fields)
        for c in conditions:
            needed_fields.add(c.field_name)
        for row in queryset.values(*needed_fields).iterator(chunk_size=2000):
            if _dq_conditions_flag(conditions, lambda f: row.get(f, ''), today, safe_parse_date, calculate_age):
                fail_count += 1
                if len(clients) < 15:
                    clients.append({
                        'client': row.get('CLIENT', ''),
                        'expl': row.get('EXPL', ''),
                        'filiale': row.get('FILIALE', ''),
                        'agence': row.get('AGENCE', ''),
                        'field_value': 'Multi-critères',
                    })
        
        ok_count = total - fail_count
        return {
            'total': total,
            'fail_count': fail_count,
            'ok_count': ok_count,
            'clients': clients,
            'message': '',
        }

    return {
        'total': total,
        'fail_count': 0,
        'ok_count': total,
        'clients': [],
        'message': '',
    }


@login_required
def quality_control_view(request):
    user = request.user
    allowed_organs = ['PASS']
    user_organe = (getattr(user, 'organe', '') or '').strip()
    if user_organe not in allowed_organs:
        messages.error(request, "Accès non autorisé au contrôle qualité.")
        return redirect('accueil')

    from kyc.forms import DataQualityRuleForm, DataQualityConditionFormSet

                                        
    user_organe = (getattr(request.user, 'organe', '') or '').strip()
    can_manage = user_organe == 'PASS'
    user_filiale = getattr(request.user, 'filiale', '')
    
    if user_organe == 'PASS':
        from kyc.models import Filiales as ModelFiliales
        filiale_choices = [f[0] for f in ModelFiliales]
    else:
        filiale_choices = [user_filiale] if user_filiale else []
    
    if request.method == 'POST' and can_manage:
        form = DataQualityRuleForm(request.POST, filiale_choices=filiale_choices)
        formset = DataQualityConditionFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            rule = form.save(commit=False)
            rule.created_by = request.user
            rule.save()
            formset.instance = rule
            formset.save()
            
                            
            DataQualityRuleAudit.objects.create(
                rule_name=rule.name,
                user=request.user,
                action='CREATION',
                details=f"Création de la règle '{rule.name}' ({rule.applicability})"
            )
            
            current_version = cache.get('quality_control_rules_version', 1)
            cache.set('quality_control_rules_version', current_version + 1, timeout=None)
            messages.success(request, 'Règle de qualité enregistrée.')
            return redirect('kyc:quality_control')
    else:
        form = DataQualityRuleForm(filiale_choices=filiale_choices)
        formset = DataQualityConditionFormSet()

    rules = list(
        DataQualityRule.objects.all()
        .order_by('id')
        .prefetch_related('conditions')
    )
    cache_ttl_seconds = 86400                                                      
    rules_version = cache.get('quality_control_rules_version', 1)
    data_refresh_bucket = timezone.localdate().isoformat()
    stats = []
    
                                                                            
    group_organs = ['PASS', 'Conformité Groupe']
    eval_filiale = None if user_organe in group_organs else user_filiale

    for rule in rules:
        rule_signature = f"{rule.id}|{rule.name}|{rule.applicability}|{rule.field_name}|{rule.control_type}|{rule.parameter}|{rule.active}|{eval_filiale}"
        rule_hash = hashlib.md5(rule_signature.encode('utf-8')).hexdigest()
        rule_cache_key = f"quality_control:stat:v{rules_version}:d{data_refresh_bucket}:{rule_hash}"
        stat = cache.get(rule_cache_key)
        if stat is None:
            stat = evaluate_data_quality_rule(rule, filiale=eval_filiale)
            cache.set(rule_cache_key, stat, timeout=cache_ttl_seconds)
        stats.append(stat)
    for stat in stats:
        total = stat.get('total', 0)
        stat['compliance_rate'] = compliance_rate_floor(stat['ok_count'], total, stat.get('fail_count', 0))
    from kyc.forms import DataQualityRuleForm

                                                                                                   
    can_pick_filiale = user_organe in group_organs
    filiale_filter = (request.GET.get('filiale') or '').strip()

    filiale_set = set()
    for rule in rules:
        for f in DataQualityRuleForm._parse_filiales(rule.filiale):
            if f:
                filiale_set.add(f)
    if can_pick_filiale:
        filiale_list = sorted(filiale_set)
    else:
        filiale_list = [user_filiale] if user_filiale else []
        filiale_filter = ''                                       

    from kyc.models import get_rule_number_map
    rule_number_map = get_rule_number_map()
    rules_with_stats = []
    for rule, stat in zip(rules, stats):
        rule_num = rule_number_map.get((rule.name or '').strip(), rule.id)
        rule._rule_number = rule_num
        parsed_filiales = DataQualityRuleForm._parse_filiales(rule.filiale)

        if filiale_filter and parsed_filiales and filiale_filter not in parsed_filiales:
            continue

        grouped_conditions = {}
        for cond in rule.conditions.all():
            group_key = (cond.field_name, cond.operator)
            if group_key not in grouped_conditions:
                grouped_conditions[group_key] = {
                    'field_name': cond.field_name,
                    'operator_display': cond.get_operator_display(),
                    'values': [],
                }

            value = (cond.value or '').strip()
            if value and value not in grouped_conditions[group_key]['values']:
                grouped_conditions[group_key]['values'].append(value)

        filiales_display = ", ".join(parsed_filiales) if parsed_filiales else "Toutes les filiales"

        conditions_ordered = [{
            'logic': cond.logic,
            'logic_display': cond.get_logic_display(),
            'field_name': cond.field_name,
            'operator_display': cond.get_operator_display(),
            'value': (cond.value or '').strip(),
        } for cond in rule.conditions.all()]

        rules_with_stats.append({
            'rule': rule,
            'stat': stat,
            'rule_number': rule_num,
            'condition_groups': list(grouped_conditions.values()),
            'conditions_ordered': conditions_ordered,
            'filiales_display': filiales_display,
        })

    field_options = get_data_quality_field_options()
    total_rules = len(rules_with_stats)
    active_rules = sum(1 for item in rules_with_stats if item['rule'].active)
    inactive_rules = total_rules - active_rules
    total_failures = sum(item['stat']['fail_count'] for item in rules_with_stats)
    total_ok = sum(item['stat']['ok_count'] for item in rules_with_stats)
    total_evaluated = sum(item['stat']['total'] for item in rules_with_stats)
    global_compliance_rate = compliance_rate_floor(total_ok, total_evaluated, total_failures)
    pp_rules_count = sum(1 for item in rules_with_stats if item['rule'].applicability == 'PP')
    pm_rules_count = sum(1 for item in rules_with_stats if item['rule'].applicability == 'PM')

    return render(request, 'quality_control.html', {
        'form': form,
        'formset': formset,
        'rules': rules_with_stats,
        'field_options': field_options,
        'total_rules': total_rules,
        'active_rules': active_rules,
        'inactive_rules': inactive_rules,
        'total_failures': total_failures,
        'global_compliance_rate': global_compliance_rate,
        'form_has_errors': bool(form.errors),
        'can_manage': can_manage,
        'user_organe': user_organe,
        'filiale_list': filiale_list,
        'can_pick_filiale': can_pick_filiale,
        'pp_rules_count': pp_rules_count,
        'pm_rules_count': pm_rules_count,
    })

@login_required
def delete_quality_rule(request, pk):
    user_organe = (getattr(request.user, 'organe', '') or '').strip()
    if user_organe not in ['PASS']:
        messages.error(request, "Accès refusé.")
        return redirect('kyc:quality_control')
        
    rule = get_object_or_404(DataQualityRule, pk=pk)
    
                                      
    if user_organe != 'PASS' and rule.created_by and rule.created_by.filiale != request.user.filiale:
        messages.error(request, "Vous ne pouvez supprimer que les règles de votre filiale.")
        return redirect('kyc:quality_control')

    rule_name = rule.name
    DataQualityRuleAudit.objects.create(
        rule_name=rule_name,
        user=request.user,
        action='SUPPRESSION',
        details=f"Suppression de la règle '{rule_name}'"
    )
    
    rule.delete()
    current_version = cache.get('quality_control_rules_version', 1)
    cache.set('quality_control_rules_version', current_version + 1, timeout=None)
    messages.success(request, f"Règle '{rule_name}' supprimée.")
    return redirect('kyc:quality_control')

@login_required
def edit_quality_rule(request, pk):
    user_organe = (getattr(request.user, 'organe', '') or '').strip()
    if user_organe not in ['PASS']:
        messages.error(request, "Accès refusé.")
        return redirect('kyc:quality_control')
        
    rule = get_object_or_404(DataQualityRule, pk=pk)
    
                                      
    if user_organe != 'PASS' and rule.created_by and rule.created_by.filiale != request.user.filiale:
        messages.error(request, "Vous ne pouvez modifier que les règles de votre filiale.")
        return redirect('kyc:quality_control')

    from kyc.forms import DataQualityRuleForm, DataQualityConditionFormSet
    
    if user_organe == 'PASS':
        from kyc.models import Filiales as ModelFiliales
        filiale_choices = [f[0] for f in ModelFiliales]
    else:
        filiale_choices = [request.user.filiale] if request.user.filiale else []
        
    if request.method == 'POST':
        form = DataQualityRuleForm(request.POST, instance=rule, filiale_choices=filiale_choices)
        formset = DataQualityConditionFormSet(request.POST, instance=rule)
        if form.is_valid() and formset.is_valid():
            changes = []
            if form.has_changed():
                for field in form.changed_data:
                    old = getattr(rule, field)
                    new = form.cleaned_data.get(field)
                    changes.append(f"{field}: {old} -> {new}")
            
            form.save()
            formset.save()
            
            DataQualityRuleAudit.objects.create(
                rule_name=rule.name,
                user=request.user,
                action='MODIFICATION',
                details="; ".join(changes) if changes else "Modification des conditions"
            )
            
            current_version = cache.get('quality_control_rules_version', 1)
            cache.set('quality_control_rules_version', current_version + 1, timeout=None)
            messages.success(request, "Règle mise à jour.")
            return redirect('kyc:quality_control')
    else:
        form = DataQualityRuleForm(instance=rule, filiale_choices=filiale_choices)
        formset = DataQualityConditionFormSet(instance=rule)
        
    return render(request, 'quality_rule_edit.html', {
        'form': form,
        'formset': formset,
        'rule': rule
    })

@login_required
def duplicate_quality_rule(request, pk):
    user_organe = (getattr(request.user, 'organe', '') or '').strip()
    if user_organe not in ['PASS']:
        messages.error(request, "Accès refusé.")
        return redirect('kyc:quality_control')

    source_rule = get_object_or_404(
        DataQualityRule.objects.prefetch_related('conditions'),
        pk=pk,
    )

    from django.forms import inlineformset_factory
    from kyc.forms import DataQualityRuleForm, DataQualityConditionForm, DataQualityConditionFormSet

    if user_organe == 'PASS':
        from kyc.models import Filiales as ModelFiliales
        filiale_choices = [f[0] for f in ModelFiliales]
    else:
        filiale_choices = [request.user.filiale] if request.user.filiale else []

    if request.method == 'POST':
        form = DataQualityRuleForm(request.POST, filiale_choices=filiale_choices)
        formset = DataQualityConditionFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            rule = form.save(commit=False)
            rule.created_by = request.user
            rule.save()
            formset.instance = rule
            formset.save()

            DataQualityRuleAudit.objects.create(
                rule_name=rule.name,
                user=request.user,
                action='DUPLICATION',
                details=(
                    f"Duplication de la règle #{source_rule.rule_number} "
                    f"'{source_rule.name}' vers '{rule.name}' ({rule.filiale or 'Toutes les filiales'})"
                ),
            )

            current_version = cache.get('quality_control_rules_version', 1)
            cache.set('quality_control_rules_version', current_version + 1, timeout=None)
            if (rule.name or '').strip() == (source_rule.name or '').strip():
                messages.success(request, "Règle dupliquée avec le même numéro. Vous pouvez ajuster ses conditions indépendamment.")
            else:
                messages.success(request, "Règle dupliquée avec un nouveau numéro car le nom a changé.")
            return redirect('kyc:quality_control')
    else:
        duplicated_rule = DataQualityRule(
            name=source_rule.name,
            applicability=source_rule.applicability,
            filiale=source_rule.filiale,
            description=source_rule.description,
            active=source_rule.active,
            control_type=source_rule.control_type,
            field_name=source_rule.field_name,
            parameter=source_rule.parameter,
        )
        form = DataQualityRuleForm(
            instance=duplicated_rule,
            filiale_choices=filiale_choices,
        )
        initial_conditions = [
            {
                'logic': condition.logic,
                'field_name': condition.field_name,
                'operator': condition.operator,
                'value': condition.value,
            }
            for condition in source_rule.conditions.all()
        ]
        DuplicateConditionFormSet = inlineformset_factory(
            DataQualityRule,
            DataQualityCondition,
            form=DataQualityConditionForm,
            fields=['logic', 'field_name', 'operator', 'value'],
            extra=max(len(initial_conditions), 1),
            can_delete=True,
        )
        formset = DuplicateConditionFormSet(
            initial=initial_conditions,
            queryset=DataQualityCondition.objects.none(),
        )

    return render(request, 'quality_rule_edit.html', {
        'form': form,
        'formset': formset,
        'rule': source_rule,
        'duplicate_mode': True,
    })

@login_required
def quality_control_audits(request):
    user_organe = (getattr(request.user, 'organe', '') or '').strip()
    if user_organe not in ['PASS']:
        messages.error(request, "Accès refusé.")
        return redirect('kyc:quality_control')
        
    if user_organe == 'PASS':
        audits = DataQualityRuleAudit.objects.all().order_by('-timestamp')
    else:
                                                                                
        audits = DataQualityRuleAudit.objects.filter(
            Q(user__filiale=request.user.filiale) | Q(user__isnull=True)
        ).order_by('-timestamp')
        
    from kyc.models import get_rule_number_map
    rule_number_map = get_rule_number_map()
    audits = list(audits)
    for audit in audits:
        raw_rule_name = (audit.rule_name or "")
        audit.rule_name_display = raw_rule_name.strip() or "N/A"
        audit.rule_number = rule_number_map.get(raw_rule_name.strip(), audit.id)
        audit.time_display = audit.timestamp.strftime("%H:%M:%S") if audit.timestamp else "--:--:--"
    return render(request, 'quality_control_audits.html', {'audits': audits})

@login_required
def export_audits_excel(request):
    user_organe = (getattr(request.user, 'organe', '') or '').strip()
    if user_organe not in ['PASS']:
        return HttpResponseForbidden()
        
    if user_organe == 'PASS':
        audits = DataQualityRuleAudit.objects.all().order_by('-timestamp')
    else:
        audits = DataQualityRuleAudit.objects.filter(
            Q(user__filiale=request.user.filiale) | Q(user__isnull=True)
        ).order_by('-timestamp')
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit des Contrôles"
    ws.append(["Date & Heure", "Utilisateur", "Règle", "Action", "Détails"])
    for audit in audits:
        ws.append([audit.timestamp.strftime("%d/%m/%Y %H:%M:%S"), audit.user.username if audit.user else "System", audit.rule_name, audit.action, audit.details])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=audit_controles.xlsx'
    wb.save(response)
    return response

@login_required
def export_audits_pdf(request):
    user_organe = (getattr(request.user, 'organe', '') or '').strip()
    if user_organe not in ['PASS']:
        return HttpResponseForbidden()
        
    if user_organe == 'PASS':
        audits = DataQualityRuleAudit.objects.all().order_by('-timestamp')
    else:
        audits = DataQualityRuleAudit.objects.filter(
            Q(user__filiale=request.user.filiale) | Q(user__isnull=True)
        ).order_by('-timestamp')
        
    template_path = 'quality_control_audits_pdf.html'
    logo_rel_path = "images/boa.png"
    logo_full_path = os.path.join(settings.MEDIA_ROOT, logo_rel_path)
    context = {
        'audits': audits,
        'logo_path': logo_full_path if os.path.exists(logo_full_path) else None
    }
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="audit_controles.pdf"'
    template = get_template(template_path)
    html = template.render(context)
    if not pisa:
        return HttpResponse("L'exportation PDF n'est pas disponible sur ce serveur (dépendances manquantes).")
    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=_pdf_link_callback)
    if pisa_status.err: return HttpResponse('Erreur PDF')
    return response

@login_required
def export_rule_failures(request, rule_id):
    rule = get_object_or_404(DataQualityRule, pk=rule_id)
    user_organe = (getattr(request.user, 'organe', '') or '').strip()
    user_filiale = getattr(request.user, 'filiale', '')
    
                                                                         
                                                                            
    group_organs = ['PASS', 'Conformité Groupe', 'Contrôle Permanent Groupe']
    if request.user.is_superuser or user_organe in group_organs:
        eval_filiale = None
    else:
        eval_filiale = (user_filiale or '').strip()
        if not eval_filiale or eval_filiale == 'GROUPE':
            messages.error(request, "Accès refusé : aucune filiale n'est associée à votre compte.")
            return redirect("accueil")

                                                                 
    model = Kyc_pp if rule.applicability == 'PP' else Kyc_pm
    queryset = model.objects.all()
    if eval_filiale:
        queryset = queryset.filter(FILIALE=eval_filiale)
        
                                                                                                
                                                                                                          
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anomalies"
    
                        
    headers = ["CLIENT", "EXPL", "FILIALE", "AGENCE"]
    if rule.control_type == 'simple':
        headers.append(rule.field_name.upper())
    else:
                                                   
        cond_fields = list(rule.conditions.values_list('field_name', flat=True))
                                                        
        seen = set()
        unique_fields = [f for f in cond_fields if not (f in seen or seen.add(f))]
        for f in unique_fields:
            headers.append(f.upper())
            
    ws.append(headers)
    
    if rule.control_type == 'simple':
        param = (rule.parameter or '').strip().lower()
        if not param or param == 'existence':
            failures = queryset.filter(Q(**{f"{rule.field_name}__isnull": True}) | Q(**{f"{rule.field_name}": ""}))
            for row in failures.values("CLIENT", "EXPL", "FILIALE", "AGENCE", rule.field_name).iterator():
                ws.append([row['CLIENT'], row['EXPL'], row['FILIALE'], row['AGENCE'], str(row.get(rule.field_name) or '')])
        
        elif param.isdigit() or (param.startswith('len') or param.startswith('long')):
            import re
            match = re.search(r'\d+', param)
            target_len = int(match.group()) if match else int(param)
            for row in queryset.values("CLIENT", "EXPL", "FILIALE", "AGENCE", rule.field_name).iterator():
                val = str(row.get(rule.field_name) or '')
                if len(val) != target_len:
                    ws.append([row['CLIENT'], row['EXPL'], row['FILIALE'], row['AGENCE'], val])
        else:
            failures = queryset.exclude(**{f"{rule.field_name}": rule.parameter})
            for row in failures.values("CLIENT", "EXPL", "FILIALE", "AGENCE", rule.field_name).iterator():
                ws.append([row['CLIENT'], row['EXPL'], row['FILIALE'], row['AGENCE'], str(row.get(rule.field_name) or '')])
                
    elif rule.control_type == 'composite':
        conditions = rule.conditions.all()
        fields_to_fetch = ["CLIENT", "EXPL", "FILIALE", "AGENCE"]
        cond_fields = [c.field_name for c in conditions]
        unique_cond_fields = list(dict.fromkeys(cond_fields))
        fields_to_fetch.extend(unique_cond_fields)
        
        today = datetime.today().date()

        for row in queryset.values(*set(fields_to_fetch)).iterator():
            if _dq_conditions_flag(conditions, lambda f: row.get(f, ''), today, _dq_safe_parse_date, _dq_calculate_age):
                line = [row['CLIENT'], row['EXPL'], row['FILIALE'], row['AGENCE']]
                for f in unique_cond_fields:
                    line.append(str(row.get(f) or ''))
                ws.append(line)

    from django.utils.text import slugify
    safe_name = slugify(rule.name).replace('-', '_') or f"anomalies_{rule.id}"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="{safe_name}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_rules_pdf(request):
                                                                       
    user_organe = (getattr(request.user, 'organe', '') or '').strip()
    user_filiale = getattr(request.user, 'filiale', '')
    group_organs = ['PASS', 'Conformité Groupe']
    
    if user_organe == 'PASS':
        rules_qs = DataQualityRule.objects.all().order_by('-created_at')
    elif user_organe in ['Conformité Groupe']:
        rules_qs = DataQualityRule.objects.all().order_by('-created_at')
    else:
                             
        rules_qs = DataQualityRule.objects.filter(
            Q(created_by__filiale=user_filiale) | Q(created_by__isnull=True)
        ).order_by('-created_at')

                                             
    import hashlib
    from django.core.cache import cache
    
    rules_with_stats = []
    eval_filiale = None if user_organe in group_organs else user_filiale
    rules_version = cache.get('quality_control_rules_version', 1)
    cache_ttl = 86400
    data_refresh_bucket = timezone.localdate().isoformat()

    for rule in rules_qs:
                                                                           
        rule_signature = f"{rule.id}|{rule.name}|{rule.applicability}|{rule.field_name}|{rule.control_type}|{rule.parameter}|{rule.active}|{eval_filiale}"
        rule_hash = hashlib.md5(rule_signature.encode('utf-8')).hexdigest()
        rule_cache_key = f"quality_control:stat:v{rules_version}:d{data_refresh_bucket}:{rule_hash}"
        
        stat = cache.get(rule_cache_key)
        if stat is None:
            stat = evaluate_data_quality_rule(rule, filiale=eval_filiale)
            cache.set(rule_cache_key, stat, timeout=cache_ttl)
            
                        
        total = stat.get('total', 0)
        stat['compliance_rate'] = compliance_rate_floor(stat['ok_count'], total, stat.get('fail_count', 0)) if total else 0
        
        from kyc.forms import DataQualityRuleForm
        parsed_filiales = DataQualityRuleForm._parse_filiales(rule.filiale)
        filiales_display = ", ".join(parsed_filiales) if parsed_filiales else "Toutes les filiales"

        rules_with_stats.append({
            'rule': rule,
            'stat': stat,
            'rule_number': rule.rule_number,
            'filiales_display': filiales_display,
        })

    template_path = 'quality_rules_pdf.html'
    logo_path = os.path.join(settings.MEDIA_ROOT, "images", "boa.png")
    context = {
        'rules': rules_with_stats,
        'user': request.user,
        'date': timezone.now(),
        'filiale': user_filiale if user_organe not in group_organs else "GROUPE",
        'logo_path': logo_path if os.path.exists(logo_path) else None,
    }
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="regles_qualite_kyc.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)
    
    if not pisa:
        return HttpResponse("L'exportation PDF n'est pas disponible sur ce serveur (dépendances manquantes).")
        
    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=_pdf_link_callback)
    if pisa_status.err:
        return HttpResponse('Erreur lors de la génération du PDF')
        
    return response


@login_required
def accueil(request):
    user = request.user

    if user.is_authenticated:
        return redirect('profil')
    return redirect('login_kyc')


@login_required
def import_page(request):
    if not request.user.is_superuser:
        return redirect('accueil')
    log_dir = os.path.join(settings.BASE_DIR, "logs")
    os.makedirs(log_dir, exist_ok=True)
    history_path = os.path.join(log_dir, "import_history.log")
    run_dir = os.path.join(log_dir, "import_runs")
    os.makedirs(run_dir, exist_ok=True)

    def read_history(limit=50):
        if not os.path.exists(history_path):
            return []
        try:
            with open(history_path, "r", encoding="utf-8") as f:
                lines = f.readlines()
            return [ln.rstrip("\n") for ln in lines[-limit:]]
        except Exception:
            return []

    def list_run_logs(limit=20):
        try:
            files = [f for f in os.listdir(run_dir) if f.endswith(".log")]
            files.sort(reverse=True)
            return files[:limit]
        except Exception:
            return []

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        data_dir = (request.POST.get("data_dir") or "").strip()
        filiales = (request.POST.get("filiales") or "").strip()
        only = (request.POST.get("only") or "").strip()
        bulk_size = (request.POST.get("bulk_size") or "").strip()
        taux_clear = request.POST.get("taux_clear") == "on"

        script = None
        if action == "run_kyc":
            script = "import_kyc.py"
        elif action == "run_premier":
            script = "import_premier.py"

        if not script:
            messages.error(request, "Action d'import inconnue.")
        else:
            env = os.environ.copy()
            if data_dir:
                env["KYC_DATA_DIR"] = data_dir
            if filiales:
                env["KYC_FILIALES"] = filiales
            if bulk_size:
                env["KYC_BULK_SIZE"] = bulk_size
            if only:
                env["KYC_ONLY"] = only
            if taux_clear:
                env["KYC_TAUX_CLEAR"] = "1"
            elif "KYC_TAUX_CLEAR" in env:
                env.pop("KYC_TAUX_CLEAR", None)

            cmd = [sys.executable, str(settings.BASE_DIR / script)]
            start_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
            detail_path = os.path.join(run_dir, f"{action}_{run_id}.log")
            try:
                result = subprocess.run(
                    cmd,
                    cwd=str(settings.BASE_DIR),
                    env=env,
                    capture_output=True,
                    text=True,
                    check=False,
                )
                status = "SUCCESS" if result.returncode == 0 else "FAILED"
                with open(detail_path, "w", encoding="utf-8") as df:
                    df.write(f"CMD: {cmd}\n")
                    df.write(f"START: {start_ts}\n")
                    df.write(f"RETURN: {result.returncode}\n")
                    df.write(f"DATA_DIR: {data_dir}\n")
                    df.write(f"FILIALES: {filiales}\n")
                    df.write(f"ONLY: {only}\n")
                    df.write(f"BULK_SIZE: {bulk_size}\n")
                    df.write(f"TAUX_CLEAR: {taux_clear}\n")
                    df.write("\n--- STDOUT ---\n")
                    df.write(result.stdout or "")
                    df.write("\n--- STDERR ---\n")
                    df.write(result.stderr or "")

                with open(history_path, "a", encoding="utf-8") as hf:
                    hf.write(f"{start_ts} | {action} | {status} | log={detail_path}\n")

                log_audit(
                    request,
                    category=AuditEvent.CAT_IMPORT,
                    action=f"Execution import ({action})",
                    target=os.path.basename(detail_path),
                    details=(f"Repertoire : {data_dir or '-'} | Filiales : {filiales or 'toutes'} | "
                             f"Perimetre : {only or '-'} | Code retour : {result.returncode}"),
                    success=(status == "SUCCESS"),
                )

                if status == "SUCCESS":
                                                                                       
                    current_v = cache.get('quality_control_rules_version', 1)
                    cache.set('quality_control_rules_version', current_v + 1, timeout=None)
                    messages.success(request, "Import terminé avec succès.")
                else:
                    messages.error(request, f"Import échoué (code {result.returncode}).")

            except Exception as e:
                messages.error(request, f"Erreur d'exécution: {e}")

    context = {
        "history": read_history(),
        "run_logs": list_run_logs(),
        "history_log_name": "import_history.log",
    }
    return render(request, 'import.html', context)


DOCUMENT_EXTRACTION_FIELD_LABELS = [
    ("prenom", "Prenom"),
    ("nom", "Nom"),
    ("date_naissance", "Date de naissance"),
    ("lieu_naissance", "Lieu de naissance"),
    ("sexe", "Sexe"),
    ("pays_naissance", "Pays de naissance"),
    ("pays_delivrance", "Pays de delivrance"),
    ("date_expiration", "Date d'expiration"),
    ("adresse", "Adresse"),
    ("origine_revenu", "Origine du revenu"),
    ("numero_identification_nationale", "Numero identification nationale"),
    ("numero_document", "Numero document"),
    ("nationalite", "Nationalite"),
]

DOCUMENT_EXTRACTION_SEARCH_FIELDS = [
    ("all", "Tous les champs"),
    ("import_batch", "Lot d'import"),
    ("original_filename", "Nom du fichier"),
    ("source_filename", "Fichier source"),
    *DOCUMENT_EXTRACTION_FIELD_LABELS,
    ("extracted_text", "Texte extrait"),
]

KYC_PP_DOCUMENT_FIELD_MAP = [
    ("NUMID", "numero_identification_nationale", "NUMID"),
    ("NUMID", "numero_document", "NUMID"),
    ("DATNAIS", "date_naissance", "DATNAIS"),
    ("PAYNAIS", "pays_naissance", "PAYNAIS"),
    ("DATVALID", "date_expiration", "DATVALID"),
    ("ADRESSE", "adresse", "ADRESSE"),
    ("ORIGINE_REV", "origine_revenu", "ORIGINE_REV"),
]

KYC_PM_DOCUMENT_FIELD_MAP = [
    ("RCSNO", "numero_document", "RCSNO"),
    ("NUMERO_FISCAL", "numero_identification_nationale", "NUMERO_FISCAL"),
    ("ADRESSE_SOCIALE", "adresse", "ADRESSE_SOCIALE"),
    ("INTITULE_COMPTE", "nom", "INTITULE_COMPTE"),
]


def _normalize_match_value(value):
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


_field_visibility_configs_cache = None
_field_visibility_configs_cache_time = 0

def _get_cached_field_visibility_configs():
    global _field_visibility_configs_cache, _field_visibility_configs_cache_time
    import time
    now = time.time()
    if _field_visibility_configs_cache is None or now - _field_visibility_configs_cache_time > 10:
        _field_visibility_configs_cache = list(KycFieldVisibilityConfig.objects.all())
        _field_visibility_configs_cache_time = now
    return _field_visibility_configs_cache


def _document_source_allows(allowed_source, document_type):
    """Vrai si le type de document peut alimenter le champ.
    allowed_source : vide/None = tous documents ; str (un code, ou plusieurs
    separes par des virgules — heritage) ; ou liste de codes (multi-sources)."""
    if not allowed_source:
        return True
    if isinstance(allowed_source, str):
        allowed = [s.strip() for s in allowed_source.split(",") if s.strip()]
    else:
        allowed = [str(s).strip() for s in allowed_source if str(s).strip()]
    return not allowed or document_type in allowed


def _get_field_sources(filiale, client_type_val):
    configs = _get_cached_field_visibility_configs()
    spec_config = next((c for c in configs if c.client_type == client_type_val and filiale in (c.filiales or [])), None)
    if spec_config:
        return spec_config.field_sources or {}
    global_config = next((c for c in configs if c.client_type == client_type_val and not c.filiales), None)
    if global_config:
        return global_config.field_sources or {}
    return {}



COUNTRY_ALIASES = {
    "SENEGAL": {"SENEGAL", "SEN", "SN", "BOASN"},
    "BENIN": {"BENIN", "BEN", "BJ", "BOABJ"},
    "COTE D IVOIRE": {"COTEDIVOIRE", "CIV", "CI", "IVOIRE", "BOACI"},
    "BURKINA FASO": {"BURKINAFASO", "BFA", "BF", "BOABF"},
    "MALI": {"MALI", "MLI", "ML", "BOAML"},
    "TOGO": {"TOGO", "TGO", "TG", "BOATG"},
    "NIGER": {"NIGER", "NER", "NE", "BOANE"},
}


def _country_key(value):
    normalized = _normalize_match_value(value)
    if not normalized:
        return ""
    for country, aliases in COUNTRY_ALIASES.items():
        if normalized in aliases or any(alias and alias in normalized for alias in aliases):
            return country
    return normalized


def _countries_are_compatible(left, right):
    left_key = _country_key(left)
    right_key = _country_key(right)
    return not left_key or not right_key or left_key == right_key


def _document_country_guard_passes(document, client):
    if document.pays_naissance and getattr(client, "PAYNAIS", "") and not _countries_are_compatible(document.pays_naissance, client.PAYNAIS):
        return False
    return True


def _is_empty_kyc_value(value):
    normalized = str(value or "").strip().lower()
    return normalized in {"", "-", "na", "n/a", "none", "null", "nan"}


def _document_identity_keys(document):
    return {
        key for key in [
            _normalize_match_value(document.numero_document),
            _normalize_match_value(document.numero_identification_nationale),
        ] if key
    }


def _values_match(left, right):
    left_value = _normalize_match_value(left)
    right_value = _normalize_match_value(right)
    return bool(left_value and right_value and left_value == right_value)


def _date_values_match(left, right):
    left_text = str(left or "").strip()
    right_text = str(right or "").strip()
    if not left_text or not right_text:
        return False

    left_formatted = format_date_for_export(left_text, empty_value="")
    right_formatted = format_date_for_export(right_text, empty_value="")
    if left_formatted and right_formatted and _normalize_match_value(left_formatted) == _normalize_match_value(right_formatted):
        return True

    return _normalize_match_value(left_text) == _normalize_match_value(right_text)


def _date_match_key(value):
    formatted = format_date_for_export(value, empty_value="")
    return _normalize_match_value(formatted or value)


def _nationality_match_key(value):
    return _country_key(value) or _normalize_match_value(value)


def _nationality_values_match(document_value, client_value):
    if not document_value or not client_value:
        return False
    if _countries_are_compatible(document_value, client_value):
        return True
    return _normalize_match_value(document_value) == _normalize_match_value(client_value)


DEFAULT_KYC_DOCUMENT_MATCH_WEIGHTS = {
                        
    "pp_fullname_weight": 35,
    "pp_birth_date_weight": 35,
    "pp_birth_place_weight": 15,
    "pp_birth_country_weight": 15,
                        
    "pm_fullname_weight": 35,
    "pm_fiscal_weight": 35,
    "pm_address_weight": 15,
    "pm_country_weight": 15,
    "combination_threshold": 65,
    "min_display_score": 30,
    "pp_fullname_field": "INTITULE_COMPTE",
    "pm_fullname_field": "INTITULE_COMPTE",
}


DOCUMENT_MATCH_GROUP_ORGANES = [
    "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST",
]


def _user_filiale_scope(user):
    """Filiale a laquelle restreindre les candidats KYC : vide (= toutes) pour un
    utilisateur groupe, sinon la filiale de l'utilisateur."""
    if not user or not getattr(user, "is_authenticated", False):
        return ""
    is_group_user = (
        getattr(user, "filiale", "") in ["BOA Group", "BOA GROUP"]
        or getattr(user, "organe", "") in DOCUMENT_MATCH_GROUP_ORGANES
        or not getattr(user, "filiale", "")
    )
    return "" if is_group_user else (user.filiale or "").strip()


def _filter_matches_for_user_scope(matches, user):
    """Restreint les correspondances au perimetre de l'utilisateur :
    - groupe (BOA Group / organes groupe) : tout ;
    - Charge Client : uniquement son portefeuille (FILIALE + EXPL = code_expl) ;
    - Directeur Agence : uniquement son agence (FILIALE + AGENCE) ;
    - autres organes filiale : toute leur filiale."""
    if not matches:
        return matches
    filiale_scope = _user_filiale_scope(user)
    if not filiale_scope:
        return matches
    organe = (getattr(user, "organe", "") or "").strip()
    agence = (getattr(user, "agence", "") or "").strip()
    code_expl = (getattr(user, "code_expl", "") or "").strip()
    filiale_key = _normalize_match_value(filiale_scope)
    scoped = []
    for match in matches:
        client = match.get("client")
        if _normalize_match_value(getattr(client, "FILIALE", "") or "") != filiale_key:
            continue
        if organe == "Chargé Client" and code_expl:
            if _normalize_match_value(getattr(client, "EXPL", "") or "") != _normalize_match_value(code_expl):
                continue
        elif organe == "Directeur Agence" and agence:
            if _normalize_match_value(getattr(client, "AGENCE", "") or "") != _normalize_match_value(agence):
                continue
        scoped.append(match)
    return scoped


def _name_tokens(value):
    """Jeu de jetons normalises d'un nom (par mot), pour comparer un nom/prenom
    du document au champ combine (ex. CLIENT = 'JEAN DUPONT') d'un client KYC."""
    words = re.split(r"[^A-Za-zÀ-ÿ0-9]+", str(value or "").upper())
    return {_normalize_match_value(word) for word in words if len(word) >= 2 and _normalize_match_value(word)}


def _name_value_matches(document_value, client_value):
    """Vrai si tous les jetons du nom/prenom extrait sont presents dans la valeur KYC."""
    document_tokens = _name_tokens(document_value)
    client_tokens = _name_tokens(client_value)
    if not document_tokens or not client_tokens:
        return False
    return document_tokens.issubset(client_tokens)


def _get_kyc_document_match_weights():
    try:
        settings_obj = KycDocumentMatchSettings.get_active()
        return {
            "pp_fullname_weight": settings_obj.pp_fullname_weight,
            "pp_birth_date_weight": settings_obj.pp_birth_date_weight,
            "pp_birth_place_weight": settings_obj.pp_birth_place_weight,
            "pp_birth_country_weight": settings_obj.pp_birth_country_weight,
            "pm_fullname_weight": settings_obj.pm_fullname_weight,
            "pm_fiscal_weight": settings_obj.pm_fiscal_weight,
            "pm_address_weight": settings_obj.pm_address_weight,
            "pm_country_weight": settings_obj.pm_country_weight,
            "combination_threshold": settings_obj.combination_threshold,
            "min_display_score": settings_obj.min_display_score,
            "pp_fullname_field": settings_obj.pp_fullname_field or "INTITULE_COMPTE",
            "pm_fullname_field": settings_obj.pm_fullname_field or "INTITULE_COMPTE",
        }
    except Exception:
        return DEFAULT_KYC_DOCUMENT_MATCH_WEIGHTS.copy()


def _document_client_identity_score(document, client, weights=None):
    """Taux de correspondance document <-> client, calcul distinct PP / PM.
    PP : NUMID (n° id national) = 100 % ; sinon nom&prenom + naissance + lieu + pays.
    PM : RCSNO (registre commerce) = 100 % ; sinon raison sociale + fiscal + adresse + pays."""
    weights = weights or DEFAULT_KYC_DOCUMENT_MATCH_WEIGHTS
    is_pm = isinstance(client, Kyc_pm)
    doc_fullname = f"{document.nom or ''} {document.prenom or ''}"

    if is_pm:
                                                                       
        client_rcsno = getattr(client, "RCSNO", "")
        if _values_match(document.numero_document, client_rcsno) or _values_match(document.numero_identification_nationale, client_rcsno):
            return 100
        fullname_field = weights.get("pm_fullname_field", "INTITULE_COMPTE")
        doc_country = document.nationalite or document.pays_naissance or document.pays_delivrance
        checks = [
            (_name_value_matches(doc_fullname, getattr(client, fullname_field, "")), weights.get("pm_fullname_weight", 0)),
            (_values_match(document.numero_identification_nationale, getattr(client, "NUMERO_FISCAL", "")), weights.get("pm_fiscal_weight", 0)),
            (_name_value_matches(document.adresse, getattr(client, "ADRESSE_SOCIALE", "")), weights.get("pm_address_weight", 0)),
            (_nationality_values_match(doc_country, getattr(client, "PAYS_JUR", "")), weights.get("pm_country_weight", 0)),
        ]
    else:
                                                                                    
        client_numid = getattr(client, "NUMID", "")
        if _values_match(document.numero_identification_nationale, client_numid) or _values_match(document.numero_document, client_numid):
            return 100
        fullname_field = weights.get("pp_fullname_field", "INTITULE_COMPTE")
        checks = [
            (_name_value_matches(doc_fullname, getattr(client, fullname_field, "")), weights.get("pp_fullname_weight", 0)),
            (_date_values_match(document.date_naissance, getattr(client, "DATNAIS", "")), weights.get("pp_birth_date_weight", 0)),
            (_nationality_values_match(document.lieu_naissance, getattr(client, "PAYNAIS", "")), weights.get("pp_birth_place_weight", 0)),
            (
                _nationality_values_match(document.nationalite, getattr(client, "PAYNAIS", ""))
                or _nationality_values_match(document.pays_naissance, getattr(client, "PAYNAIS", "")),
                weights.get("pp_birth_country_weight", 0),
            ),
        ]

    score = 0
    for matched, weight in checks:
        if matched:
            score += weight
    return score


def _document_client_haystack(document):
    return _normalize_match_value(
        " ".join([
            document.original_filename or "",
            document.source_filename or "",
            document.import_batch or "",
        ])
    )


def _document_client_tokens(document):
    raw_value = " ".join([
        document.original_filename or "",
        document.source_filename or "",
        document.import_batch or "",
    ]).upper()
    return {_normalize_match_value(token) for token in re.split(r"[^A-Z0-9]+", raw_value) if len(token) >= 4}


def _document_unique_key(document):
    identity_keys = sorted(_document_identity_keys(document))
    if identity_keys:
        country_parts = [
            _country_key(document.pays_delivrance),
            _country_key(document.pays_naissance),
        ]
        country_scope = "|".join([part for part in country_parts if part])
        if country_scope:
            return "identity:" + country_scope + ":" + "|".join(identity_keys)
        return "identity:" + "|".join(identity_keys)
    return "file:" + _normalize_match_value(
        "|".join([
            document.original_filename or "",
            document.source_filename or "",
            document.import_batch or "",
            document.page_range or "",
        ])
    )


def _client_dedup_key(client):
    normalized_idp = _normalize_match_value(getattr(client, "IDP", ""))
    if normalized_idp:
        return f"idp:{normalized_idp}"
    normalized_client = _normalize_match_value(getattr(client, "CLIENT", ""))
    if normalized_client:
        return f"client:{normalized_client}"
    return f"pk:{client.pk}"


def _build_kyc_pp_document_matches(document_queryset, limit=3000, result_limit=200, progress_callback=None, filiale_scope=None):
                                                                             
    documents_all = document_queryset.filter(extraction_status="done").order_by("-created_at")
    documents_total = documents_all.count()
    documents_for_match = list(documents_all[:limit])
    documents_truncated = documents_total > limit
    if not documents_for_match:
        return [], {"documents_checked": 0, "documents_matched": 0, "clients_matched": 0, "suggestions_count": 0,
                    "match_rate": 0, "documents_total": documents_total, "documents_truncated": False,
                    "results_truncated": False, "result_limit": result_limit}
    if progress_callback:
        progress_callback(0, len(documents_for_match), "Preparation du rapprochement")
    match_weights = _get_kyc_document_match_weights()
    min_display_score = match_weights.get("min_display_score", 30)
    fullname_field = match_weights.get("pp_fullname_field", "INTITULE_COMPTE")

    document_keys = set()
    for document in documents_for_match:
        document_keys.update(_document_identity_keys(document))

                                                                                    
    only_fields = ["id", "FILIALE", "AGENCE", "CLIENT", "IDP", "NUMID", "DATNAIS", "PAYNAIS",
                   "DATVALID", "ADRESSE", "ORIGINE_REV", "INTITULE_COMPTE", "EMPLOYEUR"]
    base_qs = Kyc_pp.objects.all()
    if filiale_scope:
        base_qs = base_qs.filter(FILIALE=filiale_scope)

    kyc_candidates = {}
    if document_keys:
        for client in base_qs.exclude(NUMID="").only(*only_fields):
            normalized_numid = _normalize_match_value(client.NUMID)
            if normalized_numid in document_keys:
                kyc_candidates.setdefault(normalized_numid, []).append(client)

    client_by_code = {}
    clients_by_birth_date = {}
    clients_by_validity_date = {}
    clients_by_nationality = {}
    clients_by_birth_place = {}
    clients_by_name_token = {}
    for client in base_qs.only(*only_fields)[:50000]:
        normalized_client = _normalize_match_value(client.CLIENT)
        if normalized_client:
            client_by_code.setdefault(normalized_client, []).append(client)
        birth_key = _date_match_key(client.DATNAIS)
        if birth_key:
            clients_by_birth_date.setdefault(birth_key, []).append(client)
        validity_key = _date_match_key(client.DATVALID)
        if validity_key:
            clients_by_validity_date.setdefault(validity_key, []).append(client)
        nationality_key = _nationality_match_key(client.PAYNAIS)
        if nationality_key:
            clients_by_nationality.setdefault(nationality_key, []).append(client)
            clients_by_birth_place.setdefault(nationality_key, []).append(client)
                                                                                
        for token in _name_tokens(getattr(client, fullname_field, "")):
            clients_by_name_token.setdefault(token, []).append(client)

    matches = []
    client_match_index = {}
    matched_client_ids = set()
    matched_document_ids = set()
    for index, document in enumerate(documents_for_match, start=1):
        if progress_callback:
            progress_callback(index, len(documents_for_match), f"Analyse document {index}/{len(documents_for_match)}")
        candidate_clients = []
        for identity_key in _document_identity_keys(document):
            candidate_clients.extend(kyc_candidates.get(identity_key, []))

        for client_token in _document_client_tokens(document):
            candidate_clients.extend(client_by_code.get(client_token, []))

        birth_key = _date_match_key(document.date_naissance)
        validity_key = _date_match_key(document.date_expiration)
        nationality_key = _nationality_match_key(document.nationalite or document.pays_naissance)
        birth_place_key = _nationality_match_key(document.lieu_naissance)
        combination_pool = {}
        for client in clients_by_birth_date.get(birth_key, []):
            combination_pool[client.pk] = client
        for client in clients_by_validity_date.get(validity_key, []):
            combination_pool[client.pk] = client
        if birth_key or validity_key:
            for client in clients_by_birth_place.get(birth_place_key, []):
                combination_pool[client.pk] = client
        if birth_key or validity_key:
            for client in clients_by_nationality.get(nationality_key, []):
                combination_pool[client.pk] = client
                                                                                     
                                                                                          
        doc_name_tokens = _name_tokens(document.nom) | _name_tokens(document.prenom)
        if doc_name_tokens and match_weights.get("pp_fullname_weight"):
            token_hits = {}
            token_objs = {}
            for token in doc_name_tokens:
                for client in clients_by_name_token.get(token, []):
                    token_hits[client.pk] = token_hits.get(client.pk, 0) + 1
                    token_objs[client.pk] = client
            for pk, hits in token_hits.items():
                if hits >= len(doc_name_tokens):
                    candidate_clients.append(token_objs[pk])
                else:
                    combination_pool.setdefault(pk, token_objs[pk])
        for client in combination_pool.values():
            if _document_client_identity_score(document, client, match_weights) >= match_weights["combination_threshold"]:
                candidate_clients.append(client)

        unique_clients = {}
        for client in candidate_clients:
            if not _document_country_guard_passes(document, client):
                continue
            unique_clients[client.pk] = client

        for client in unique_clients.values():
            suggestions = []
            used_kyc_fields = set()
            client_filiale = getattr(client, "FILIALE", "").strip()
            field_sources = _get_field_sources(client_filiale, "pp")
            for kyc_field, document_field, label in KYC_PP_DOCUMENT_FIELD_MAP:
                allowed_source = field_sources.get(kyc_field)
                if not _document_source_allows(allowed_source, document.document_type):
                    continue
                document_value = getattr(document, document_field, "")
                if not document_value or not _is_empty_kyc_value(getattr(client, kyc_field, "")):
                    continue
                if kyc_field in used_kyc_fields:
                    continue
                used_kyc_fields.add(kyc_field)
                suggestions.append({
                    "field": kyc_field,
                    "label": label,
                    "document_value": document_value,
                })

            match_rate = _document_client_identity_score(document, client, match_weights)
            if match_rate < min_display_score:
                continue

            candidate_match = {
                "client": client,
                "document": document,
                "suggestions": suggestions,
                "match_rate": match_rate,
            }
            if not _build_kyc_pp_match_action_items(candidate_match):
                continue

            client_dedup_key = _client_dedup_key(client)
            if client_dedup_key in client_match_index:
                existing_match = matches[client_match_index[client_dedup_key]]
                if _document_unique_key(existing_match["document"]) != _document_unique_key(document):
                    continue
                existing_fields = {suggestion["field"] for suggestion in existing_match["suggestions"]}
                for suggestion in suggestions:
                    if suggestion["field"] not in existing_fields:
                        existing_match["suggestions"].append(suggestion)
                        existing_fields.add(suggestion["field"])
                existing_extra_actions = existing_match.setdefault("extra_action_items", [])
                existing_action_keys = {
                    (action.get("kind"), (action.get("field") or "").strip().upper())
                    for action in _build_kyc_pp_match_action_items(existing_match)
                }
                for action in _build_kyc_pp_match_action_items(candidate_match):
                    action_key = (action.get("kind"), (action.get("field") or "").strip().upper())
                    if action_key not in existing_action_keys:
                        existing_extra_actions.append(action)
                        existing_action_keys.add(action_key)
                existing_match["match_rate"] = max(existing_match["match_rate"], match_rate)
                continue

            matched_client_ids.add(client.pk)
            matched_document_ids.add(document.pk)
            client_match_index[client_dedup_key] = len(matches)
            matches.append(candidate_match)

    suggestions_count = sum(len(match["suggestions"]) for match in matches)
    match_rate = round((len(matched_document_ids) / len(documents_for_match)) * 100, 1)

    results_truncated = bool(result_limit and len(matches) > result_limit)
    summary = {
        "documents_checked": len(documents_for_match),
        "documents_matched": len(matched_document_ids),
        "clients_matched": len(client_match_index),
        "suggestions_count": suggestions_count,
        "match_rate": match_rate,
        "documents_total": documents_total,
        "documents_truncated": documents_truncated,
        "results_truncated": results_truncated,
        "result_limit": result_limit,
    }
    if result_limit:
        return matches[:result_limit], summary
    return matches, summary


def _build_kyc_pm_document_matches(document_queryset, limit=3000, result_limit=200, progress_callback=None, filiale_scope=None):
                                                                             
    documents_all = document_queryset.filter(extraction_status="done").order_by("-created_at")
    documents_total = documents_all.count()
    documents_for_match = list(documents_all[:limit])
    documents_truncated = documents_total > limit
    if not documents_for_match:
        return [], {"documents_checked": 0, "documents_matched": 0, "clients_matched": 0, "suggestions_count": 0,
                    "match_rate": 0, "documents_total": documents_total, "documents_truncated": False,
                    "results_truncated": False, "result_limit": result_limit}
    if progress_callback:
        progress_callback(0, len(documents_for_match), "Preparation du rapprochement")
    match_weights = _get_kyc_document_match_weights()
    min_display_score = match_weights.get("min_display_score", 30)
    pm_fullname_field = match_weights.get("pm_fullname_field", "INTITULE_COMPTE")

    document_keys = set()
    for document in documents_for_match:
        if document.numero_document:
            document_keys.add(_normalize_match_value(document.numero_document))
        if document.numero_identification_nationale:
            document_keys.add(_normalize_match_value(document.numero_identification_nationale))

    pm_only_fields = ["id", "FILIALE", "AGENCE", "CLIENT", "IDM", "RCSNO", "NUMERO_FISCAL", "ADRESSE_SOCIALE", "INTITULE_COMPTE", "PAYS_JUR"]
    base_qs = Kyc_pm.objects.all()
    if filiale_scope:
        base_qs = base_qs.filter(FILIALE=filiale_scope)

    kyc_candidates = {}
    if document_keys:
        for client in base_qs.exclude(RCSNO="").only(*pm_only_fields):
            norm_rcs = _normalize_match_value(client.RCSNO)
            if norm_rcs in document_keys:
                kyc_candidates.setdefault(norm_rcs, []).append(client)
        for client in base_qs.exclude(NUMERO_FISCAL="").only(*pm_only_fields):
            norm_nif = _normalize_match_value(client.NUMERO_FISCAL)
            if norm_nif in document_keys:
                kyc_candidates.setdefault(norm_nif, []).append(client)

    client_by_name = {}
    clients_by_name_token = {}
    for client in base_qs.only(*pm_only_fields)[:50000]:
        norm_name = _normalize_match_value(client.CLIENT)
        if norm_name:
            client_by_name.setdefault(norm_name, []).append(client)
        norm_intitule = _normalize_match_value(client.INTITULE_COMPTE)
        if norm_intitule and norm_intitule != norm_name:
            client_by_name.setdefault(norm_intitule, []).append(client)
                                                                                  
        for token in _name_tokens(getattr(client, pm_fullname_field, "")):
            clients_by_name_token.setdefault(token, []).append(client)

    matches = []
    matched_client_ids = set()
    matched_document_ids = set()
    
    for index, document in enumerate(documents_for_match, start=1):
        if progress_callback:
            progress_callback(index, len(documents_for_match), f"Analyse document {index}/{len(documents_for_match)}")
        
        candidate_clients = []
        if document.numero_document:
            candidate_clients.extend(kyc_candidates.get(_normalize_match_value(document.numero_document), []))
        if document.numero_identification_nationale:
            candidate_clients.extend(kyc_candidates.get(_normalize_match_value(document.numero_identification_nationale), []))
            
        for client_token in _document_client_tokens(document):
            candidate_clients.extend(client_by_name.get(client_token, []))

                                                                                     
        doc_name_tokens = _name_tokens(document.nom) | _name_tokens(document.prenom)
        if doc_name_tokens and match_weights.get("pm_fullname_weight"):
            token_hits = {}
            token_objs = {}
            for token in doc_name_tokens:
                for client in clients_by_name_token.get(token, []):
                    token_hits[client.pk] = token_hits.get(client.pk, 0) + 1
                    token_objs[client.pk] = client
            for pk, hits in token_hits.items():
                if hits >= len(doc_name_tokens):
                    candidate_clients.append(token_objs[pk])

        unique_clients = {}
        for client in candidate_clients:
            unique_clients[client.pk] = client
            
        for client in unique_clients.values():
            suggestions = []
            used_kyc_fields = set()
            client_filiale = getattr(client, "FILIALE", "").strip()
            field_sources = _get_field_sources(client_filiale, "pm")
            for kyc_field, document_field, label in KYC_PM_DOCUMENT_FIELD_MAP:
                allowed_source = field_sources.get(kyc_field)
                if not _document_source_allows(allowed_source, document.document_type):
                    continue
                document_value = getattr(document, document_field, "")
                if not document_value or not _is_empty_kyc_value(getattr(client, kyc_field, "")):
                    continue
                if kyc_field in used_kyc_fields:
                    continue
                used_kyc_fields.add(kyc_field)
                suggestions.append({
                    "field": kyc_field,
                    "label": label,
                    "value": str(document_value),
                })
                
                                                                                     
            match_rate = _document_client_identity_score(document, client, match_weights)
            if match_rate < min_display_score:
                continue

            extra_action_items = []
            for kyc_field, document_field, label in KYC_PM_DOCUMENT_FIELD_MAP:
                allowed_source = field_sources.get(kyc_field)
                if not _document_source_allows(allowed_source, document.document_type):
                    continue
                document_value = getattr(document, document_field, "")
                if document_value and _is_empty_kyc_value(getattr(client, kyc_field, "")):
                    extra_action_items.append({
                        "kind": "suggest",
                        "field": kyc_field,
                        "label": label,
                        "value": str(document_value),
                    })
                    
            matches.append({
                "client": client,
                "document": document,
                "suggestions": suggestions,
                "extra_action_items": extra_action_items,
                "match_rate": match_rate,
            })
            matched_client_ids.add(client.pk)
            matched_document_ids.add(document.pk)

    matches.sort(key=lambda m: m["match_rate"], reverse=True)
    results_truncated = bool(result_limit and len(matches) > result_limit)
    summary = {
        "documents_checked": len(documents_for_match),
        "documents_matched": len(matched_document_ids),
        "clients_matched": len(matched_client_ids),
        "suggestions_count": sum(len(m["suggestions"]) for m in matches),
        "match_rate": int(len(matched_document_ids) / len(documents_for_match) * 100) if documents_for_match else 0,
        "documents_total": documents_total,
        "documents_truncated": documents_truncated,
        "results_truncated": results_truncated,
        "result_limit": result_limit,
    }
    if result_limit:
        return matches[:result_limit], summary
    return matches, summary


LAST_KYC_PP_MATCH_SESSION_KEY = "document_extraction_last_kyc_pp_match_params"
LAST_KYC_PP_MATCH_RESULT_SESSION_KEY = "document_extraction_last_kyc_pp_match_result"
KYC_PP_MATCHED_BATCHES_SESSION_KEY = "document_extraction_kyc_pp_matched_batches"
LAST_UPLOADED_DOCUMENT_BATCH_SESSION_KEY = "document_extraction_last_uploaded_batch"
KYC_PP_MATCH_RESULT_VERSION = 3


def _filtered_document_extractions_from_params(params, user=None):
    documents = KycDocumentExtraction.objects.select_related("uploaded_by").all().order_by("-created_at", "-id")
    
    if user:
        users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
        is_group_user = (user.filiale in ["BOA Group", "BOA GROUP"]) or (user.organe in users_groupe) or (not user.filiale)
        if not is_group_user:
            documents = documents.filter(uploaded_by__filiale=user.filiale)

    client_type = params.get("client_type", "pp")
    documents = documents.filter(client_type=client_type)
    selected_document_type = (params.get("document_type", "") or "").strip()
    selected_import_batch = (params.get("import_batch") or "").strip()
    search_query = (params.get("q") or "").strip()
    search_field = params.get("field") or "all"
    allowed_search_fields = {field for field, _ in DOCUMENT_EXTRACTION_SEARCH_FIELDS}

                                                                              
                                                                               
    if selected_document_type:
        documents = documents.filter(document_type=selected_document_type)

    if selected_import_batch:
        documents = documents.filter(import_batch=selected_import_batch)

    if search_field not in allowed_search_fields:
        search_field = "all"

    if search_query:
        if search_field == "all":
            search_filter = Q()
            for field_name, _ in DOCUMENT_EXTRACTION_SEARCH_FIELDS:
                if field_name == "all":
                    continue
                search_filter |= Q(**{f"{field_name}__icontains": search_query})
            documents = documents.filter(search_filter)
        else:
            documents = documents.filter(**{f"{search_field}__icontains": search_query})

    return documents


def _filtered_document_extractions_from_request(request):
    return _filtered_document_extractions_from_params(request.GET, user=request.user)


def _clean_document_match_params(params):
    return {
        key: value
        for key, value in params.items()
        if key not in {"page", "extraction_id", "match_kyc", "match_job", "show_match_results", "result_modal", "child_modal", "tab"} and value not in (None, "")
    }


def _document_match_scope_params(params):
    return {
        key: value
        for key, value in _clean_document_match_params(params).items()
        if key not in KYC_PP_MATCH_FILTER_FIELDS
    }


def _serialize_kyc_pp_matches(matches, summary, params):
    return {
        "version": KYC_PP_MATCH_RESULT_VERSION,
        "params": params,
        "summary": summary,
        "matches": [
            {
                "client_id": match["client"].pk,
                "document_id": match["document"].pk,
                "suggestions": match["suggestions"],
                "extra_action_items": match.get("extra_action_items") or [],
                "match_rate": match["match_rate"],
            }
            for match in matches
        ],
    }


def _hydrate_kyc_pp_match_result(result):
    if not result:
        return [], None, {}
    if result.get("version") != KYC_PP_MATCH_RESULT_VERSION:
        return [], None, result.get("params") or {}

    serialized_matches = result.get("matches") or []
    client_ids = [match.get("client_id") for match in serialized_matches if match.get("client_id")]
    document_ids = [match.get("document_id") for match in serialized_matches if match.get("document_id")]
    
    params = result.get("params") or {}
    client_type = params.get("client_type", "pp")
    
    if client_type == "pm":
        clients = Kyc_pm.objects.in_bulk(client_ids)
    else:
        clients = Kyc_pp.objects.in_bulk(client_ids)
        
    documents = KycDocumentExtraction.objects.in_bulk(document_ids)

    matches = []
    for serialized_match in serialized_matches:
        client = clients.get(serialized_match.get("client_id"))
        document = documents.get(serialized_match.get("document_id"))
        if not client or not document:
            continue
        matches.append({
            "client": client,
            "document": document,
            "suggestions": serialized_match.get("suggestions") or [],
            "extra_action_items": serialized_match.get("extra_action_items") or [],
            "match_rate": serialized_match.get("match_rate") or 0,
        })

    return matches, result.get("summary"), params


DOCUMENT_MATCH_STATUS_FILTERS = ("active", "pending", "validated", "rejected", "all")


def _annotate_matches_with_decisions(matches, client_type, status_filter="active"):
    """Superpose le statut de decision (a valider / valide / rejete) sur chaque
    correspondance et applique le filtre. 'active' = tout sauf les rejetees."""
    if not matches:
        return matches
    doc_ids = [m["document"].pk for m in matches if m.get("document")]
    client_ids = [m["client"].pk for m in matches if m.get("client")]
    decisions = {}
    if doc_ids and client_ids:
        for dec in KycMatchDecision.objects.filter(
            client_type=client_type, document_id__in=doc_ids, client_id__in=client_ids
        ).select_related("decided_by"):
            decisions[(dec.document_id, dec.client_id)] = dec

    annotated = []
    for match in matches:
        document = match.get("document")
        client = match.get("client")
        if not document or not client:
            continue
        decision = decisions.get((document.pk, client.pk))
        status = decision.status if decision else "pending"
        match["decision_status"] = status
        match["decision"] = decision
        if status_filter == "all":
            annotated.append(match)
        elif status_filter in ("pending", "validated", "rejected"):
            if status == status_filter:
                annotated.append(match)
        else:                                     
            if status != "rejected":
                annotated.append(match)
    return annotated


def _merge_kyc_pp_match_lists(match_lists):
    merged = []
    index_by_key = {}

    for matches in match_lists:
        for match in matches:
            client = match.get("client")
            client_id = getattr(client, "IDP", None) or getattr(client, "IDM", None)
            normalized_id = _normalize_match_value(client_id or "")
            client_pk = getattr(client, "pk", None)
                                                                                    
                                                                                      
            key = (
                ("idp", normalized_id, _document_unique_key(match.get("document")))
                if normalized_id
                else ("client", client_pk, _document_unique_key(match.get("document")))
            )
            if key not in index_by_key:
                index_by_key[key] = len(merged)
                merged.append(match)
                continue

            existing_match = merged[index_by_key[key]]
            existing_fields = {
                (suggestion.get("field") or "").strip().upper()
                for suggestion in existing_match.get("suggestions", [])
            }
            for suggestion in match.get("suggestions", []):
                suggestion_field = (suggestion.get("field") or "").strip().upper()
                if suggestion_field and suggestion_field not in existing_fields:
                    existing_match.setdefault("suggestions", []).append(suggestion)
                    existing_fields.add(suggestion_field)

            existing_actions = existing_match.setdefault("extra_action_items", [])
            action_keys = {
                (action.get("kind"), (action.get("field") or "").strip().upper())
                for action in existing_actions
            }
            for action in match.get("extra_action_items", []):
                action_key = (action.get("kind"), (action.get("field") or "").strip().upper())
                if action_key not in action_keys:
                    existing_actions.append(action)
                    action_keys.add(action_key)
            existing_match["match_rate"] = max(existing_match.get("match_rate", 0), match.get("match_rate", 0))

    return merged


def _user_can_access_document_match_job(user, job):
    return bool(user.is_superuser or job.created_by_id == user.pk)


def _run_document_match_job(job_id):
    close_old_connections()
    try:
        job = KycDocumentMatchJob.objects.get(pk=job_id)
        scope_params = job.scope_params or {}
        job.status = "running"
        job.started_at = timezone.now()
        job.message = "Preparation du rapprochement"
        job.save(update_fields=["status", "started_at", "message", "updated_at"])

        last_saved_step = {"value": -1}

        def progress_callback(current, total, message):
                                                                                      
                                                                                       
                                                                               
            step = 1 if total <= 20 else 5
            if (last_saved_step["value"] >= 0
                    and current != total
                    and current - last_saved_step["value"] < step):
                return
            last_saved_step["value"] = current
            KycDocumentMatchJob.objects.filter(pk=job_id).update(
                progress_current=current,
                progress_total=total,
                message=message,
                updated_at=timezone.now(),
            )

        documents = _filtered_document_extractions_from_params(scope_params)
        client_type = scope_params.get("client_type", "pp")
        filiale_scope = (scope_params.get("filiale_scope") or "").strip()
        if client_type == "pm":
            matches, summary = _build_kyc_pm_document_matches(
                documents,
                progress_callback=progress_callback,
                filiale_scope=filiale_scope,
            )
        else:
            matches, summary = _build_kyc_pp_document_matches(
                documents,
                progress_callback=progress_callback,
                filiale_scope=filiale_scope,
            )
        result = _serialize_kyc_pp_matches(matches, summary, scope_params)
        KycDocumentMatchJob.objects.filter(pk=job_id).update(
            status="completed",
            progress_current=summary.get("documents_checked", 0),
            progress_total=summary.get("documents_checked", 0),
            message="Rapprochement termine",
            result=result,
            completed_at=timezone.now(),
            updated_at=timezone.now(),
        )
    except Exception as exc:
        KycDocumentMatchJob.objects.filter(pk=job_id).update(
            status="failed",
            message="Echec du rapprochement",
            error=str(exc),
            completed_at=timezone.now(),
            updated_at=timezone.now(),
        )
    finally:
        close_old_connections()


@login_required
def start_document_extraction_match_job(request):
    user = request.user
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
    is_group_user = (user.filiale in ["BOA Group", "BOA GROUP"]) or (user.organe in users_groupe) or (not user.filiale)
    legacy_can_insert = (user.organe == "DSI") or (user.organe == "PASS" and is_group_user)
    if not KycScreeningAccess.perms_for(user, legacy_can_insert=legacy_can_insert)["can_run_matching"]:
        messages.error(request, "Vous n'avez pas l'autorisation de lancer un rapprochement.")
        return redirect("document_extraction")
    scope_params = _document_match_scope_params(request.GET)
                                                                                         
    filiale_scope = _user_filiale_scope(request.user)
    if filiale_scope:
        scope_params["filiale_scope"] = filiale_scope
    existing_job = (
        KycDocumentMatchJob.objects
        .filter(created_by=request.user, scope_params=scope_params, status="running")
        .order_by("-created_at")
        .first()
    )
    job = existing_job or KycDocumentMatchJob.objects.create(
        created_by=request.user,
        scope_params=scope_params,
        message="Rapprochement en attente",
    )
    if not existing_job:
        threading.Thread(target=_run_document_match_job, args=(job.pk,), daemon=True).start()

    redirect_params = dict(scope_params)
    redirect_params["match_job"] = job.pk
    return redirect(f"{reverse('document_extraction')}?{urlencode(redirect_params)}#suivi")


@login_required
def document_extraction_match_job_status(request, job_id):
    job = get_object_or_404(KycDocumentMatchJob, pk=job_id)
    if not _user_can_access_document_match_job(request.user, job):
        return JsonResponse({"error": "Acces non autorise"}, status=403)

    total = job.progress_total or 0
    current = job.progress_current or 0
    percent = min(100, int(current / total * 100)) if total else (100 if job.status == "completed" else 0)
    redirect_params = dict(job.scope_params or {})
    redirect_params["match_job"] = job.pk
    result_params = dict(redirect_params)
    result_params["show_match_results"] = "1"
    result_params["result_modal"] = "1"

    return JsonResponse({
        "id": job.pk,
        "status": job.status,
        "message": job.message,
        "error": job.error,
        "current": current,
        "total": total,
        "percent": percent,
        "redirect_url": f"{reverse('document_extraction')}?{urlencode(redirect_params)}#suivi",
        "result_url": f"{reverse('document_extraction')}?{urlencode(result_params)}#suivi",
    })


@login_required
def document_extraction_ocr_status(request):
    """Statut OCR d'un lot : compteurs par etat + progression du job en cours."""
    import_batch = (request.GET.get("batch") or "").strip()
    if not import_batch:
        return JsonResponse({"error": "Parametre batch manquant"}, status=400)

    counts = {"pending": 0, "processing": 0, "done": 0, "failed": 0}
    for row in (
        KycDocumentExtraction.objects
        .filter(import_batch=import_batch)
        .values("extraction_status")
        .annotate(n=Count("id"))
    ):
        counts[row["extraction_status"]] = row["n"]
    total = sum(counts.values())

    job = (
        KycDocumentOcrJob.objects
        .filter(import_batch=import_batch)
        .order_by("-created_at")
        .first()
    )
    job_payload = None
    if job:
        job_payload = {
            "id": job.pk,
            "status": job.status,
            "mode": job.mode,
            "message": job.message,
            "error": job.error[:500],
            "percent": job.progress_percent,
            "current": job.progress_current,
            "total": job.progress_total,
        }

    finished = counts["pending"] == 0 and counts["processing"] == 0 and (
        job is None or job.status in ("completed", "failed")
    )
    percent = int(((counts["done"] + counts["failed"]) / total) * 100) if total else (
        job_payload["percent"] if job_payload else 0
    )

    return JsonResponse({
        "batch": import_batch,
        "counts": counts,
        "total": total,
        "percent": percent,
        "finished": finished,
        "job": job_payload,
    })


KYC_PP_MATCH_FILTER_FIELDS = {
    "match_client": "CLIENT",
    "match_idp": "IDP",
    "match_idm": "IDM",
    "match_filiale": "FILIALE",
    "match_agence": "AGENCE",
}


def _get_kyc_pp_match_filters(params):
    return {
        key: (params.get(key) or "").strip()
        for key in KYC_PP_MATCH_FILTER_FIELDS
    }


def _filter_kyc_pp_matches(matches, params):
    filters = _get_kyc_pp_match_filters(params)
    active_filters = {
        key: value.lower()
        for key, value in filters.items()
        if value
    }
    if not active_filters:
        return matches

    filtered_matches = []
    for match in matches:
        client = match["client"]
        keep_match = True
        for param_name, filter_value in active_filters.items():
            client_field = KYC_PP_MATCH_FILTER_FIELDS[param_name]
            client_value = str(getattr(client, client_field, "") or "").lower()
            if filter_value not in client_value:
                keep_match = False
                break
        if keep_match:
            filtered_matches.append(match)
    return filtered_matches


def _build_kyc_pp_match_action_items(match):
    document = match["document"]
    client = match["client"]
    actions = []
    action_keys = set()
    fields_with_actions = set()

    is_pm = isinstance(client, Kyc_pm)
    field_map = KYC_PM_DOCUMENT_FIELD_MAP if is_pm else KYC_PP_DOCUMENT_FIELD_MAP

    client_filiale = getattr(client, "FILIALE", "").strip()
    client_type_val = "pm" if is_pm else "pp"
    field_sources = _get_field_sources(client_filiale, client_type_val)
                                                                                         
    custom_labels = _kyc_custom_field_labels(client_type_val, client_filiale)

    def field_display(field_name):
        return custom_labels.get(field_name) or field_name

    def add_action(kind, field, text):
        normalized_field = (field or "").strip().upper()
        key = (kind, normalized_field) if normalized_field else (kind, "", _normalize_match_value(text))
        if not text or key in action_keys:
            return
        action_keys.add(key)
        fields_with_actions.add(normalized_field)
        actions.append({"kind": kind, "field": normalized_field, "text": text})

    for suggestion in match.get("suggestions", []):
        field = (suggestion.get("field") or "").strip().upper()
        document_value = suggestion.get("document_value", "") or suggestion.get("value", "")
        if not field or not document_value:
            continue
        add_action("complete", field, f"{field_display(field)}: {document_value}")

    fields_in_order = []
    fields_seen = set()
    for kyc_field, _, _ in field_map:
        if kyc_field not in fields_seen:
            fields_seen.add(kyc_field)
            fields_in_order.append(kyc_field)

    for kyc_field in fields_in_order:
        mapped_fields = [
            (document_field, label)
            for field_name, document_field, label in field_map
            if field_name == kyc_field
        ]
        document_values = []
        for document_field, label in mapped_fields:
            allowed_source = field_sources.get(kyc_field)
            if not _document_source_allows(allowed_source, document.document_type):
                continue
            val = getattr(document, document_field, "")
            if val:
                document_values.append((val, label))
                
        kyc_value = getattr(client, kyc_field, "")
        if not document_values or _is_empty_kyc_value(kyc_value):
            continue

        if not is_pm and kyc_field in {"DATNAIS", "DATVALID"}:
            values_match = any(_date_values_match(document_value, kyc_value) for document_value, _ in document_values)
        elif not is_pm and kyc_field == "PAYNAIS":
            values_match = any(_nationality_values_match(document_value, kyc_value) for document_value, _ in document_values)
        else:
            values_match = any(_values_match(document_value, kyc_value) for document_value, _ in document_values)

        if values_match:
            continue

        document_value, label = document_values[0]
        if kyc_field in custom_labels:
            modify_label = f"{custom_labels[kyc_field]} ({kyc_field})"
        else:
            modify_label = f"{label} ({kyc_field})"
        add_action("modify", kyc_field, f"{modify_label}: {kyc_value or '-'} -> {document_value}")

    expired_match = match.get("expired_document_match")
    if expired_match and "DATVALID" not in fields_with_actions:
        allowed_source = field_sources.get("DATVALID")
        if not allowed_source or (expired_match.document
                                  and _document_source_allows(allowed_source, expired_match.document.document_type)):
            add_action(
                "modify",
                "DATVALID",
                f"{field_display('DATVALID')}: {expired_match.old_validity_date or '-'} -> {expired_match.document_validity_date or '-'}",
            )

    for action in match.get("extra_action_items") or []:
        add_action(action.get("kind") or "modify", action.get("field") or "", action.get("text") or "")

    return actions


@login_required
def export_document_extraction_matches(request):
    requested_scope_params = _document_match_scope_params(request.GET)
    matches = []

    last_match_result = request.session.get(LAST_KYC_PP_MATCH_RESULT_SESSION_KEY)
    stored_matches, _, stored_params = _hydrate_kyc_pp_match_result(last_match_result)
    client_type = requested_scope_params.get("client_type", "pp")

    if stored_matches and stored_params == requested_scope_params:
        matches = stored_matches
    else:
        documents = _filtered_document_extractions_from_params(requested_scope_params, user=request.user)
        if client_type == "pm":
            matches, _ = _build_kyc_pm_document_matches(documents, result_limit=None)
        else:
            matches, _ = _build_kyc_pp_document_matches(documents, result_limit=None)

    matches = _merge_kyc_pp_match_lists([matches])
    matches = _filter_matches_for_user_scope(matches, request.user)
    matches = _filter_kyc_pp_matches(matches, request.GET)
                                                                                        
    is_match_validator = (KycMatchValidatorRole.user_can_validate(request.user)
                          or KycMatchValidatorRole.user_can_reject(request.user))
    export_status = (request.GET.get("match_status") or "active").strip()
    if export_status not in DOCUMENT_MATCH_STATUS_FILTERS:
        export_status = "active"
    if not is_match_validator:
        export_status = "validated"
    matches = _annotate_matches_with_decisions(matches, client_type, export_status)

    selected_import_batch = (requested_scope_params.get("import_batch") or "").strip()
    expired_document_matches_qs = KycExpiredDocumentScanMatch.objects.select_related("client", "document").filter(
        status="a_valider"
    )
    if not is_match_validator:
        expired_document_matches_qs = expired_document_matches_qs.none()
    if selected_import_batch:
        expired_document_matches_qs = expired_document_matches_qs.filter(document__import_batch=selected_import_batch)
    expired_match_by_client_id = {}
    for expired_match in expired_document_matches_qs.order_by("-match_rate", "-scan_date")[:500]:
        if expired_match.client_id in expired_match_by_client_id:
            continue
        expired_match_by_client_id[expired_match.client_id] = expired_match

    for match in matches:
        expired_match = expired_match_by_client_id.pop(match["client"].pk, None)
        if expired_match:
            match["expired_document_match"] = expired_match

    matched_idp_keys = {
        _normalize_match_value(getattr(match["client"], "IDM" if client_type == "pm" else "IDP", "")) or str(match["client"].pk)
        for match in matches
    }
    standalone_expired_matches = []
    if client_type == "pp":
        for expired_match in expired_match_by_client_id.values():
            expired_idp_key = _normalize_match_value(getattr(expired_match.client, "IDP", "") or expired_match.idp)
            expired_key = expired_idp_key or str(expired_match.client_id)
            if expired_key in matched_idp_keys:
                continue
            matched_idp_keys.add(expired_key)
            standalone_expired_matches.append(expired_match)

    response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
    filename = timezone.localtime(timezone.now()).strftime(f"correspondances_kyc_{client_type}_%Y%m%d_%H%M.csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write("\ufeff")

    writer = csv.writer(response, delimiter=";")
    writer.writerow([
        "CLIENT",
        "IDM" if client_type == "pm" else "IDP",
        "FILIALE",
        "AGENCE",
        "TYPE DE DOCUMENT",
        "TAUX CORRESPONDANCE",
        "A completer / A modifier",
        "Numero document",
        "Numero d'identification nationale",
    ])

    for match in matches:
        document = match["document"]
        client = match["client"]
        action_items = _build_kyc_pp_match_action_items(match)
        client_id_val = getattr(client, "IDM", "") if client_type == "pm" else getattr(client, "IDP", "")

        writer.writerow([
            client.CLIENT,
            client_id_val,
            client.FILIALE,
            client.AGENCE,
            _document_type_label(document.document_type),
            match.get("match_rate", 0),
            " | ".join(action["text"] for action in action_items),
            document.numero_document,
            document.numero_identification_nationale,
        ])

    for expired_match in standalone_expired_matches:
        client = expired_match.client
        document = expired_match.document
        writer.writerow([
            getattr(client, "CLIENT", "") or expired_match.client_code,
            getattr(client, "IDP", "") or expired_match.idp,
            getattr(client, "FILIALE", "") or expired_match.filiale,
            getattr(client, "AGENCE", "") or expired_match.agence,
            _document_type_label(document.document_type) if document else "",
            expired_match.match_rate or 0,
            f"DATVALID: {expired_match.old_validity_date or '-'} -> {expired_match.document_validity_date or '-'}",
            getattr(document, "numero_document", "") if document else "",
            getattr(document, "numero_identification_nationale", "") if document else "",
        ])
    return response


def _build_import_batch_name(request, uploaded_files):
    requested_name = (request.POST.get("batch_name") or "").strip()
    if requested_name:
        return requested_name[:120]

    first_file = uploaded_files[0].name if uploaded_files else "documents"
    timestamp = timezone.localtime(timezone.now()).strftime("%Y%m%d-%H%M%S")
    return f"LOT-{timestamp}-{os.path.splitext(os.path.basename(first_file))[0]}"[:120]


DOCUMENT_TYPE_UNRECOGNIZED = "non_reconnu"


def _document_type_label_map():
    """Libelles de tous les types : codes fixes + types configures en admin +
    sentinelle 'non reconnu'. Sert a l'affichage et au mode lot mixte."""
    labels = dict(DOCUMENT_EXTRACTION_TYPE_CHOICES)
    for dt in KycDocumentType.objects.all():
        labels.setdefault(dt.code, dt.label)
    labels[DOCUMENT_TYPE_UNRECOGNIZED] = "Type non reconnu"
    labels["auto"] = "Automatique (lot mixte)"
    return labels


def _document_type_label(code, label_map=None):
    label_map = label_map or _document_type_label_map()
    return label_map.get(code, code or "-")


def _document_type_correction_options(client_type=None):
    """Options du menu de correction manuelle : types configures (filtres par
    type de client si fourni) + 'non reconnu'."""
    options = []
    seen = set()
    for dt in KycDocumentType.objects.all().order_by("label"):
        if client_type and dt.client_type and dt.client_type != client_type:
            continue
        if dt.code in seen:
            continue
        seen.add(dt.code)
        options.append({"code": dt.code, "label": dt.label})
    options.append({"code": DOCUMENT_TYPE_UNRECOGNIZED, "label": "Type non reconnu"})
    return options


def _document_type_filter_options():
    """Options du filtre par type dans la recherche (types configures + non reconnu)."""
    options = []
    seen = set()
    for dt in KycDocumentType.objects.all().order_by("label"):
        if dt.code in seen:
            continue
        seen.add(dt.code)
        options.append({"code": dt.code, "label": dt.label})
    options.append({"code": DOCUMENT_TYPE_UNRECOGNIZED, "label": "Type non reconnu"})
    return options


def _format_document_extraction_record(record):
    fields = {
        field_name: getattr(record, field_name, "")
        for field_name, _ in DOCUMENT_EXTRACTION_FIELD_LABELS
    }
    return {
        "id": record.pk,
        "filename": record.original_filename or os.path.basename(record.uploaded_file.name),
        "source_filename": record.source_filename,
        "file_url": record.uploaded_file.url if record.uploaded_file else "",
        "document_type": _document_type_label(record.document_type),
        "document_type_code": record.document_type,
        "import_batch": record.import_batch,
        "page_number": record.page_number,
        "page_range": record.page_range,
        "text": record.extracted_text,
        "fields": fields,
        "warnings": [warning for warning in record.extraction_warnings.splitlines() if warning],
        "field_rows": [
            {"label": label, "value": fields.get(field_name)}
            for field_name, label in DOCUMENT_EXTRACTION_FIELD_LABELS
            if fields.get(field_name)
        ],
    }


def _fill_document_extraction_fields(record, extraction):
    extracted_fields = extraction.get("fields") or {}
    for field_name, _ in DOCUMENT_EXTRACTION_FIELD_LABELS:
        setattr(record, field_name, extracted_fields.get(field_name, ""))
    record.extracted_text = extraction.get("text") or ""
    record.extraction_warnings = "\n".join(extraction.get("warnings") or [])
    record.page_number = extraction.get("page_number") or record.page_number
    record.page_range = extraction.get("page_range") or record.page_range


DOCUMENT_EXTRACTION_TYPE_LABELS = dict(DOCUMENT_EXTRACTION_TYPE_CHOICES)


def _apply_detected_document_type(record, extraction, requested_type):
    """Affecte le type detecte. Deux modes :
    - "auto" (lot mixte) : chaque document est classe individuellement ;
      si rien n'est reconnu -> 'non_reconnu' + avertissement.
    - type impose : on garde le choix, mais on signale un ecart de detection."""
    label_map = _document_type_label_map()
    detected_type = extraction.get("detected_document_type") or ""

    if requested_type == "auto":
        if detected_type and detected_type in label_map:
            record.document_type = detected_type
        else:
            record.document_type = DOCUMENT_TYPE_UNRECOGNIZED
            warnings = extraction.setdefault("warnings", [])
            warning = "Type non reconnu automatiquement : a corriger manuellement."
            if warning not in warnings:
                warnings.append(warning)
            record.extraction_warnings = "\n".join(warnings)
        return

    if detected_type not in label_map:
        return

    if detected_type != requested_type:
        warning = (
            "Type ajuste automatiquement: le document semble etre "
            f"{label_map.get(detected_type, detected_type)} alors que "
            f"{label_map.get(requested_type, requested_type)} avait ete selectionne."
        )
        warnings = extraction.setdefault("warnings", [])
        if warning not in warnings:
            warnings.append(warning)
        record.extraction_warnings = "\n".join(warnings)

    record.document_type = detected_type


def _hash_django_file(django_file):
    """SHA-256 du contenu d'un fichier uploade (lecture par blocs, curseur remis a zero)."""
    hasher = hashlib.sha256()
    try:
        django_file.seek(0)
    except Exception:
        pass
    for chunk in iter(lambda: django_file.read(1024 * 1024), b""):
        hasher.update(chunk)
    try:
        django_file.seek(0)
    except Exception:
        pass
    return hasher.hexdigest()


def _hash_bytes(content_bytes):
    return hashlib.sha256(content_bytes).hexdigest()


def _find_duplicate_document(file_hash):
    if not file_hash:
        return None
    return (
        KycDocumentExtraction.objects
        .filter(file_hash=file_hash)
        .order_by("-created_at")
        .first()
    )


def _save_uploaded_document_record(uploaded_file, document_type, user, import_batch="", source_filename="",
                                   client_type="pp", file_hash=""):
    """Sauvegarde rapide SANS OCR : le document est mis en file d'attente
    (extraction_status=pending) et traite par `manage.py process_document_ocr`."""
    record = KycDocumentExtraction(
        document_type=document_type,
        original_filename=os.path.basename(uploaded_file.name),
        source_filename=source_filename or os.path.basename(uploaded_file.name),
        import_batch=import_batch,
        uploaded_by=user,
        client_type=client_type,
        extraction_status="pending",
        file_hash=file_hash or _hash_django_file(uploaded_file),
    )
    record.uploaded_file.save(uploaded_file.name, uploaded_file, save=False)
    record.save()
    return record, None


                                                                 
ZIP_MAX_MEMBERS = 2000
ZIP_MAX_TOTAL_BYTES = 500 * 1024 * 1024                       
SINGLE_FILE_MAX_BYTES = 50 * 1024 * 1024                                

                                                                               
                                                                              
                                                                               
                                                                      
ALLOWED_DOCUMENT_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".tif", ".tiff", ".zip"}

                                                                                
                                                       
_FILE_SIGNATURES = {
    ".pdf": [b"%PDF"],
    ".png": [b"\x89PNG\r\n\x1a\n"],
    ".jpg": [b"\xff\xd8\xff"],
    ".jpeg": [b"\xff\xd8\xff"],
    ".tif": [b"II*\x00", b"MM\x00*"],
    ".tiff": [b"II*\x00", b"MM\x00*"],
    ".zip": [b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08"],
}


def _validate_uploaded_document(uploaded_file):
    """Retourne un message d'erreur si le fichier est refuse, sinon None.

    Verifie l'extension (liste blanche) puis la signature binaire. Le pointeur
    de lecture est remis a zero apres inspection pour ne pas perturber la suite.
    """
    extension = os.path.splitext(uploaded_file.name)[1].lower()
    if extension not in ALLOWED_DOCUMENT_EXTENSIONS:
        autorisees = ", ".join(sorted(ALLOWED_DOCUMENT_EXTENSIONS))
        return f"Type de fichier non autorise ({uploaded_file.name}): extensions acceptees = {autorisees}."

    signatures = _FILE_SIGNATURES.get(extension)
    if signatures:
        try:
            uploaded_file.seek(0)
            header = uploaded_file.read(8)
        finally:
            uploaded_file.seek(0)
        if not any(header.startswith(sig) for sig in signatures):
            return f"Contenu incoherent avec l'extension ({uploaded_file.name}): fichier refuse."
    return None


def _save_zip_document_record(zip_file, member_name, document_type, user, import_batch, archive_name,
                              client_type="pp", skip_duplicates=True):
    safe_name = os.path.basename(member_name)
    _, extension = os.path.splitext(safe_name)
    if extension.lower() not in SUPPORTED_EXTENSIONS:
        return None, f"Format ignore dans le ZIP: {member_name}"

    with zip_file.open(member_name) as member:
        content_bytes = member.read()

    file_hash = _hash_bytes(content_bytes)
    if skip_duplicates:
        duplicate = _find_duplicate_document(file_hash)
        if duplicate:
            return None, f"Doublon ignore ({safe_name} deja charge dans le lot {duplicate.import_batch or '?'})"

    content = ContentFile(content_bytes, name=safe_name)
    record = KycDocumentExtraction(
        document_type=document_type,
        original_filename=safe_name,
        source_filename=os.path.basename(archive_name or "archive.zip"),
        import_batch=import_batch,
        uploaded_by=user,
        client_type=client_type,
        extraction_status="pending",
        file_hash=file_hash,
    )
    record.uploaded_file.save(safe_name, content, save=False)
    record.save()
    return record, None


def _save_grouped_pdf_job(uploaded_file, document_type, user, import_batch, pages_per_document, client_type="pp"):
    """Mode PDF groupe : on stocke le fichier et on cree un job OCR dedie.
    Les enregistrements par piece sont crees par le worker lors du decoupage."""
    if os.path.splitext(uploaded_file.name)[1].lower() != ".pdf":
        raise ValueError("Le mode document groupe accepte uniquement un fichier PDF.")

    shared_file_name = f"grouped_{uuid.uuid4().hex}_{os.path.basename(uploaded_file.name)}"
    container = KycDocumentExtraction(
        document_type=document_type,
        original_filename=os.path.basename(uploaded_file.name),
        source_filename=os.path.basename(uploaded_file.name),
        import_batch=import_batch,
        uploaded_by=user,
        client_type=client_type,
        extraction_status="pending",
        file_hash=_hash_django_file(uploaded_file),
    )
    container.uploaded_file.save(shared_file_name, uploaded_file, save=False)
    container.save()

    job = KycDocumentOcrJob.objects.create(
        import_batch=import_batch,
        mode="grouped_pdf",
        client_type=client_type,
        document_type=document_type,
        pages_per_document=pages_per_document,
        grouped_source_file=container.uploaded_file.name,
        grouped_original_name=os.path.basename(uploaded_file.name),
        created_by=user,
        message="PDF groupe en attente de decoupage OCR",
    )
    return container, job


                                                                           
                                                                                
                                                                              
                                                                                 
KYC_PP_CONFIG_FIELDS = [
    "CLIENT", "EXPL", "FILIALE", "AGENCE", "LIB_AGENCE", "IDP", "PAYNAIS",
    "PROFESSION", "SALAIRE", "NUMID", "CODAPE", "TEL", "DATNAIS", "ADRESSE",
    "DATVALID", "ORIGINE_REV", "INTITULE_COMPTE", "EMPLOYEUR", "PAYS_RESID",
    "LIEU_DELIVRANCE_CIN", "BOITE_POSTALE", "CONSENT_BIC", "DATOUV", "PPE",
    "DEVISE", "RESID", "DATEREV", "RISQUE",
]

KYC_PM_CONFIG_FIELDS = [
    "CLIENT", "EXPL", "FILIALE", "AGENCE", "LIB_AGENCE", "IDM", "CODAPE",
    "AGEC", "CAPITAL", "CA", "RESULTAT", "RCSNO", "ORIGINE_REV", "TEL",
    "INTITULE_COMPTE", "ADRESSE_SOCIALE", "NUMERO_FISCAL", "PAYS_JUR",
    "ACTIONNAIRE", "MANDATAIRE", "BOITE_POSTALE", "CONSENT_BIC", "DATOUV",
    "DEVISE", "RESID", "DATEREV", "PPE", "RISQUE",
]

                                                                            
DOCUMENT_PP_FIELD_LABELS = {
    "CLIENT": "Nom & Prénom",
    "NUMID": "Numéro de document / NIN/NPI",
    "DATNAIS": "Date de naissance",
    "DATVALID": "Date d'expiration",
    "PAYNAIS": "Nationalité / Pays de naissance",
    "ADRESSE": "Adresse",
    "ORIGINE_REV": "Origine des revenus",
    "LIEU_DELIVRANCE_CIN": "Lieu de délivrance CIN",
    "BOITE_POSTALE": "Boîte postale",
}

DOCUMENT_PM_FIELD_LABELS = {
    "CLIENT": "Raison sociale / Dénomination",
    "RCSNO": "Registre du commerce (RCS/RCCM)",
    "NUMERO_FISCAL": "Numéro fiscal (NIF)",
    "ADRESSE_SOCIALE": "Adresse sociale / Siège",
    "BOITE_POSTALE": "Boîte postale",
}


def _kyc_custom_field_labels(client_type, filiale=""):
    """Intitules metier des champs KYC definis dans /kyc-field-config/ :
    la regle filiale prime, sinon la regle globale, sinon {}."""
    configs = [c for c in _get_cached_field_visibility_configs() if c.client_type == client_type]
    if filiale:
        for c in configs:
            if filiale in (c.filiales or []) and (c.field_labels or {}):
                return dict(c.field_labels)
    for c in configs:
        if not c.filiales and (c.field_labels or {}):
            return dict(c.field_labels)
    return {}


@login_required
def document_extraction(request):
    extraction = None
                                                                                        
    valid_document_types = {dt.code: dt.label for dt in KycDocumentType.objects.all()}

    user = request.user
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
    is_group_user = (user.filiale in ["BOA Group", "BOA GROUP"]) or (user.organe in users_groupe) or (not user.filiale)
    legacy_can_insert = (user.organe == "DSI") or (user.organe == "PASS" and is_group_user)
                                                                        
                                                               
    screening_perms = KycScreeningAccess.perms_for(user, legacy_can_insert=legacy_can_insert)
    can_insert_batches = screening_perms["can_upload_batches"]
    can_run_matching = screening_perms["can_run_matching"]

    liste_filiales = [choice[0] for choice in Filiales]

                                                              
    if is_group_user:
        selected_filiale = request.GET.get("filiale", "").strip()
    else:
        selected_filiale = getattr(user, "filiale", "").strip()

                                                                               
                                                                                   
                                                                                     
    pp_custom_labels = _kyc_custom_field_labels("pp", selected_filiale)
    pm_custom_labels = _kyc_custom_field_labels("pm", selected_filiale)
    pp_fields = [(f, pp_custom_labels.get(f) or DOCUMENT_PP_FIELD_LABELS.get(f, f)) for f in KYC_PP_CONFIG_FIELDS]
    pm_fields = [(f, pm_custom_labels.get(f) or DOCUMENT_PM_FIELD_LABELS.get(f, f)) for f in KYC_PM_CONFIG_FIELDS]

    if request.method == "POST":
        action = request.POST.get("action")
                                                                                     
                                                                              
        if action == "decide_match":
            decision_value = (request.POST.get("decision") or "").strip()                                  
            document_id = (request.POST.get("document_id") or "").strip()
            client_pk = (request.POST.get("client_id") or "").strip()
            decide_client_type = (request.POST.get("match_client_type") or "pp").strip()
            return_url = request.POST.get("return_url") or f"{reverse('document_extraction')}#consulter"

            can_validate = KycMatchValidatorRole.user_can_validate(request.user)
            can_reject = KycMatchValidatorRole.user_can_reject(request.user)
            if decision_value == "validated" and not can_validate:
                messages.error(request, "Votre profil n'est pas autorise a valider les correspondances.")
                return redirect(return_url)
            if decision_value == "rejected" and not can_reject:
                messages.error(request, "Votre profil n'est pas autorise a rejeter les correspondances.")
                return redirect(return_url)
            if decision_value not in ("validated", "rejected", "pending") or not document_id.isdigit() or not client_pk.isdigit():
                messages.error(request, "Decision de correspondance invalide.")
                return redirect(return_url)

            document = get_object_or_404(KycDocumentExtraction, pk=int(document_id))
            client_model = Kyc_pm if decide_client_type == "pm" else Kyc_pp
            client = client_model.objects.filter(pk=int(client_pk)).first()
            client_code = getattr(client, "CLIENT", "") if client else ""
            client_filiale = getattr(client, "FILIALE", "") if client else ""
            client_agence = getattr(client, "AGENCE", "") if client else ""
            try:
                match_rate_val = int(float(request.POST.get("match_rate") or 0))
            except ValueError:
                match_rate_val = 0

            decision, _created = KycMatchDecision.objects.get_or_create(
                document=document, client_type=decide_client_type, client_id=int(client_pk),
                defaults={"client_code": client_code, "filiale": client_filiale, "agence": client_agence},
            )
            decision.client_code = client_code or decision.client_code
            decision.filiale = client_filiale or decision.filiale
            decision.agence = client_agence or decision.agence
            decision.match_rate = match_rate_val or decision.match_rate
            decision.status = decision_value
            decision.decided_by = request.user
            decision.decided_at = timezone.now()
            decision.save()
            labels = {"validated": "validee", "rejected": "rejetee", "pending": "remise a valider"}
            messages.success(request, f"Correspondance {labels.get(decision_value, '')}.")
            return redirect(return_url)

                                                                                                   
        if not can_insert_batches:
            messages.error(request, "Vous n'avez pas l'autorisation d'effectuer cette action.")
            return redirect("document_extraction")

        if action == "retry_failed_ocr":
            retry_batch = (request.POST.get("import_batch") or "").strip()
            if not retry_batch:
                messages.error(request, "Lot introuvable pour la relance OCR.")
                return redirect("document_extraction")
            retry_scope = (request.POST.get("scope") or "failed").strip()
            retry_statuses = ["failed", "done"] if retry_scope == "all" else ["failed"]
            requeued = KycDocumentExtraction.objects.filter(
                import_batch=retry_batch, extraction_status__in=retry_statuses
            ).update(extraction_status="pending", extraction_warnings="")
            has_active_job = KycDocumentOcrJob.objects.filter(
                import_batch=retry_batch, mode="files", status__in=("pending", "running")
            ).exists()
            if requeued and not has_active_job:
                KycDocumentOcrJob.objects.create(
                    import_batch=retry_batch,
                    mode="files",
                    created_by=request.user,
                    progress_total=requeued,
                    message="Retraitement OCR de tout le lot" if retry_scope == "all" else "Relance OCR des documents en echec",
                )
            if requeued:
                messages.success(request, f"{requeued} document(s) remis en file d'attente OCR pour le lot {retry_batch}.")
            else:
                messages.info(request, "Aucun document a relancer pour ce lot.")
            return redirect(f"{reverse('document_extraction')}?{urlencode({'uploaded_batch': retry_batch})}#charger")
        if action == "correct_document_type":
            extraction_id = (request.POST.get("extraction_id") or "").strip()
            new_type = (request.POST.get("new_document_type") or "").strip()
            return_batch = (request.POST.get("import_batch") or "").strip()
            label_map = _document_type_label_map()
            if not extraction_id.isdigit() or new_type not in label_map:
                messages.error(request, "Correction de type invalide.")
            else:
                record = get_object_or_404(KycDocumentExtraction, pk=int(extraction_id))
                old_type = record.document_type
                record.document_type = new_type
                                                                                       
                kept_warnings = [
                    w for w in record.extraction_warnings.splitlines()
                    if w and "Type non reconnu" not in w and "Type ajuste" not in w
                ]
                record.extraction_warnings = "\n".join(kept_warnings)
                record.save(update_fields=["document_type", "extraction_warnings"])
                                                                                            
                if new_type not in ("auto", DOCUMENT_TYPE_UNRECOGNIZED) and record.extracted_text:
                    try:
                        client_filiale = getattr(request.user, "filiale", "") or ""
                        learn_document_keywords(record.extracted_text, new_type, filiale=client_filiale.strip() or None)
                    except Exception:
                        pass
                messages.success(
                    request,
                    f"Type corrige : {_document_type_label(old_type, label_map)} -> {_document_type_label(new_type, label_map)}.",
                )
            if return_batch:
                return redirect(f"{reverse('document_extraction')}?{urlencode({'import_batch': return_batch})}#base")
            return redirect(f"{reverse('document_extraction')}#base")
        if action == "save_document_field_sources":
                                      
            if is_group_user:
                target_filiale = request.GET.get("filiale", "").strip()
                apply_to_all = request.POST.get("apply_to_all_filiales") == "1"
            else:
                target_filiale = getattr(user, "filiale", "").strip()
                apply_to_all = False

                                                               
            if not target_filiale:
                pp_config_list = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pp") if not c.filiales]
                pm_config_list = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pm") if not c.filiales]
            else:
                pp_config_list = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pp") if target_filiale in (c.filiales or [])]
                pm_config_list = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pm") if target_filiale in (c.filiales or [])]
                if not pp_config_list:
                    pp_config_list = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pp") if not c.filiales]
                if not pm_config_list:
                    pm_config_list = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pm") if not c.filiales]

            pp_active_db_fields = set()
            if pp_config_list:
                pp_active_db_fields.update(pp_config_list[0].empty_check_fields or [])
            if not pp_active_db_fields:
                pp_active_db_fields = {"CLIENT", "NUMID", "DATNAIS", "DATVALID", "PAYNAIS", "ADRESSE", "ORIGINE_REV"}

            pm_active_db_fields = set()
            if pm_config_list:
                pm_active_db_fields.update(pm_config_list[0].empty_check_fields or [])
            if not pm_active_db_fields:
                pm_active_db_fields = {"CLIENT", "RCSNO", "NUMERO_FISCAL", "ADRESSE_SOCIALE"}

            filtered_pp_fields = [
                (f_name, label)
                for f_name, label in pp_fields
                if f_name in pp_active_db_fields
            ]
            filtered_pm_fields = [
                (f_name, label)
                for f_name, label in pm_fields
                if f_name in pm_active_db_fields
            ]

            for client_type_item in ["pp", "pm"]:
                fields_list = filtered_pp_fields if client_type_item == "pp" else filtered_pm_fields
                
                if apply_to_all:
                                                   
                    global_configs = [c for c in KycFieldVisibilityConfig.objects.filter(client_type=client_type_item) if not c.filiales]
                    if global_configs:
                        config = global_configs[0]
                    else:
                        config = KycFieldVisibilityConfig(client_type=client_type_item, filiales=[])
                    
                    sources_dict = dict(config.field_sources or {})
                    for field_name, _ in fields_list:
                                                                                            
                        vals = [v.strip() for v in request.POST.getlist(f"source_{client_type_item}_{field_name}") if v.strip()]
                        if vals:
                            sources_dict[field_name] = vals if len(vals) > 1 else vals[0]
                        else:
                            sources_dict.pop(field_name, None)
                    config.field_sources = sources_dict
                    config.save()

                                                                    
                    for oc in KycFieldVisibilityConfig.objects.filter(client_type=client_type_item):
                        oc.field_sources = sources_dict
                        oc.save()
                else:
                                                              
                    if not target_filiale:
                                                               
                        global_configs = [c for c in KycFieldVisibilityConfig.objects.filter(client_type=client_type_item) if not c.filiales]
                        if global_configs:
                            config = global_configs[0]
                        else:
                            config = KycFieldVisibilityConfig(client_type=client_type_item, filiales=[])
                    else:
                                                                                   
                        filiale_configs = [c for c in KycFieldVisibilityConfig.objects.filter(client_type=client_type_item) if target_filiale in (c.filiales or [])]
                        if filiale_configs:
                            config = filiale_configs[0]
                        else:
                            config = KycFieldVisibilityConfig(client_type=client_type_item, filiales=[target_filiale])
                    
                    sources_dict = dict(config.field_sources or {})
                    for field_name, _ in fields_list:
                                                                                            
                        vals = [v.strip() for v in request.POST.getlist(f"source_{client_type_item}_{field_name}") if v.strip()]
                        if vals:
                            sources_dict[field_name] = vals if len(vals) > 1 else vals[0]
                        else:
                            sources_dict.pop(field_name, None)
                    config.field_sources = sources_dict
                    config.save()
            
                                                            
            global _field_visibility_configs_cache
            _field_visibility_configs_cache = None

            if apply_to_all:
                messages.success(request, "Configuration des sources documentaires enregistrée et appliquée à TOUTES les filiales.")
            else:
                messages.success(request, "Configuration des sources documentaires par champ enregistrée avec succès.")
            return_url = request.POST.get("return_url") or f"{reverse('document_extraction')}#sources"
            return redirect(return_url)
        uploaded_files = request.FILES.getlist("documents")
        if not uploaded_files and request.FILES.get("document"):
            uploaded_files = [request.FILES.get("document")]
        document_type = request.POST.get("document_type") or "piece_identite"
        import_mode = request.POST.get("import_mode") or "single"
        client_type = request.POST.get("client_type") or "pp"
        try:
            pages_per_document = max(int(request.POST.get("pages_per_document") or 1), 1)
        except ValueError:
            pages_per_document = 1

        if document_type != "auto" and document_type not in valid_document_types:
            messages.error(request, "Veuillez choisir un type de document valide.")
            return redirect("document_extraction")

        skip_duplicates = request.POST.get("skip_duplicates", "1") == "1"

        if not uploaded_files:
            messages.error(request, "Veuillez selectionner au moins un document a analyser.")
        else:
            import_batch = _build_import_batch_name(request, uploaded_files)
            created_records = []
            errors = []
            duplicates_skipped = 0
            grouped_job = None

            oversized = [f for f in uploaded_files if f.size > SINGLE_FILE_MAX_BYTES]
            for f in oversized:
                errors.append(f"Fichier ignore ({f.name}): taille > {SINGLE_FILE_MAX_BYTES // (1024*1024)} Mo.")
            uploaded_files = [f for f in uploaded_files if f.size <= SINGLE_FILE_MAX_BYTES]

                                                                                 
                                                      
            type_valides = []
            for f in uploaded_files:
                type_error = _validate_uploaded_document(f)
                if type_error:
                    errors.append(type_error)
                else:
                    type_valides.append(f)
            uploaded_files = type_valides

            if not uploaded_files:
                pass
            elif import_mode == "grouped_pdf":
                try:
                    _, grouped_job = _save_grouped_pdf_job(
                        uploaded_files[0],
                        document_type,
                        request.user,
                        import_batch,
                        pages_per_document,
                        client_type=client_type,
                    )
                except Exception as exc:
                    errors.append(str(exc))
            else:
                for uploaded_file in uploaded_files:
                    extension = os.path.splitext(uploaded_file.name)[1].lower()
                    if extension == ".zip":
                        try:
                            with zipfile.ZipFile(uploaded_file) as archive:
                                                                  
                                infos = [i for i in archive.infolist() if not i.is_dir()]
                                total_uncompressed = sum(i.file_size for i in infos)
                                if len(infos) > ZIP_MAX_MEMBERS:
                                    errors.append(f"Archive ignoree ({uploaded_file.name}): plus de {ZIP_MAX_MEMBERS} fichiers.")
                                    continue
                                if total_uncompressed > ZIP_MAX_TOTAL_BYTES:
                                    errors.append(f"Archive ignoree ({uploaded_file.name}): taille decompressee > {ZIP_MAX_TOTAL_BYTES // (1024*1024)} Mo.")
                                    continue
                                for member_name in archive.namelist():
                                    if member_name.endswith("/"):
                                        continue
                                    record, error = _save_zip_document_record(
                                        archive,
                                        member_name,
                                        document_type,
                                        request.user,
                                        import_batch,
                                        uploaded_file.name,
                                        client_type=client_type,
                                        skip_duplicates=skip_duplicates,
                                    )
                                    if record:
                                        created_records.append(record)
                                    if error:
                                        if error.startswith("Doublon ignore"):
                                            duplicates_skipped += 1
                                        else:
                                            errors.append(error)
                        except zipfile.BadZipFile:
                            errors.append(f"Archive ZIP invalide: {uploaded_file.name}")
                    else:
                        try:
                            file_hash = _hash_django_file(uploaded_file)
                            if skip_duplicates:
                                duplicate = _find_duplicate_document(file_hash)
                                if duplicate:
                                    duplicates_skipped += 1
                                    continue
                            record, _ = _save_uploaded_document_record(
                                uploaded_file,
                                document_type,
                                request.user,
                                import_batch=import_batch,
                                client_type=client_type,
                                file_hash=file_hash,
                            )
                            created_records.append(record)
                        except Exception as exc:
                            errors.append(f"{uploaded_file.name}: {exc}")

                if created_records:
                    KycDocumentOcrJob.objects.create(
                        import_batch=import_batch,
                        mode="files",
                        client_type=client_type,
                        document_type=document_type,
                        created_by=request.user,
                        progress_total=len(created_records),
                        message="Lot en attente de traitement OCR",
                    )

            if created_records or grouped_job:
                if grouped_job:
                    messages.success(
                        request,
                        f"PDF groupe charge dans le lot {import_batch}. Le decoupage et l'OCR "
                        "s'executent en arriere-plan : suivez la progression ci-dessous.",
                    )
                else:
                    messages.success(
                        request,
                        f"{len(created_records)} document(s) charge(s) dans le lot {import_batch}. "
                        "L'analyse OCR s'execute en arriere-plan : suivez la progression ci-dessous.",
                    )
                request.session[LAST_UPLOADED_DOCUMENT_BATCH_SESSION_KEY] = import_batch
                request.session.modified = True
            if duplicates_skipped:
                messages.info(request, f"{duplicates_skipped} doublon(s) ignore(s) (fichier identique deja charge).")
            if errors:
                messages.warning(request, f"{len(errors)} element(s) non importe(s): " + " | ".join(errors[:5]))
            if created_records or grouped_job:
                return redirect(f"{reverse('document_extraction')}?{urlencode({'uploaded_batch': import_batch, 'client_type': client_type})}#charger")

    documents = _filtered_document_extractions_from_request(request)
    selected_document_type = request.GET.get("document_type", "")
    selected_import_batch = (request.GET.get("import_batch") or "").strip()
    uploaded_batch = (request.GET.get("uploaded_batch") or "").strip()
    uploaded_batch_from_session = False
    if not uploaded_batch:
                                                                                  
                                                                                      
        uploaded_batch = (request.session.get(LAST_UPLOADED_DOCUMENT_BATCH_SESSION_KEY) or "").strip()
        uploaded_batch_from_session = bool(uploaded_batch)
    search_query = (request.GET.get("q") or "").strip()
    search_field = request.GET.get("field") or "all"
    document_type_labels = _document_type_label_map()
    if selected_document_type and selected_document_type not in document_type_labels:
        selected_document_type = ""
    if search_field not in {field for field, _ in DOCUMENT_EXTRACTION_SEARCH_FIELDS}:
        search_field = "all"

    selected_extraction_id = request.GET.get("extraction_id")
    if selected_extraction_id and selected_extraction_id.isdigit():
        selected_record = get_object_or_404(KycDocumentExtraction, pk=selected_extraction_id)
        extraction = _format_document_extraction_record(selected_record)

    uploaded_documents = KycDocumentExtraction.objects.none()
    uploaded_documents_count = 0
    uploaded_quality_alerts = []
    uploaded_quality_alerts_count = 0
    uploaded_type_distribution = []
    if uploaded_batch:
        uploaded_documents_queryset = KycDocumentExtraction.objects.filter(import_batch=uploaded_batch)
        uploaded_documents_count = uploaded_documents_queryset.count()
        uploaded_quality_alerts_count = uploaded_documents_queryset.exclude(extraction_warnings="").count()
                                                      
        for row in (
            uploaded_documents_queryset.values("document_type").annotate(n=Count("id")).order_by("-n")
        ):
            code = row["document_type"]
            uploaded_type_distribution.append({
                "code": code,
                "label": _document_type_label(code, document_type_labels),
                "count": row["n"],
                "is_unrecognized": code in ("", DOCUMENT_TYPE_UNRECOGNIZED, "auto"),
            })
        uploaded_documents = list(uploaded_documents_queryset.order_by("-created_at")[:50])
        for document in uploaded_documents:
            document.type_label = _document_type_label(document.document_type, document_type_labels)
            document.type_is_unrecognized = document.document_type in ("", DOCUMENT_TYPE_UNRECOGNIZED, "auto")
            warnings = [warning for warning in document.extraction_warnings.splitlines() if warning]
            if warnings:
                uploaded_quality_alerts.append({
                    "filename": document.original_filename or os.path.basename(document.uploaded_file.name),
                    "document_type": document.type_label,
                    "warnings": warnings[:3],
                })
            if len(uploaded_quality_alerts) >= 5:
                break

    requested_kyc_pp_matching = request.GET.get("match_kyc") == "1"
    client_type = request.GET.get("client_type", "pp")
    selected_match_job = None
    selected_match_job_id = request.GET.get("match_job")
    show_match_job_results = request.GET.get("show_match_results") == "1"
    show_match_result_modal = request.GET.get("result_modal") == "1"
    show_match_child_modal = request.GET.get("child_modal") == "1"
    wants_results_tab = request.GET.get("tab") == "results"
    if selected_match_job_id and selected_match_job_id.isdigit():
        selected_match_job = get_object_or_404(KycDocumentMatchJob, pk=selected_match_job_id)
        if not _user_can_access_document_match_job(request.user, selected_match_job):
            selected_match_job = None

                                                                               
                                                             
    consulted_match_job_ids = set(request.session.get("consulted_match_job_ids") or [])
    if selected_match_job and selected_match_job.status == "completed" and show_match_job_results:
        if selected_match_job.pk not in consulted_match_job_ids:
            consulted_match_job_ids.add(selected_match_job.pk)
            request.session["consulted_match_job_ids"] = sorted(consulted_match_job_ids)
            request.session.modified = True
    selected_match_job_consulted = bool(
        selected_match_job and selected_match_job.status == "completed"
        and selected_match_job.pk in consulted_match_job_ids
    )

    last_match_result = request.session.get(LAST_KYC_PP_MATCH_RESULT_SESSION_KEY)
    active_match_params = None
    is_global_consultation = not request.GET

    if requested_kyc_pp_matching:
        active_match_params = _document_match_scope_params(request.GET)
        request.session[LAST_KYC_PP_MATCH_SESSION_KEY] = active_match_params

    kyc_pp_matches = []
    kyc_pp_match_summary = {
        "documents_checked": 0,
        "documents_matched": 0,
        "clients_matched": 0,
        "suggestions_count": 0,
        "match_rate": 0,
    }
    if selected_match_job:
        active_match_params = selected_match_job.scope_params or {}
        if selected_match_job.status == "completed" and show_match_job_results:
            kyc_pp_matches, stored_summary, stored_params = _hydrate_kyc_pp_match_result(selected_match_job.result)
            if stored_summary is not None:
                kyc_pp_match_summary = stored_summary
                active_match_params = stored_params
                request.session[LAST_KYC_PP_MATCH_SESSION_KEY] = active_match_params
                request.session[LAST_KYC_PP_MATCH_RESULT_SESSION_KEY] = selected_match_job.result
                request.session.modified = True
        elif selected_match_job.status in {"pending", "running", "completed"}:
            kyc_pp_match_summary = {
                "documents_checked": selected_match_job.progress_total,
                "documents_matched": 0,
                "clients_matched": 0,
                "suggestions_count": 0,
                "match_rate": 0,
            }
    elif requested_kyc_pp_matching:
        match_documents = _filtered_document_extractions_from_params(active_match_params, user=request.user)
        sync_filiale_scope = _user_filiale_scope(request.user)
        if client_type == "pm":
            kyc_pp_matches, kyc_pp_match_summary = _build_kyc_pm_document_matches(match_documents, filiale_scope=sync_filiale_scope)
        else:
            kyc_pp_matches, kyc_pp_match_summary = _build_kyc_pp_document_matches(match_documents, filiale_scope=sync_filiale_scope)
        request.session[LAST_KYC_PP_MATCH_RESULT_SESSION_KEY] = _serialize_kyc_pp_matches(
            kyc_pp_matches,
            kyc_pp_match_summary,
            active_match_params,
        )
        matched_batch = (active_match_params.get("import_batch") or "").strip()
        if matched_batch:
            matched_batches = set(request.session.get(KYC_PP_MATCHED_BATCHES_SESSION_KEY) or [])
            matched_batches.add(matched_batch)
            request.session[KYC_PP_MATCHED_BATCHES_SESSION_KEY] = sorted(matched_batches)
        request.session.modified = True
    elif wants_results_tab or is_global_consultation:
                                                                                    
                                                                                   
                                                                                 
                                         
        completed_jobs = KycDocumentMatchJob.objects.filter(
            status="completed",
        ).order_by("-completed_at", "-created_at")
        hydrated_match_lists = []
        documents_checked_total = 0
        for job in completed_jobs:
            job_matches, job_summary, _ = _hydrate_kyc_pp_match_result(job.result)
            if job_matches:
                hydrated_match_lists.append(job_matches)
            if job_summary:
                documents_checked_total += job_summary.get("documents_checked", 0)

        kyc_pp_matches = _merge_kyc_pp_match_lists(hydrated_match_lists)
        if kyc_pp_matches:
            matched_idp_keys = {
                _normalize_match_value(getattr(match["client"], "IDM" if client_type == "pm" else "IDP", "")) or str(match["client"].pk)
                for match in kyc_pp_matches
            }
            kyc_pp_match_summary = {
                "documents_checked": documents_checked_total,
                "documents_matched": len({match["document"].pk for match in kyc_pp_matches}),
                "clients_matched": len(matched_idp_keys),
                "suggestions_count": sum(len(_build_kyc_pp_match_action_items(match)) for match in kyc_pp_matches),
                "match_rate": 0,
            }
            request.session[LAST_KYC_PP_MATCH_SESSION_KEY] = {}
            request.session[LAST_KYC_PP_MATCH_RESULT_SESSION_KEY] = _serialize_kyc_pp_matches(
                kyc_pp_matches,
                kyc_pp_match_summary,
                {},
            )
            request.session.modified = True
            active_match_params = {}
    elif not is_global_consultation:
        kyc_pp_matches, stored_summary, stored_params = _hydrate_kyc_pp_match_result(last_match_result)
        stored_batch = (stored_params.get("import_batch") or "").strip()
        if selected_import_batch and stored_batch != selected_import_batch:
            kyc_pp_matches, stored_summary, stored_params = [], None, {}
        if stored_summary is not None:
            kyc_pp_match_summary = stored_summary
            active_match_params = stored_params

    kyc_pp_matches = _merge_kyc_pp_match_lists([kyc_pp_matches])
                                                                                        
                                                        
    scoped_matches = _filter_matches_for_user_scope(kyc_pp_matches, request.user)
    if len(scoped_matches) != len(kyc_pp_matches):
        kyc_pp_matches = scoped_matches
        kyc_pp_match_summary = dict(kyc_pp_match_summary or {})
        kyc_pp_match_summary.update({
            "documents_matched": len({m["document"].pk for m in kyc_pp_matches}),
            "clients_matched": len({m["client"].pk for m in kyc_pp_matches}),
            "suggestions_count": sum(len(_build_kyc_pp_match_action_items(m)) for m in kyc_pp_matches),
        })
    run_kyc_pp_matching = active_match_params is not None
                                                                                                
    can_validate_matches = KycMatchValidatorRole.user_can_validate(request.user)
    can_reject_matches = KycMatchValidatorRole.user_can_reject(request.user)
    is_match_validator = can_validate_matches or can_reject_matches
    match_status = (request.GET.get("match_status") or "active").strip()
    if match_status not in DOCUMENT_MATCH_STATUS_FILTERS:
        match_status = "active"
    if not is_match_validator:
                                                                              
        match_status = "validated"
    kyc_pp_matches = _annotate_matches_with_decisions(kyc_pp_matches, client_type, match_status)
    kyc_pp_match_total_count = len(kyc_pp_matches)
    kyc_pp_match_filters = _get_kyc_pp_match_filters(request.GET)
    kyc_pp_matches = _filter_kyc_pp_matches(kyc_pp_matches, request.GET)
    expired_document_matches = KycExpiredDocumentScanMatch.objects.select_related("client", "document").filter(
        status="a_valider"
    )
    if not is_match_validator:
                                                                                         
        expired_document_matches = expired_document_matches.none()
    _expired_filiale_scope = _user_filiale_scope(request.user)
    if _expired_filiale_scope:
        expired_document_matches = expired_document_matches.filter(filiale__iexact=_expired_filiale_scope)
        _user_organe = (getattr(request.user, "organe", "") or "").strip()
        _user_agence = (getattr(request.user, "agence", "") or "").strip()
        if _user_organe in ("Chargé Client", "Directeur Agence") and _user_agence:
            expired_document_matches = expired_document_matches.filter(agence__iexact=_user_agence)
    if selected_import_batch:
        expired_document_matches = expired_document_matches.filter(document__import_batch=selected_import_batch)
    unique_expired_document_matches = {}
    for expired_match in expired_document_matches.order_by("-match_rate", "-scan_date")[:500]:
        if expired_match.client_id in unique_expired_document_matches:
            continue
        unique_expired_document_matches[expired_match.client_id] = expired_match
        if len(unique_expired_document_matches) >= 100:
            break
    expired_match_by_client_id = unique_expired_document_matches
    for match in kyc_pp_matches:
        expired_match = expired_match_by_client_id.pop(match["client"].pk, None)
        if expired_match:
            match["expired_document_match"] = expired_match
        match["action_items"] = _build_kyc_pp_match_action_items(match)
        match["action_summary"] = " | ".join(action["text"] for action in match["action_items"])
                                                                                    
        _match_labels = _kyc_custom_field_labels(client_type, getattr(match["client"], "FILIALE", "") or "")
        for _suggestion in match.get("suggestions", []):
            _sf = (_suggestion.get("field") or "").strip().upper()
            _suggestion["field_label"] = _match_labels.get(_sf) or _sf
    expired_document_matches = list(expired_match_by_client_id.values())
    matched_batches = set(request.session.get(KYC_PP_MATCHED_BATCHES_SESSION_KEY) or [])
    if is_group_user:
        match_jobs_qs = KycDocumentMatchJob.objects.all()
        extractions_qs = KycDocumentExtraction.objects.all()
    else:
        match_jobs_qs = KycDocumentMatchJob.objects.filter(created_by__filiale=user.filiale)
        extractions_qs = KycDocumentExtraction.objects.filter(uploaded_by__filiale=user.filiale)

                                                                                   
                                                                                      
    recent_global_jobs = list(
        match_jobs_qs.filter(scope_params__import_batch__isnull=True)
        .order_by("-created_at")[:8]
    )

    def _global_job_covering(since_dt):
        """Dernier job global lance apres `since_dt` (donc couvrant ce lot)."""
        if not since_dt:
            return None
        for gj in recent_global_jobs:
            if gj.created_at >= since_dt:
                return gj
        return None

    uploaded_batch_job_done = False
    uploaded_batch_running_job = None
    uploaded_batch_result_url = ""
    if uploaded_batch:
        uploaded_batch_running_job = (
            match_jobs_qs
            .filter(scope_params__import_batch=uploaded_batch, status__in=["pending", "running"])
            .order_by("-created_at")
            .first()
        )
        if uploaded_batch_running_job:
            follow_params = dict(uploaded_batch_running_job.scope_params or {})
            follow_params["match_job"] = uploaded_batch_running_job.pk
            uploaded_batch_running_job.follow_url = f"{reverse('document_extraction')}?{urlencode(follow_params)}#suivi"
        uploaded_batch_job_done = match_jobs_qs.filter(
            scope_params__import_batch=uploaded_batch,
            status="completed",
        ).exists()
        uploaded_batch_completed_job = (
            match_jobs_qs
            .filter(scope_params__import_batch=uploaded_batch, status="completed")
            .order_by("-completed_at", "-created_at")
            .first()
        )
        if uploaded_batch_completed_job:
            result_params = dict(uploaded_batch_completed_job.scope_params or {})
            result_params["match_job"] = uploaded_batch_completed_job.pk
            result_params["show_match_results"] = "1"
            uploaded_batch_result_url = f"{reverse('document_extraction')}?{urlencode(result_params)}#consulter"
        if not uploaded_batch_job_done and not uploaded_batch_running_job:
                                                                        
            latest_doc_dt = extractions_qs.filter(import_batch=uploaded_batch).aggregate(m=Max("created_at"))["m"]
            covering_job = _global_job_covering(latest_doc_dt)
            if covering_job is not None:
                cover_params = dict(covering_job.scope_params or {})
                cover_params["match_job"] = covering_job.pk
                if covering_job.status == "completed":
                    uploaded_batch_job_done = True
                    cover_params["show_match_results"] = "1"
                    uploaded_batch_result_url = f"{reverse('document_extraction')}?{urlencode(cover_params)}#consulter"
                elif covering_job.status in {"pending", "running"}:
                    covering_job.follow_url = f"{reverse('document_extraction')}?{urlencode(cover_params)}#suivi"
                    uploaded_batch_running_job = covering_job
    uploaded_batch_matching_done = bool(uploaded_batch and (uploaded_batch in matched_batches or uploaded_batch_job_done))
    if uploaded_batch_from_session and uploaded_batch_matching_done:
                                                                                   
                                                                               
        request.session.pop(LAST_UPLOADED_DOCUMENT_BATCH_SESSION_KEY, None)
        request.session.modified = True
        uploaded_batch = ""
        uploaded_batch_running_job = None
        uploaded_batch_result_url = ""
        uploaded_batch_matching_done = False
    show_document_modal = (bool(extraction) and not show_match_child_modal) or request.GET.get("lot_view") == "1"
    recent_match_jobs = list(match_jobs_qs.order_by("-created_at")[:12])
    for job in recent_match_jobs:
        follow_params = dict(job.scope_params or {})
        follow_params["match_job"] = job.pk
        result_params = dict(follow_params)
        result_params["show_match_results"] = "1"
        result_params["result_modal"] = "1"
        job.follow_url = f"{reverse('document_extraction')}?{urlencode(follow_params)}#suivi"
        job.result_url = f"{reverse('document_extraction')}?{urlencode(result_params)}#suivi"
    upload_batch_queue_all = list(
        extractions_qs
        .exclude(import_batch="")
        .values("import_batch")
        .annotate(documents_count=Count("id"), latest_created_at=Max("created_at"))
        .order_by("-latest_created_at")[:12]
    )
    upload_batch_queue = []
    for batch in upload_batch_queue_all:
        batch_name = batch["import_batch"]
        batch["documents_url"] = f"{reverse('document_extraction')}?{urlencode({'import_batch': batch_name})}#base"
        latest_job = (
            match_jobs_qs
            .filter(scope_params__import_batch=batch_name)
            .order_by("-created_at")
            .first()
        )
        if latest_job is None:
                                                                                   
            latest_job = _global_job_covering(batch.get("latest_created_at"))
        batch["job"] = latest_job
        batch["status"] = "pending"
        batch["status_label"] = "En attente"
        batch["start_url"] = f"{reverse('start_document_extraction_match_job')}?{urlencode({'import_batch': batch_name})}"
        batch["follow_url"] = ""
        batch["result_url"] = ""
        if latest_job:
            follow_params = dict(latest_job.scope_params or {})
            follow_params["match_job"] = latest_job.pk
            result_params = dict(follow_params)
            result_params["show_match_results"] = "1"
            result_params["result_modal"] = "1"
            batch["follow_url"] = f"{reverse('document_extraction')}?{urlencode(follow_params)}#suivi"
            batch["result_url"] = f"{reverse('document_extraction')}?{urlencode(result_params)}#suivi"
            if latest_job.status == "completed":
                batch["status"] = "completed"
                batch["status_label"] = "Termine"
            elif latest_job.status in {"pending", "running"}:
                batch["status"] = "running"
                batch["status_label"] = "En cours"
            elif latest_job.status == "failed":
                batch["status"] = "failed"
                batch["status_label"] = "Echec"
        if batch["status"] in {"pending", "failed"}:
            upload_batch_queue.append(batch)

    if not is_match_validator:
                                                                                   
                                                                       
        decided_doc_ids = KycMatchDecision.objects.filter(
            client_type=client_type, status="validated"
        ).values_list("document_id", flat=True).distinct()
        documents = documents.filter(pk__in=decided_doc_ids)

    paginator = Paginator(documents, 20)
    page_obj = paginator.get_page(request.GET.get("page"))
    for document in page_obj:
        document.type_label = _document_type_label(document.document_type, document_type_labels)
        document.type_is_unrecognized = document.document_type in ("", DOCUMENT_TYPE_UNRECOGNIZED, "auto")

    query_params = request.GET.copy()
    query_params.pop("page", None)
    query_params.pop("extraction_id", None)
    query_params.pop("match_kyc", None)

    match_query_params = request.GET.copy()
    match_query_params["match_kyc"] = "1"
    match_query_params.pop("page", None)
    match_query_params.pop("extraction_id", None)

    if active_match_params is not None:
        export_match_params = dict(active_match_params)
    else:
        export_match_params = _document_match_scope_params(request.GET)
    match_filter_hidden_params = {
        key: value
        for key, value in export_match_params.items()
        if key not in KYC_PP_MATCH_FILTER_FIELDS
        and key not in {"page", "extraction_id"}
        and value not in (None, "")
    }
    match_reset_params = dict(match_filter_hidden_params)
    if run_kyc_pp_matching:
        match_reset_params["match_kyc"] = "1"
    export_match_params.update({
        key: value
        for key, value in kyc_pp_match_filters.items()
        if value
    })
    export_match_querystring = urlencode(export_match_params)
    match_reset_querystring = urlencode(match_reset_params)
    selected_match_job_result_url = ""
    selected_match_job_follow_url = f"{reverse('document_extraction')}#suivi"
    selected_match_job_parent_modal_url = ""
    if selected_match_job:
        follow_params = dict(selected_match_job.scope_params or {})
        follow_params["match_job"] = selected_match_job.pk
        selected_match_job_follow_url = f"{reverse('document_extraction')}?{urlencode(follow_params)}#suivi"
        result_params = dict(selected_match_job.scope_params or {})
        result_params["match_job"] = selected_match_job.pk
        result_params["show_match_results"] = "1"
        result_params["result_modal"] = "1"
        selected_match_job_result_url = f"{reverse('document_extraction')}?{urlencode(result_params)}#suivi"
        selected_match_job_parent_modal_url = selected_match_job_result_url
    latest_global_match_job = (
        KycDocumentMatchJob.objects
        .filter(created_by=request.user, scope_params={}, status="completed")
        .order_by("-completed_at", "-created_at")
        .first()
    )
    all_match_results_url = f"{reverse('document_extraction')}#consulter"
    if latest_global_match_job:
        all_results_params = {
            "match_job": latest_global_match_job.pk,
            "show_match_results": "1",
            "result_modal": "1",
        }
        all_match_results_url = f"{reverse('document_extraction')}?{urlencode(all_results_params)}#suivi"

    if selected_match_job and show_match_result_modal:
        base_child_params = dict(selected_match_job.scope_params or {})
        base_child_params["match_job"] = selected_match_job.pk
        base_child_params["show_match_results"] = "1"
        base_child_params["result_modal"] = "1"
        base_child_params["child_modal"] = "1"
        for match in kyc_pp_matches:
            child_params = dict(base_child_params)
            child_params["extraction_id"] = match["document"].pk
            match["child_detail_url"] = f"{reverse('document_extraction')}?{urlencode(child_params)}#suivi"
    pp_sources = {}
    pm_sources = {}
    
    if not selected_filiale:
        pp_config = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pp") if not c.filiales]
        pm_config = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pm") if not c.filiales]
    else:
        pp_config = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pp") if selected_filiale in (c.filiales or [])]
        pm_config = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pm") if selected_filiale in (c.filiales or [])]
        
        if not pp_config:
            pp_config = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pp") if not c.filiales]
        if not pm_config:
            pm_config = [c for c in KycFieldVisibilityConfig.objects.filter(client_type="pm") if not c.filiales]
            
    pp_active_db_fields = set()
    if pp_config:
        pp_sources = getattr(pp_config[0], "field_sources", {}) or {}
        pp_active_db_fields.update(pp_config[0].empty_check_fields or [])
    if not pp_active_db_fields:
        pp_active_db_fields = {"CLIENT", "NUMID", "DATNAIS", "DATVALID", "PAYNAIS", "ADRESSE", "ORIGINE_REV"}

    pm_active_db_fields = set()
    if pm_config:
        pm_sources = getattr(pm_config[0], "field_sources", {}) or {}
        pm_active_db_fields.update(pm_config[0].empty_check_fields or [])
    if not pm_active_db_fields:
        pm_active_db_fields = {"CLIENT", "RCSNO", "NUMERO_FISCAL", "ADRESSE_SOCIALE"}

    def _normalize_source_map(source_map):
        """Valeurs toujours en liste pour le template (heritage str accepte)."""
        normalized = {}
        for field_name, value in (source_map or {}).items():
            if isinstance(value, str):
                vals = [s.strip() for s in value.split(",") if s.strip()]
            else:
                vals = [str(s).strip() for s in value if str(s).strip()]
            if vals:
                normalized[field_name] = vals
        return normalized

    pp_sources = _normalize_source_map(pp_sources)
    pm_sources = _normalize_source_map(pm_sources)

    filtered_pp_fields = [
        (f_name, label)
        for f_name, label in pp_fields
        if f_name in pp_active_db_fields
    ]
    filtered_pm_fields = [
        (f_name, label)
        for f_name, label in pm_fields
        if f_name in pm_active_db_fields
    ]

    document_field_source_sections = [
        {
            "title": "Particuliers (PP)",
            "client_type": "pp",
            "fields": filtered_pp_fields,
            "sources": pp_sources,
        },
        {
            "title": "Entreprises (PM)",
            "client_type": "pm",
            "fields": filtered_pm_fields,
            "sources": pm_sources,
        }
    ]

    context = {
        "extraction": extraction,
        "documents": page_obj,
        "documents_count": documents.count(),
        "uploaded_batch": uploaded_batch,
        "uploaded_documents": uploaded_documents,
        "uploaded_documents_count": uploaded_documents_count,
        "uploaded_quality_alerts": uploaded_quality_alerts,
        "uploaded_quality_alerts_count": uploaded_quality_alerts_count,
        "uploaded_batch_matching_done": uploaded_batch_matching_done,
        "uploaded_batch_running_job": uploaded_batch_running_job,
        "uploaded_batch_result_url": uploaded_batch_result_url,
        "upload_batch_queue": upload_batch_queue,
        "show_document_modal": show_document_modal,
        "kyc_pp_matches": kyc_pp_matches,
        "kyc_pp_match_summary": kyc_pp_match_summary,
        "kyc_pp_match_total_count": kyc_pp_match_total_count,
        "kyc_pp_match_filtered_count": len(kyc_pp_matches),
        "can_validate_matches": can_validate_matches,
        "can_reject_matches": can_reject_matches,
        "match_status": match_status,
        "match_status_options": [
            ("active", "A traiter"),
            ("validated", "Validees"),
            ("rejected", "Rejetees"),
            ("all", "Toutes"),
        ],
        "match_decision_return_url": request.get_full_path(),
        "expired_document_matches": expired_document_matches,
        "kyc_pp_match_filters": kyc_pp_match_filters,
        "match_filter_hidden_params": match_filter_hidden_params,
        "run_kyc_pp_matching": run_kyc_pp_matching,
        "show_match_actions": run_kyc_pp_matching and not wants_results_tab and (not selected_match_job or show_match_job_results),
        "wants_results_tab": wants_results_tab,
        "selected_match_job": selected_match_job,
        "show_match_job_results": show_match_job_results,
        "show_match_result_modal": show_match_result_modal,
        "show_match_child_modal": show_match_child_modal,
        "selected_match_job_result_url": selected_match_job_result_url,
        "selected_match_job_follow_url": selected_match_job_follow_url,
        "selected_match_job_parent_modal_url": selected_match_job_parent_modal_url,
        "all_match_results_url": all_match_results_url,
        "recent_match_jobs": recent_match_jobs,
        "match_scope_import_batch": (active_match_params or {}).get("import_batch", ""),
        "match_querystring": match_query_params.urlencode(),
        "match_reset_querystring": match_reset_querystring,
        "export_match_querystring": export_match_querystring,
        "document_type_choices": [(dt.code, dt.label) for dt in KycDocumentType.objects.all()],
        "all_document_types": KycDocumentType.objects.all(),
        "uploaded_type_distribution": uploaded_type_distribution,
        "document_type_correction_options": _document_type_correction_options(client_type),
        "document_type_filter_options": _document_type_filter_options(),
        "client_type": client_type,
        "selected_document_type": selected_document_type,
        "selected_import_batch": selected_import_batch,
        "search_fields": DOCUMENT_EXTRACTION_SEARCH_FIELDS,
        "search_field": search_field,
        "search_query": search_query,
        "field_labels": DOCUMENT_EXTRACTION_FIELD_LABELS,
        "page_querystring": query_params.urlencode(),
        "is_group_user": is_group_user,
        "liste_filiales": liste_filiales,
        "selected_filiale": selected_filiale,
        "document_field_source_sections": document_field_source_sections,
        "document_field_source_return_url": request.get_full_path(),
        "selected_match_job_consulted": selected_match_job_consulted,
        "can_insert_batches": can_insert_batches,
        "can_run_matching": can_run_matching,
        "screening_perms": screening_perms,
                                                          
        "upload_max_files": settings.DATA_UPLOAD_MAX_NUMBER_FILES,
        "single_file_max_mb": SINGLE_FILE_MAX_BYTES // (1024 * 1024),
        "zip_max_members": ZIP_MAX_MEMBERS,
        "zip_max_total_mb": ZIP_MAX_TOTAL_BYTES // (1024 * 1024),
        "matching_doc_limit": 3000,
    }
    return render(request, 'document_extraction.html', context)


@login_required
def config_document_types(request):
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "save":
            type_id = request.POST.get("type_id")
            code = request.POST.get("code", "").strip().lower()
            label = request.POST.get("label", "").strip()
            filiale = request.POST.get("filiale", "").strip()
            client_type = request.POST.get("client_type", "pp").strip()
            min_score_str = request.POST.get("min_score", "2.0")
            keywords = request.POST.get("keywords", "").strip()
            
            try:
                min_score = float(min_score_str)
            except ValueError:
                min_score = 2.0
                
            if not code or not label:
                messages.error(request, "Le code technique et le libellé sont requis.")
            else:
                if type_id:
                    doc_type = get_object_or_404(KycDocumentType, pk=type_id)
                    if doc_type.code not in ['piece_identite', 'passeport']:
                        doc_type.code = code
                        doc_type.filiale = filiale
                    doc_type.label = label
                    doc_type.client_type = client_type
                    doc_type.min_score = min_score
                    doc_type.keywords = keywords
                    doc_type.save()
                    messages.success(request, f"Le type de document '{label}' a été mis à jour.")
                else:
                    if KycDocumentType.objects.filter(code=code, filiale=filiale, client_type=client_type).exists():
                        messages.error(request, f"Un type de document avec le code '{code}' existe déjà pour cette filiale et ce type de client.")
                    else:
                        KycDocumentType.objects.create(
                            code=code,
                            label=label,
                            filiale=filiale,
                            client_type=client_type,
                            min_score=min_score,
                            keywords=keywords
                        )
                        messages.success(request, f"Le type de document '{label}' a été créé avec succès.")
            return redirect("config_document_types")
            
        elif action == "delete":
            type_id = request.POST.get("type_id")
            if type_id:
                doc_type = get_object_or_404(KycDocumentType, pk=type_id)
                if doc_type.code in ['piece_identite', 'passeport']:
                    messages.error(request, "Les types de documents système ne peuvent pas être supprimés.")
                else:
                    label = doc_type.label
                    doc_type.delete()
                    messages.success(request, f"Le type de document '{label}' a été supprimé.")
            return redirect("config_document_types")
            
    document_types = KycDocumentType.objects.all()
    context = {
        "document_types": document_types,
        "filiale_choices": Filiales,
        "client_type_choices": CLIENT_TYPE_CHOICES,
    }
    return render(request, "config_document_types.html", context)


@login_required
def import_log_download(request, filename):
    if not request.user.is_superuser:
        return redirect('accueil')
    if not filename or ".." in filename or "/" in filename or "\\" in filename:
        raise Http404("Invalid file")

    log_dir = os.path.join(settings.BASE_DIR, "logs")
    run_dir = os.path.join(log_dir, "import_runs")
    allowed_paths = [
        os.path.join(run_dir, filename),
        os.path.join(log_dir, filename),
    ]
    file_path = next((p for p in allowed_paths if os.path.isfile(p)), None)
    if not file_path:
        raise Http404("File not found")

    return FileResponse(open(file_path, "rb"), as_attachment=True, filename=filename)


                                                                               
                                                                               
                                                                            
                                                                           
                             
 
                                                                               
                                                                             
                      
PUBLIC_MEDIA_PREFIXES = ("images/",)

                                                                        
                                
EXTRACTION_MEDIA_PREFIX = "document_extraction/"


def _user_is_group_scope(user):
    """Vrai si l'utilisateur a une visibilite groupe (toutes filiales)."""
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
    filiale = getattr(user, "filiale", "") or ""
    return (filiale in ["BOA Group", "BOA GROUP"]) or (getattr(user, "organe", "") in users_groupe) or (not filiale)


def _user_can_access_extraction_file(user, relative_path):
    """Autorise l'acces a un document televersé selon le perimetre de l'utilisateur.

    Le modele KycDocumentExtraction ne porte pas de filiale : le cloisonnement
    se fait via l'agent ayant televersé le document (uploaded_by).
    """
    if user.is_superuser or _user_is_group_scope(user):
        return True

    records = KycDocumentExtraction.objects.filter(uploaded_file=relative_path).select_related("uploaded_by")
    user_filiale = (getattr(user, "filiale", "") or "").strip()
    for record in records:
        if record.uploaded_by_id == user.pk:
            return True
        uploader_filiale = (getattr(record.uploaded_by, "filiale", "") or "").strip()
        if user_filiale and uploader_filiale == user_filiale:
            return True
                                                                                   
    return False


def serve_protected_media(request, path):
    from django.contrib.auth.views import redirect_to_login

    media_root = Path(settings.MEDIA_ROOT).resolve()
                                                                                
                                                                          
    try:
        file_path = (media_root / path).resolve()
        file_path.relative_to(media_root)
    except (ValueError, OSError):
        raise Http404("Invalid path")

    if not file_path.is_file():
        raise Http404("File not found")

    relative_path = file_path.relative_to(media_root).as_posix()

    is_public = relative_path.startswith(PUBLIC_MEDIA_PREFIXES)
    if not is_public:
        if not request.user.is_authenticated:
            return redirect_to_login(request.get_full_path(), settings.LOGIN_URL)
        if relative_path.startswith(EXTRACTION_MEDIA_PREFIX):
            if not _user_can_access_extraction_file(request.user, relative_path):
                log_audit(
                    request,
                    category=AuditEvent.CAT_SECURITE,
                    action="Acces refuse a un document KYC",
                    target=relative_path,
                    details="Tentative d'acces hors perimetre a un document televersé.",
                )
                return HttpResponseForbidden("Acces non autorise")

                                                                         
                                                                               
    as_attachment = relative_path.startswith(EXTRACTION_MEDIA_PREFIX)
    response = FileResponse(open(file_path, "rb"), as_attachment=as_attachment)
    response["X-Content-Type-Options"] = "nosniff"
    if not is_public:
        response["Cache-Control"] = "private, no-store"
    return response


@login_required
def profile_update(request):
    if request.method == "POST":
        form = ProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            return redirect("profile")                          
    else:
        form = ProfileForm(instance=request.user)
    return render(request, "profile_update.html", {"form": form})


def rechercher_et_noter_agent(request):
    agent = None
    form = None
    message = None

    if request.method == "POST":
        expl = request.POST.get('expl', '')
        if expl:
            try:
                agent = ProfileV.objects.get(code_expl=expl)
            except ProfileV.DoesNotExist:
                message = "Agent introuvable."

            if agent:
                                                                   
                if request.user.is_authenticated and request.user.groups.filter(name='Contrôle permanent').exists():
                    form = NoteForm(request.POST or None)
                    if form.is_valid():
                        note = form.save(commit=False)
                        note.agent = agent
                        note.date_notation = timezone.now()                                   
                        note.save()
                        message = "Notation enregistrée avec succès."
                        form = None                                                                               
                else:
                    message = "Vous n'avez pas la permission de noter cet agent."

    return render(request, 'accueil.html', {'agent': agent, 'form': form, 'message': message})


def password_reset_request(request):
    if request.method == "POST":
        password_reset_form = PasswordResetForm(request.POST)
        if password_reset_form.is_valid():
            data = password_reset_form.cleaned_data['email']
            associated_users = ProfileV.objects.filter(Q(username=data))
            if associated_users.exists():
                for user in associated_users:
                    subject = "Password Reset Requested"
                    email_template_name = "password_reset_email.txt"
                    c = {
                        "email": user.username,
                        'domain': request.get_host(),
                        'site_name': 'Plateforme KYC BOA',
                        "uid": urlsafe_base64_encode(force_bytes(user.pk)),
                        'token': default_token_generator.make_token(user),
                        'protocol': 'https' if request.is_secure() else 'http',
                    }
                    email = render_to_string(email_template_name, c)
                    try:
                        send_mail(subject, email, 'mamadou@mamadou.sn', [user.username], fail_silently=False)
                    except BadHeaderError:
                        return HttpResponse('Invalid header found.')
                    return redirect('/password_resete/done')

                    messages.success(request, 'A message with reset password instructions has been sent to your inbox.')
                    return redirect("accueil")
            else:
                errors = 'Votre mail ne figure pas dans notre base.'
                return render(request=request, template_name="password_reset.html",
                              context={"password_reset_form": password_reset_form, "errors": errors})

    password_reset_form = PasswordResetForm()
    return render(request=request, template_name="password_reset.html",
                  context={"password_reset_form": password_reset_form})


@login_required
def profil(request):
    roles_exclus = ["Chargé Client"]
    notes = Notation.objects.filter(flux_stock='Flux')
    user = request.user
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))

                                                                 
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    notation = notation.filter(agent__filiale=user.filiale, agent__code_expl=user.code_expl)

                                                              
    derniere_notation = (Notation.objects
                         .filter(agent=user)
                         .order_by('-date_notation')
                         .first())

    context = {'roles_exclus': roles_exclus,
               'notation': notation,
               'derniere_notation': derniere_notation,
               }
    return render(request, 'profil.html', context)


@login_required
def profile(request):
    if request.method == 'POST':
        user_form = ProfileModify(request.POST, instance=request.user)
        if user_form.is_valid():
            user_form.save()
            messages.success(request, 'Votre profil a été modifié avec succès')
            return redirect('/perso/profil')

    else:
        user_form = ProfileModify(instance=request.user)
    return render(request, 'modify_profil.html', {'user_form': user_form})


@method_decorator(login_required, name='dispatch')
class ChangePasswordView(SuccessMessageMixin, PasswordChangeView):
    template_name = 'modify_pw.html'
    success_message = "Votre mot de passe a été changé avec succès"
    success_url = reverse_lazy('profil')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
                                        
        context['roles_exclus'] = ["Chargé Client"]
        return context


@login_required
                                                                               
                                                                              
                                                                                
                                                                   
                                                                        


@login_required
def perso(request):
                                      
    roles_exclus = ["Chargé Client"]
    user = request.user

                                                                                           
    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Risques', 'DAI', 'Qualité','DSI']:
        agents = ProfileV.objects.filter(filiale=user.filiale)
    else:
        agents = ProfileV.objects.all()
                                            
    query = request.GET.get('q', '')
    if query:
        agents = agents.filter(code_expl__icontains=query)

    return render(request, 'mon_profile.html', {'agents': agents, 'query': query, 'roles_exclus': roles_exclus})


@login_required
def agent(request):
    roles_exclus = ["Chargé Client"]
    user = request.user

                                                                                           
    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Risques', 'DAI', 'Qualité','DSI']:
        notes = Notation.objects.filter(note_par__filiale=user.filiale, flux_stock='Flux')
    elif user.organe in ['Directeur Agence']:
        notes = Notation.objects.filter(note_par__filiale=user.filiale, agent__agence=user.agence, flux_stock='Flux')
    else:
        notes = Notation.objects.filter(flux_stock='Flux')

                                                             
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))

                                                                 
    notes = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    notes = notes.order_by('-date_notation')
                                            
    query = request.GET.get('q', '')
    if query:
        agents = ProfileV.objects.filter(code_expl__icontains=query)                                                         
        if agents.exists():
            notes = notes.all().filter(agent__in=agents)
        else:
                                                                                  
            notes = notes.none()

    return render(request, 'agent.html', {
        'notes': notes,
        'query': query,
        'roles_exclus': roles_exclus,
    })


@login_required
def export_agents_excel(request):
    def strip_tz(value):
        if hasattr(value, 'tzinfo'):
            return value.replace(tzinfo=None)
        return value

    user = request.user

    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Risques', 'DAI', 'Qualité','DSI']:
        donnees = Notation.objects.filter(note_par__filiale=user.filiale, flux_stock='Flux')
    elif user.organe in ['Directeur Agence']:
        donnees = Notation.objects.filter(note_par__filiale=user.filiale, agent__agence=user.agence, flux_stock='Flux')
    else:
        donnees = Notation.objects.filter(flux_stock='Flux')

    latest_notes = donnees.values('agent').annotate(latest_date=Max('date_notation'))

                                                                 
    donnees = donnees.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    donnees = donnees.order_by('-date_notation')

                                
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notation_Flux"

              
    headers = ["FILIALE", "EXPLOITANT", "NOM", "AGENCE", "EMAIL", "NOTE", "Dernière notation", "Noté par le contrôleur",
               "Flux/Stock"]
    ws.append(headers)

             
    for d in donnees:
        agent_nom = f"{d.agent.first_name} {d.agent.last_name}".strip()
        agent_email = d.agent.email or d.agent.username
        ws.append([
            d.agent.filiale, d.agent.code_expl, agent_nom, d.agent.agence, agent_email, d.note,
            strip_tz(d.date_notation), d.note_par.email, d.flux_stock

        ])

                                                 
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

                              
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Notation_Flux_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_agents_excel_s(request):
    def strip_tz(value):
        if hasattr(value, 'tzinfo'):
            return value.replace(tzinfo=None)
        return value

    user = request.user

    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Risques', 'DAI', 'Qualité','DSI']:
        donnees = Notation.objects.filter(note_par__filiale=user.filiale, flux_stock='Stock')
    elif user.organe in ['Directeur Agence']:
        donnees = Notation.objects.filter(note_par__filiale=user.filiale, agent__agence=user.agence, flux_stock='Stock')
    else:
        donnees = Notation.objects.filter(flux_stock='Stock')

    latest_notes = donnees.values('agent').annotate(latest_date=Max('date_notation'))

                                                                 
    donnees = donnees.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    donnees = donnees.order_by('-date_notation')

                                
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notation_Stock"

              
    headers = ["FILIALE", "EXPLOITANT", "NOM", "AGENCE", "EMAIL", "NOTE", "Dernière notation", "Noté par le contrôleur",
               "Flux/Stock"]
    ws.append(headers)

             
    for d in donnees:
        agent_nom = f"{d.agent.first_name} {d.agent.last_name}".strip()
        agent_email = d.agent.email or d.agent.username
        ws.append([
            d.agent.filiale, d.agent.code_expl, agent_nom, d.agent.agence, agent_email, d.note,
            strip_tz(d.date_notation), d.note_par.email, d.flux_stock

        ])

                                                 
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

                              
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Notation_Stock_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def perso_stock(request):
                                      
    user = request.user

                                                                                           
    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Risques', 'DAI', 'Qualité','DSI']:
        notes = Notation.objects.filter(filiale=user.filiale, flux_stock='Stock')
    else:
        notes = Notation.objects.all().filter(flux_stock='Flux')
                                            
    query = request.GET.get('q', '')
    if query:
        notes = notes.filter(agent__code_expl__icontains=query)

    return render(request, 'agent_stock.html', {'notes': notes, 'query': query})


@login_required
def agent_stock(request):
    user = request.user

                                                                                           
    if user.organe in ['Contrôle Permanent', 'Directeur Réseau', 'Conformité','Risques', 'DAI', 'Qualité','DSI']:
        notes = Notation.objects.filter(note_par__filiale=user.filiale, flux_stock='Stock')
    else:
        notes = Notation.objects.filter(flux_stock='Stock')

                                                             
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))
    notes = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])

                                                                 
    notes = notes.order_by('-date_notation')

                                            
    query = request.GET.get('q', '')
    if query:
        agents = ProfileV.objects.filter(code_expl__icontains=query)                                                         
        if agents.exists():
            notes = notes.filter(agent__in=agents)
        else:
                                                                                  
            notes = notes.none()

    return render(request, 'agent_stock.html', {'notes': notes, 'query': query})


@login_required
def notes(request):
    agent = None
    roles_exclus = ["Chargé Client", "Directeur Agence"]
    form = NotationForm()                                

    if request.method == 'POST':
        if 'search_agent' in request.POST:
                                                        
            code_exploitant = request.POST.get('code_exploitant')
            try:
                agent = ProfileV.objects.get(code_expl=code_exploitant, filiale=request.user.filiale)

                                                              
                form = NotationForm(initial={'agent': agent})
            except ProfileV.DoesNotExist:
                agent = None
                error_message = "L'agent avec ce code exploitant n'existe pas."
                return render(request, 'notation.html', {'form': form, 'error_message': error_message})

        else:
                                                  
            form = NotationForm(request.POST)
            if form.is_valid():
                user = request.user
                                                          
                notation = form.save(commit=False)
                notation.filiale = request.user.filiale
                notation.note_par = request.user
                notation.date_notation = timezone.now()
                notation.save()
                messages.success(request, 'La notation a bien été sauvegardée.')

                return redirect('agent')
    else:
        form = NotationForm()                                                                

    return render(request, 'notation.html', {'form': form, 'agent': agent, 'roles_exclus': roles_exclus})


@login_required
def agent_detail(request, agent_id):
    agent = get_object_or_404(ProfileV, id=agent_id)
    notations = agent.notations.all().order_by('-date_notation')
    return render(request, 'agent_detail.html', {'agent': agent, 'notations': notations})


@login_required
def historique(request):
    roles_exclus = ["Chargé Client", "Directeur Agence"]
    query = request.GET.get('q')

    if query:
                                                             
        notations = Notation.objects.filter(note_par=request.user, agent__code_expl__icontains=query).order_by(
            "-date_notation")
    else:
                                                                  
        notations = Notation.objects.filter(note_par=request.user).order_by("-date_notation")

                                     
    context = {
        'notations': notations,
        'query': query,
    }                                          
    return render(request, 'historique.html', {'notations': notations, 'roles_exclus': roles_exclus})


@login_required
def test(request):
    return render(request, 'test.html')


@login_required
def register(request):
    roles_exclus = ["Chargé Client"]
    current_user = request.user

                                           
    if current_user.organe not in ["PASS", "DSI"]:
        messages.error(request, "Vous nâ€™avez pas la permission de créer un compte utilisateur.")
        return redirect('user_list')

    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            new_user = form.save(commit=False)

                                                                              
            if current_user.organe == "DSI":
                new_user.filiale = current_user.filiale

            new_user.save()
            messages.success(request, "Utilisateur créé avec succès.")
            return redirect('user_list')
    else:
        form = CustomUserCreationForm(current_user=current_user)                                                        

    return render(request, 'register.html', {'form': form, 'roles_exclus': roles_exclus})


                                                                       
def is_pass_user(user):
    return user.organe == 'PASS'


                                             


@login_required
def user_list(request):
    user = request.user
    query = request.GET.get('q', '')
    page_number = request.GET.get('page')

                                   
    if user.organe == "PASS":
        users_base = ProfileV.objects.all().order_by('last_name')
    elif user.organe == "DSI":
        users_base = ProfileV.objects.filter(filiale=user.filiale).order_by('last_name')
    else:
        users_base = ProfileV.objects.none()

                                  
    if query:
        users_base = users_base.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(username__icontains=query)
        )

                                                      
    active_sessions = Session.objects.filter(expire_date__gte=now())
    user_ids = []
    for s in active_sessions:
        data = s.get_decoded()
        if '_auth_user_id' in data:
            user_ids.append(data['_auth_user_id'])

    connected_users = users_base.filter(id__in=list(set(user_ids)))

                                 
    paginator = Paginator(users_base, 10)
    page_obj = paginator.get_page(page_number)

    context = {
        'total_users': users_base.count(),
        'connected_count': connected_users.count(),
        'connected_users': connected_users,
        'page_obj': page_obj,
        'query': query,
    }
    return render(request, 'user_list.html', context)



@login_required
def edit_user(request, user_id):
    current_user = request.user
    target_user = get_object_or_404(ProfileV, pk=user_id)

                             
    if current_user.organe != "PASS":
        if current_user.organe == "DSI":
            if current_user.filiale != target_user.filiale:
                messages.error(request, "Vous ne pouvez modifier que les utilisateurs de votre filiale.")
                return redirect('user_list')
        else:
            messages.error(request, "Vous nâ€™avez pas la permission de modifier cet utilisateur.")
            return redirect('user_list')

    if request.method == "POST":
        form = UserEditForm(request.POST, instance=target_user, current_user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Utilisateur modifié avec succès.")
            return redirect('user_list')
    else:
        form = UserEditForm(instance=target_user, current_user=request.user)

    return render(request, 'edit_user.html', {'form': form, 'user': target_user})


@login_required
def change_user_password(request, user_id):
    current_user = request.user
    target_user = get_object_or_404(ProfileV, pk=user_id)

                               
                                                                    
                                                                                    
                              
    if current_user.organe != "PASS":
        if current_user.organe == "DSI":
            if current_user.filiale != target_user.filiale:
                messages.error(request, "Vous ne pouvez changer le mot de passe que des utilisateurs de votre filiale.")
                return redirect('user_list')
        else:
            messages.error(request, "Vous nâ€™avez pas la permission de changer ce mot de passe.")
            return redirect('user_list')

                                   
    if request.method == 'POST':
        form = PasswordChangeForm(target_user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)                         
            messages.success(request, "Le mot de passe a été modifié avec succès.")
            return redirect('user_list')
    else:
        form = PasswordChangeForm(target_user)

    return render(request, 'change_user_password.html', {'form': form, 'user': target_user})


@login_required
def reset_user_password(request, user_id):
    current_user = request.user
    target_user = get_object_or_404(ProfileV, pk=user_id)

                               
    if current_user.organe != "PASS":
        if current_user.organe == "DSI":
            if current_user.filiale != target_user.filiale:
                messages.error(request,
                               "Vous ne pouvez réinitialiser que les mots de passe des utilisateurs de votre filiale.")
                return redirect('user_list')
        else:
            messages.error(request, "Vous nâ€™avez pas la permission de réinitialiser ce mot de passe.")
            return redirect('user_list')

                                   
    if request.method == 'POST':
        form = ResetPasswordForm(request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['new_password']
            target_user.password = make_password(new_password)
            target_user.force_password_change = form.cleaned_data.get('force_password_change', False)
            target_user.save()
            log_audit(
                request,
                category=AuditEvent.CAT_SECURITE,
                action="Reinitialisation de mot de passe",
                target=target_user.username,
                details=(f"Compte {target_user.username} ({target_user.organe or '-'} / "
                         f"{target_user.filiale or '-'}) | Changement force a la prochaine connexion : "
                         f"{'oui' if target_user.force_password_change else 'non'}"),
            )
            messages.success(request, "Le mot de passe a été réinitialisé avec succès.")
            return redirect('user_list')
    else:
        form = ResetPasswordForm(initial={
            'force_password_change': target_user.force_password_change
        })

    return render(request, 'reset_user_password.html', {'form': form, 'target_user': target_user})


@login_required
def user_statistics_view(request):
    roles_exclus = ["Chargé Client"]
    current_user = request.user                        

                                              
    if current_user.organe == "PASS":
        users = ProfileV.objects.all()

    elif current_user.organe == "DSI":
        users = ProfileV.objects.filter(filiale=current_user.filiale)

    else:
        messages.error(request, "Vous nâ€™avez pas la permission dâ€™accéder Ã  cette page.")
        return render(request, 'user_statistics.html', {
            'total_users': 0,
            'connected_count': 0,
            'connected_users': [],
            'users': [],
            'roles_exclus': roles_exclus,
            'connection_history_labels': json.dumps([]),
            'connection_history_values': json.dumps([]),
            'connection_history_rows': [],
            'history_days': 0,
        })

                                          
    total_users = users.count()

                           
    active_sessions = Session.objects.filter(expire_date__gte=now())
    user_ids = []
    for session in active_sessions:
        data = session.get_decoded()
        if '_auth_user_id' in data:
            user_ids.append(data['_auth_user_id'])

                            
    user_ids = list(set(user_ids))

                                                               
    connected_users = users.filter(id__in=user_ids)
    connected_count = connected_users.count()

    visible_login_events = UserLoginHistory.objects.filter(user_id__in=users.values("id"))
    first_login_at = visible_login_events.order_by("login_at").values_list("login_at", flat=True).first()

    end_date = timezone.localdate()
    if first_login_at:
        if timezone.is_aware(first_login_at):
            start_date = timezone.localtime(first_login_at).date()
        else:
            start_date = first_login_at.date()
    else:
        start_date = end_date - timedelta(days=6)

    daily_connections = (
        visible_login_events
        .filter(login_at__date__range=(start_date, end_date))
        .annotate(day=TruncDate("login_at"))
        .values("day")
        .annotate(count=Count("user_id", distinct=True))
        .order_by("day")
    )
    daily_connections_map = {row["day"]: row["count"] for row in daily_connections}

    cursor = start_date
    chart_labels = []
    chart_values = []
    history_rows = []

    while cursor <= end_date:
        count = daily_connections_map.get(cursor, 0)
        chart_labels.append(cursor.strftime("%d/%m"))
        chart_values.append(count)
        history_rows.append(
            {
                "date": cursor.strftime("%d/%m/%Y"),
                "count": count,
            }
        )
        cursor += timedelta(days=1)

                                     
    context = {
        'total_users': total_users,
        'connected_count': connected_count,
        'connected_users': connected_users,
        'users': users,
        'roles_exclus': roles_exclus,
        'connection_history_labels': json.dumps(chart_labels),
        'connection_history_values': json.dumps(chart_values),
        'connection_history_rows': list(reversed(history_rows)),
        'history_days': len(history_rows),
    }

    return render(request, 'user_statistics.html', context)


@login_required
def ppe(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                    "Contrôle Permanent Groupe", "PASS", "GUEST"]
    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')
    filiale_txt = (request.GET.get('col_filiale') or request.GET.get('filiale_txt', '')).strip()
    agence_txt = (request.GET.get('col_agence') or request.GET.get('agence_txt', '')).strip()
    lib_agence = (request.GET.get('col_lib_agence') or request.GET.get('lib_agence', '')).strip()
    expl_txt = (request.GET.get('col_expl') or request.GET.get('expl_txt', '')).strip()
    client_txt = (request.GET.get('col_client') or request.GET.get('client', '')).strip()
    risque_txt = (request.GET.get('col_risque') or request.GET.get('risque', '')).strip()

                                                                          
    donnees = Kyc_pp.objects.filter(PPE="O")

                                        
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    if user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    if user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    if user.organe in users_groupe:
        donnees = donnees

                             
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)
    if filiale_txt:
        donnees = donnees.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        donnees = donnees.filter(AGENCE__icontains=agence_txt)
    if expl_txt:
        donnees = donnees.filter(EXPL__icontains=expl_txt)
    if client_txt:
        donnees = donnees.filter(CLIENT__icontains=client_txt)
    if risque_txt:
        donnees = donnees.filter(RISQUE__icontains=risque_txt)
    if lib_agence:
        donnees = donnees.filter(LIB_AGENCE__icontains=lib_agence)
    donnees = apply_datouv_period_filter(donnees, request)

                                                 
    if user.organe == "Directeur Agence":
        exploitants = donnees.filter(AGENCE=user.agence).values_list('EXPL', flat=True).distinct()
        agences = donnees.filter(AGENCE=user.agence).values_list('AGENCE', flat=True).distinct()
    elif user.organe in users_filiale:
        agences = donnees.filter(FILIALE=user.filiale).values_list('AGENCE', flat=True).distinct()
        exploitants = donnees.filter(FILIALE=user.filiale).values_list('EXPL', flat=True).distinct()
    else:
        exploitants = donnees.values_list('EXPL', flat=True).distinct()
        agences = donnees.values_list('AGENCE', flat=True).distinct()

    filiales = donnees.values_list('FILIALE', flat=True).distinct()

                                                    
    ITEMS_PER_PAGE = 25
    paginator = Paginator(donnees.order_by('id'), ITEMS_PER_PAGE)
    page_number = request.GET.get('page')
    try:
        objets_page = paginator.page(page_number)
    except PageNotAnInteger:
        objets_page = paginator.page(1)
    except EmptyPage:
        objets_page = paginator.page(paginator.num_pages)

    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']

              
    total_ppe = paginator.count
    missing_ppe = get_incomplete_clients_queryset(donnees, 'pp').count()
    complete_ppe = max(0, total_ppe - missing_ppe)
    compliance_rate = round((complete_ppe / total_ppe) * 100, 1) if total_ppe > 0 else 100.0

                                      
    risque_counts = list(
        donnees.values('RISQUE')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    risque_items = []
    for rc in risque_counts:
        label = rc['RISQUE'] or "Non renseigné"
        pct = round((rc['count'] / total_ppe) * 100, 1) if total_ppe > 0 else 0.0
        risque_items.append({
            'label': label,
            'value': rc['count'],
            'suffix': f'({pct}%)'
        })

    export_params = request.GET.copy()
    export_params['incompletes'] = '1'
    export_incomplets_url = f"{reverse('export_ppe')}?{export_params.urlencode()}"

    kpi_cards = [
        {
            'tone': 'emerald',
            'label': 'Répartition par Risque',
            'value': total_ppe,
            'subtitle': 'Clients PPE par classe de risque',
            'show_modal': True,
            'items': risque_items
        },
        {
            'tone': 'red',
            'label': 'PPE Incomplets',
            'value': missing_ppe,
            'subtitle': 'Dossiers avec données manquantes',
            'export_url': export_incomplets_url
        },
        {
            'tone': 'blue',
            'label': 'Taux de conformité',
            'value': compliance_rate,
            'suffix': '%',
            'subtitle': 'Dossiers complets / total',
        }
    ]

    context = {
        'donnees': objets_page,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
        'get_params': query_params.urlencode(),
        'kpi_cards': kpi_cards,
    }

    return render(request, 'ppe.html', context)



@login_required
def export_ppe(request):
    user = request.user
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

                               
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')
    filiale_txt = request.GET.get('filiale_txt', '').strip()
    agence_txt = request.GET.get('agence_txt', '').strip()
    lib_agence = request.GET.get('lib_agence', '').strip()
    expl_txt = request.GET.get('expl_txt', '').strip()
    client_txt = request.GET.get('client', '').strip()
    risque_txt = request.GET.get('risque', '').strip()

                               
    donnees = Kyc_pp.objects.filter(PPE="O")

    incompletes_only = request.GET.get('incompletes', '') == '1'
    if incompletes_only:
        donnees = get_incomplete_clients_queryset(donnees, 'pp')

                                        
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
                                                                                            

                                                    
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)
    donnees = apply_datouv_period_filter(donnees, request)

                                
    wb = Workbook()
    ws = wb.active
    ws.title = "Non rens PPE"

    headers = [
        "FILIALE", "AGENCE", "EXPL", "CLIENT", "CODAPE", "IDP",
        "PAYNAIS", "PROFESSION", "ADRESSE", "PAYS_RESID", "NUMID",
        "SALAIRE", "ORIGINE_REV", "DATVALID", "TEL"
    ]
    ws.append(headers)

    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.CODAPE, d.IDP,
            d.PAYNAIS, d.PROFESSION, d.ADRESSE, d.PAYS_RESID,
            d.NUMID, d.SALAIRE, d.ORIGINE_REV, d.DATVALID, d.TEL
        ])

    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    if incompletes_only:
        filename = f"PPE_incomplets_{date_str}.xlsx"
    else:
        filename = f"PPE_non_rens_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def _apply_pp_header_filters(queryset, request, include_pays_resid=False, include_devise=False):
    field_map = {
        "lib_agence": "LIB_AGENCE",
        "client": "CLIENT",
        "idp": "IDP",
        "numid": "NUMID",
        "datnais": "DATNAIS",
        "paynais": "PAYNAIS",
        "adresse": "ADRESSE",
        "codape": "CODAPE",
        "profession": "PROFESSION",
        "salaire": "SALAIRE",
        "origine_rev": "ORIGINE_REV",
        "datvalid": "DATVALID",
        "tel": "TEL",
        "datouv": "DATOUV",
        "agence": "AGENCE",
        "expl": "EXPL",
    }
    if include_pays_resid:
        field_map["pays_resid"] = "PAYS_RESID"
    if include_devise:
        field_map["devise"] = "DEVISE"

    for param, field in field_map.items():
        value = (request.GET.get(f"col_{param}") or request.GET.get(param, "")).strip()
        if value:
            queryset = queryset.filter(**{f"{field}__icontains": value})
    return queryset


@login_required
def non_resid(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]
    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

                                          
                                                                             
                                                                                                 
    donnees = Kyc_pp.objects.filter(RESID="N")
                                                                               
                                                        

                                                
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    elif user.organe in users_groupe:
                                                                                                 
        pass

                                     
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)
    donnees = _apply_pp_header_filters(donnees, request, include_pays_resid=True)
    donnees = apply_datouv_period_filter(donnees, request)

                                                                  
                                                              

                                                                                  
    exploitants = donnees.values_list('EXPL', flat=True).distinct()
    agences = donnees.values_list('AGENCE', flat=True).distinct()

                                                                                 
    if user.organe == "Directeur Agence":
                                                                                                       
        agences = donnees.values_list('AGENCE', flat=True).distinct()
        exploitants = donnees.values_list('EXPL', flat=True).distinct()
    elif user.organe in users_filiale:
                                                                                        
        agences = donnees.values_list('AGENCE', flat=True).distinct()
        exploitants = donnees.values_list('EXPL', flat=True).distinct()

    filiales = donnees.values_list('FILIALE', flat=True).distinct()
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']

                                                     
                                                     
    queryset = donnees.order_by('id')

                               
    ITEMS_PER_PAGE = 25

                               
    paginator = Paginator(queryset, ITEMS_PER_PAGE)
    page_number = request.GET.get('page')

    try:
        objets_page = paginator.page(page_number)
    except PageNotAnInteger:
                                                                   
        objets_page = paginator.page(1)
    except EmptyPage:
                                                                 
        objets_page = paginator.page(paginator.num_pages)

    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']

                          
    total_non_resid = donnees.count()
    missing_non_resid = get_incomplete_clients_queryset(donnees, 'pp').count()
    complete_non_resid = max(0, total_non_resid - missing_non_resid)
    compliance_rate = round((complete_non_resid / total_non_resid) * 100, 1) if total_non_resid > 0 else 100.0

                          
    country_counts = list(
        donnees.values('PAYS_RESID')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    country_items = []
    for cc in country_counts:
        country_code = cc['PAYS_RESID'] or "Non renseigné"
        pct = round((cc['count'] / total_non_resid) * 100, 1) if total_non_resid > 0 else 0.0
        country_items.append({
            'label': country_code,
            'value': cc['count'],
            'suffix': f'({pct}%)'
        })

    export_params = request.GET.copy()
    export_params['incompletes'] = '1'
    export_incomplets_url = f"{reverse('export_non_resid_pp')}?{export_params.urlencode()}"

    kpi_cards = [
        {
            'tone': 'emerald',
            'label': 'Répartition par Pays',
            'value': total_non_resid,
            'subtitle': 'Clients non résidents PP',
            'show_modal': True,
            'items': country_items
        },
        {
            'tone': 'red',
            'label': 'PP Incomplets',
            'value': missing_non_resid,
            'subtitle': 'Dossiers avec données manquantes',
            'export_url': export_incomplets_url,
        },
        {
            'tone': 'blue',
            'label': 'Taux de complétude',
            'value': compliance_rate,
            'suffix': '%',
            'subtitle': 'Dossiers complets / total',
        }
    ]

    context = {
                                                      
        "donnees": objets_page,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
        'get_params': query_params.urlencode(),
        'kpi_cards': kpi_cards,
    }
    return render(request, 'non_resid.html', context)

@login_required
def export_non_resid_pp(request):
        user = request.user

        users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
        users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                        "Contrôle Permanent Groupe", "PASS", "GUEST"]

                                                              
        filiale_param = request.GET.get('filiale', '')
        agence_param = request.GET.get('agence', '')
        expl_param = request.GET.get('expl', '')

                           
        donnees = Kyc_pp.objects.filter(RESID="N")


                                                                          
        if user.organe == "Chargé Client":
            donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
        elif user.organe == "Directeur Agence":
            donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
        elif user.organe in users_filiale:
            donnees = donnees.filter(FILIALE=user.filiale)
        elif user.organe in users_groupe:
            pass                              

                                                           
        if filiale_param:
            donnees = donnees.filter(FILIALE__icontains=filiale_param)
        if agence_param:
            donnees = donnees.filter(AGENCE__icontains=agence_param)
        if expl_param:
            donnees = donnees.filter(EXPL__icontains=expl_param)
        donnees = _apply_pp_header_filters(donnees, request, include_pays_resid=True)
        donnees = apply_datouv_period_filter(donnees, request)

        if request.GET.get('incompletes') == '1':
            donnees = get_incomplete_clients_queryset(donnees, 'pp')

                                

                                    
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Comptes Devise PP"                         

                  
        headers = ["FILIALE", "AGENCE", "EXPL", "CLIENT", "CODAPE", "IDP", "PAYNAIS", "PROFESSION", "ADRESSE", "PAYS_RESID",
                   "NUMID", "SALAIRE", "ORIGINE_REV", "DATVALID", "TEL"]
        ws.append(headers)

                 
        for d in donnees:
            ws.append([
                d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.CODAPE, d.IDP,
                d.PAYNAIS, d.PROFESSION, d.ADRESSE, d.PAYS_RESID, d.NUMID, d.SALAIRE, d.ORIGINE_REV, d.DATVALID, d.TEL
            ])

                                                     
        for col_num, column_title in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 15

                                  
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(output.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
        if request.GET.get('incompletes') == '1':
            filename = f"Comptes_non_resid_PP_incomplets_{date_str}.xlsx"
        else:
            filename = f"Comptes_non_resid_PP_{date_str}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


@login_required

def non_resid_pm(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]
    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

                                          
                                                                             
                                                                                                 
    donnees = Kyc_pm.objects.filter(RESID__exact="N")
                                                                               
                                                        

                                                
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    elif user.organe in users_groupe:
                                                                                                 
        pass

                                     
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)
    donnees = apply_datouv_period_filter(donnees, request)

                                                                  
                                                              

                                                                                  
    exploitants = donnees.values_list('EXPL', flat=True).distinct()
    agences = donnees.values_list('AGENCE', flat=True).distinct()

                                                                                 
    if user.organe == "Directeur Agence":
                                                                                                       
        agences = donnees.values_list('AGENCE', flat=True).distinct()
        exploitants = donnees.values_list('EXPL', flat=True).distinct()
    elif user.organe in users_filiale:
                                                                                        
        agences = donnees.values_list('AGENCE', flat=True).distinct()
        exploitants = donnees.values_list('EXPL', flat=True).distinct()

    filiales = donnees.values_list('FILIALE', flat=True).distinct()

                                                     
                                                     
    queryset = donnees.order_by('id')

                               
    ITEMS_PER_PAGE = 25

                               
    paginator = Paginator(queryset, ITEMS_PER_PAGE)
    page_number = request.GET.get('page')
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']

    try:
        objets_page = paginator.page(page_number)
    except PageNotAnInteger:
                                                                   
        objets_page = paginator.page(1)
    except EmptyPage:
                                                                 
        objets_page = paginator.page(paginator.num_pages)

                          
    total_non_resid = donnees.count()
    missing_non_resid = get_incomplete_clients_queryset(donnees, 'pm').count()
    complete_non_resid = max(0, total_non_resid - missing_non_resid)
    compliance_rate = round((complete_non_resid / total_non_resid) * 100, 1) if total_non_resid > 0 else 100.0

                          
    country_counts = list(
        donnees.values('PAYS_JUR')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    country_items = []
    for cc in country_counts:
        country_code = cc['PAYS_JUR'] or "Non renseigné"
        pct = round((cc['count'] / total_non_resid) * 100, 1) if total_non_resid > 0 else 0.0
        country_items.append({
            'label': country_code,
            'value': cc['count'],
            'suffix': f'({pct}%)'
        })

    export_params = request.GET.copy()
    export_params['incompletes'] = '1'
    export_incomplets_url = f"{reverse('export_non_resid_pm')}?{export_params.urlencode()}"

    kpi_cards = [
        {
            'tone': 'emerald',
            'label': 'Répartition par Pays',
            'value': total_non_resid,
            'subtitle': 'Clients non résidents PM',
            'show_modal': True,
            'items': country_items
        },
        {
            'tone': 'red',
            'label': 'PM Incomplets',
            'value': missing_non_resid,
            'subtitle': 'Dossiers avec données manquantes',
            'export_url': export_incomplets_url,
        },
        {
            'tone': 'blue',
            'label': 'Taux de complétude',
            'value': compliance_rate,
            'suffix': '%',
            'subtitle': 'Dossiers complets / total',
        }
    ]

    context = {
                                                      
        "donnees": objets_page,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
        'get_params': query_params.urlencode(),
        'kpi_cards': kpi_cards,
    }
    return render(request, 'non_resid_pm.html', context)


@login_required
def export_non_resid_pm(request):
    user = request.user

    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                    "Contrôle Permanent Groupe", "PASS", "GUEST"]

                                                          
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

                       
    donnees = Kyc_pm.objects.filter(RESID="N")

                                                                      
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    elif user.organe in users_groupe:
        pass                              

                                                       
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)
    donnees = apply_datouv_period_filter(donnees, request)

    if request.GET.get('incompletes') == '1':
        donnees = get_incomplete_clients_queryset(donnees, 'pm')

                            

                                
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comptes Devise PP"                         

              
    headers = ["FILIALE", "AGENCE", "EXPL", "CLIENT", "AGEC", "CODAPE", "IDM", "RCSNO", "CAPITAL", "CA",
               "RESULTAT", "TEL"]
    ws.append(headers)

             
    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.AGEC, d.CODAPE, d.IDM,
            d.RCSNO, d.CAPITAL, d.CA, d.RESULTAT, d.TEL
        ])

                                                 
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

                              
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    if request.GET.get('incompletes') == '1':
        filename = f"Comptes_non_resid_PM_incomplets_{date_str}.xlsx"
    else:
        filename = f"Comptes_non_resid_PM_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def scoring(request):
                             
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau", 'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                    "Contrôle Permanent Groupe", "PASS", "GUEST"]

    user = request.user
    today = date.today()

                                     
    periode_param = request.GET.get("periode", "")
    filiale_param = request.GET.get("filiale", "")
    agence_param = request.GET.get("agence", "")
    expl_param = request.GET.get("expl", "")
    filiale_txt = request.GET.get("filiale_txt", "").strip()
    agence_txt = request.GET.get("agence_txt", "").strip()
    lib_agence = request.GET.get("lib_agence", "").strip()
    expl_txt = request.GET.get("expl_txt", "").strip()
    client_txt = request.GET.get("client", "").strip()
    daterev_txt = request.GET.get("daterev", "").strip()

                                                                   
    qs = DATEREV.objects.all()

                                     
    organe = getattr(user, "organe", "")
    is_group_user = (organe in users_groupe) or (user.filiale in ["BOA Group", "BOA GROUP"]) or (not user.filiale)
    if not is_group_user:
        if organe == "Chargé Client":
            qs = qs.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
        elif organe == "Directeur Agence":
            qs = qs.filter(FILIALE=user.filiale, AGENCE=user.agence)
        elif organe in users_filiale:
            qs = qs.filter(FILIALE=user.filiale)
                                                    

                             
    if periode_param == "today":
        qs = qs.filter(DATEREV__lte=today)
    elif periode_param == "3m":
        qs = qs.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=90))
    elif periode_param == "6m":
        qs = qs.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=180))
    elif periode_param == "1y":
        qs = qs.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=365))

                                          
    if filiale_param:
        qs = qs.filter(FILIALE=filiale_param)

    if agence_param:
        qs = qs.filter(AGENCE=agence_param)

    if expl_param:
        qs = qs.filter(EXPL=expl_param)
    if filiale_txt:
        qs = qs.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        qs = qs.filter(AGENCE__icontains=agence_txt)
    if lib_agence:
        qs = qs.filter(LIB_AGENCE__icontains=lib_agence)
    if expl_txt:
        qs = qs.filter(EXPL__icontains=expl_txt)
    if client_txt:
        qs = qs.filter(CLIENT__icontains=client_txt)
    if daterev_txt:
        daterev_txt = daterev_txt.strip()
        parsed = parse_date(daterev_txt)
        if not parsed:
            for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                try:
                    parsed = datetime.strptime(daterev_txt, fmt).date()
                    break
                except (ValueError, TypeError):
                    continue
        if parsed:
            qs = qs.filter(DATEREV=parsed)

                                                                                           
                                                                                    
    filiales_opts = DATEREV.objects.values_list("FILIALE", flat=True).distinct().order_by("FILIALE")
    agences_opts = qs.values_list("AGENCE", flat=True).distinct().order_by("AGENCE")
    exploitants_opts = qs.values_list("EXPL", flat=True).distinct().order_by("EXPL")

                          
    donnees_queryset = qs.order_by("FILIALE", "AGENCE", "EXPL", "CLIENT")

    paginator = Paginator(donnees_queryset, 100)
    page = request.GET.get('page')
    try:
        donnees_page = paginator.page(page)
    except:
        donnees_page = paginator.page(1)

    context = {
        "donnees": donnees_page,
        "filiales": filiales_opts,
        "agences": agences_opts,
        "exploitants": exploitants_opts,
        "periode": periode_param,
        "filiale_param": filiale_param,
        "agence_param": agence_param,
        "expl_param": expl_param,
        "can_pick_filiale": is_group_user,
        "can_pick_agence": (is_group_user or organe in users_filiale or organe == "Directeur Agence"),
        "can_pick_expl": (is_group_user or organe in users_filiale or organe == "Directeur Agence"),
        "get_params": request.GET.urlencode(),
    }
    return render(request, "scoring.html", context)

from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
from .models import DATEREV

@login_required
def export_csv_scoring(request):
    user = request.user
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

                                                         
    periode_param = request.GET.get("periode", "")
    filiale_param = request.GET.get("filiale", "")
    agence_param = request.GET.get("agence", "")
    expl_param = request.GET.get("expl", "")
    filiale_txt = request.GET.get("filiale_txt", "").strip()
    agence_txt = request.GET.get("agence_txt", "").strip()
    lib_agence = request.GET.get("lib_agence", "").strip()
    expl_txt = request.GET.get("expl_txt", "").strip()
    client_txt = request.GET.get("client", "").strip()
    daterev_txt = request.GET.get("daterev", "").strip()
    filiale_txt = request.GET.get("filiale_txt", "").strip()
    agence_txt = request.GET.get("agence_txt", "").strip()
    lib_agence = request.GET.get("lib_agence", "").strip()
    expl_txt = request.GET.get("expl_txt", "").strip()
    client_txt = request.GET.get("client", "").strip()
    daterev_txt = request.GET.get("daterev", "").strip()

    base_qs = DATEREV.objects.filter(DATEREV__isnull=False)

    if getattr(user, "organe", "") == "Chargé Client":
        base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        base_qs = base_qs.filter(FILIALE=user.filiale)
    elif user.organe in users_groupe:
                                                                     
        pass
    else:
                                                      
        base_qs = DATEREV.objects.none()

                                           
    qs_period = base_qs
    today = date.today()
    if periode_param == "today":
        qs_period = qs_period.filter(DATEREV__lte=today)
    elif periode_param == "3m":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=90))
    elif periode_param == "6m":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=180))
    elif periode_param == "1y":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=365))

             
    if filiale_param:
        qs_period = qs_period.filter(FILIALE=filiale_param)

            
    if agence_param:
        qs_period = qs_period.filter(AGENCE=agence_param)

                
    if expl_param:
        qs_period = qs_period.filter(EXPL=expl_param)
    if filiale_txt:
        qs_period = qs_period.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        qs_period = qs_period.filter(AGENCE__icontains=agence_txt)
    if lib_agence:
        qs_period = qs_period.filter(LIB_AGENCE__icontains=lib_agence)
    if expl_txt:
        qs_period = qs_period.filter(EXPL__icontains=expl_txt)
    if client_txt:
        qs_period = qs_period.filter(CLIENT__icontains=client_txt)
    if daterev_txt:
        parsed = parse_date(daterev_txt)
        if not parsed:
            for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                try:
                    parsed = datetime.strptime(daterev_txt, fmt).date()
                    break
                except (ValueError, TypeError):
                    continue
        if parsed:
            qs_period = qs_period.filter(DATEREV=parsed)
    if filiale_txt:
        qs_period = qs_period.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        qs_period = qs_period.filter(AGENCE__icontains=agence_txt)
    if lib_agence:
        qs_period = qs_period.filter(LIB_AGENCE__icontains=lib_agence)
    if expl_txt:
        qs_period = qs_period.filter(EXPL__icontains=expl_txt)
    if client_txt:
        qs_period = qs_period.filter(CLIENT__icontains=client_txt)
    if daterev_txt:
        parsed = parse_date(daterev_txt)
        if not parsed:
            for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                try:
                    parsed = datetime.strptime(daterev_txt, fmt).date()
                    break
                except (ValueError, TypeError):
                    continue
        if parsed:
            qs_period = qs_period.filter(DATEREV=parsed)

    donnees = qs_period

                                
    wb = Workbook()
    ws = wb.active
    ws.title = "Notation_Stock"

                                                   
    headers = ['FILIALE', 'AGENCE', 'EXPL', 'CLIENT', 'DATEREV', 'PPE', 'RISQUE']
    ws.append(headers)

    for d in donnees:
        daterev = d.DATEREV
        if hasattr(daterev, 'tzinfo'):
            daterev = daterev.replace(tzinfo=None)
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, daterev, d.PPE, d.RISQUE
        ])

    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"scoring_export_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required
def export_csv_scoring_ppe(request):
    user = request.user
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

                                                         
    periode_param = request.GET.get("periode", "")
    filiale_param = request.GET.get("filiale", "")
    agence_param = request.GET.get("agence", "")
    expl_param = request.GET.get("expl", "")

    if user.organe == 'Conformité':
        base_qs = DATEREV.objects.filter(FILIALE=user.filiale, PPE__icontains="O", DATEREV__isnull=False)
    elif user.organe == "Conformité Groupe":
        base_qs = DATEREV.objects.filter(PPE__icontains="O", DATEREV__isnull=False)

                                           
    qs_period = base_qs
    today = date.today()
    if periode_param == "today":
        qs_period = qs_period.filter(DATEREV__lte=today)
    elif periode_param == "3m":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=90))
    elif periode_param == "6m":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=180))
    elif periode_param == "1y":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=365))

             
    if filiale_param:
        qs_period = qs_period.filter(FILIALE=filiale_param)

            
    if agence_param:
        qs_period = qs_period.filter(AGENCE=agence_param)

                
    if expl_param:
        qs_period = qs_period.filter(EXPL=expl_param)

    donnees = qs_period

                                
    wb = Workbook()
    ws = wb.active
    ws.title = "Notation_Stock"

                                                   
    headers = ['FILIALE', 'AGENCE', 'EXPL', 'CLIENT', 'DATEREV', 'PPE', 'RISQUE']
    ws.append(headers)

    for d in donnees:
        daterev = d.DATEREV
        if hasattr(daterev, 'tzinfo'):
            daterev = daterev.replace(tzinfo=None)
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, daterev, d.PPE, d.RISQUE
        ])

    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"scoring_PPE_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required
def clients_scorer(request):
    from .models import Notation, Kyc_pm, Kyc_pp
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau", 'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                    "Contrôle Permanent Groupe", "PASS", "GUEST"]

    user = request.user
    notes = Notation.objects.filter(flux_stock='Flux')
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])
    if user.is_authenticated and getattr(user, 'filiale', ''):
        notation = notation.filter(agent__filiale=user.filiale)
    if user.is_authenticated and getattr(user, 'code_expl', ''):
        notation = notation.filter(agent__code_expl=user.code_expl)

    periode_param = request.GET.get("periode", "")
    filiale_param = request.GET.get("filiale", "")
    agence_param = request.GET.get("agence", "")
    expl_param = request.GET.get("expl", "")
    risque_param = request.GET.get("risque", "")
    client_type = request.GET.get('type_client', 'all')
    if client_type not in ('pp', 'pm', 'all'):
        client_type = 'all'

    col_filiale = request.GET.get('col_filiale', '')
    col_agence = request.GET.get('col_agence', '')
    col_lib_agence = request.GET.get('col_lib_agence', '')
    col_expl = request.GET.get('col_expl', '')
    col_client = request.GET.get('col_client', '')
    col_daterev = request.GET.get('col_daterev', '')
    col_ppe = request.GET.get('col_ppe', '')
    col_risque = request.GET.get('col_risque', '')

    today = date.today()
    today_str = today.isoformat()
    FIELDS = ("FILIALE", "AGENCE", "LIB_AGENCE", "EXPL", "CLIENT", "DATEREV", "PPE", "RISQUE")

                                                                                    
    if client_type == 'pm':
        qsets = [Kyc_pm.objects.all()]
    elif client_type == 'pp':
        qsets = [Kyc_pp.objects.all()]
    else:         
        qsets = [Kyc_pp.objects.all(), Kyc_pm.objects.all()]

    def _f(qs_list, *args, **kw):
        return [q.filter(*args, **kw) for q in qs_list]

    def _distinct(qs_list, field):
        vals = set()
        for q in qs_list:
            vals.update(q.values_list(field, flat=True).distinct())
        return sorted(v for v in vals if v not in (None, ''))

    is_group_user = (user.organe in users_groupe) or (user.filiale in ["BOA Group", "BOA GROUP"]) or (not user.filiale)
    if not is_group_user:
        if getattr(user, "organe", "") == "Chargé Client":
            qsets = _f(qsets, FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
        elif user.organe == "Directeur Agence":
            qsets = _f(qsets, FILIALE=user.filiale, AGENCE=user.agence)
        elif user.organe in users_filiale:
            qsets = _f(qsets, FILIALE=user.filiale)

                                                                                         
                                                                                                      
    if periode_param == "today":
        qsets = [q.exclude(DATEREV='').filter(DATEREV__lt=today_str) for q in qsets]
    elif periode_param == "3m":
        qsets = [q.exclude(DATEREV='').filter(DATEREV__gte=today_str, DATEREV__lte=(today + timedelta(days=90)).isoformat()) for q in qsets]
    elif periode_param == "6m":
        qsets = [q.exclude(DATEREV='').filter(DATEREV__gte=today_str, DATEREV__lte=(today + timedelta(days=180)).isoformat()) for q in qsets]
    elif periode_param == "1y":
        qsets = [q.exclude(DATEREV='').filter(DATEREV__gte=today_str, DATEREV__lte=(today + timedelta(days=365)).isoformat()) for q in qsets]
    elif periode_param == "no_date":
        qsets = _f(qsets, DATEREV='')

    can_pick_filiale = is_group_user
    selected_filiale = filiale_param if can_pick_filiale else getattr(user, "filiale", "")
    _user_scope = "groupe" if is_group_user else getattr(user, 'filiale', 'filiale')
    filiales_cache_key = "filiales_opts_kyc:" + hashlib.md5(
        f"{client_type}_{periode_param}_{_user_scope}".encode("utf-8")).hexdigest()
    filiales_opts = cache.get(filiales_cache_key)
    if filiales_opts is None:
        filiales_opts = _distinct(qsets, "FILIALE")
        cache.set(filiales_cache_key, filiales_opts, 300)

    if selected_filiale:
        qsets = _f(qsets, FILIALE=selected_filiale)

    can_pick_agence = (user.organe in users_groupe) or (user.organe in users_filiale) or (user.organe == "Directeur Agence")
    selected_agence = getattr(user, "agence", "") if user.organe == "Directeur Agence" else agence_param
    agences_cache_key = "agences_opts_kyc:" + hashlib.md5(
        f"{client_type}_{periode_param}_{selected_filiale}_{_user_scope}".encode("utf-8")).hexdigest()
    agences_opts = cache.get(agences_cache_key)
    if agences_opts is None:
        agences_opts = _distinct(qsets, "AGENCE")
        cache.set(agences_cache_key, agences_opts, 300)

    if selected_agence:
        qsets = _f(qsets, AGENCE=selected_agence)

    qsets_agence = qsets                                                               

    can_pick_expl = getattr(user, "organe", "") != "Chargé Client"
    selected_expl = getattr(user, "code_expl", "") if getattr(user, "organe", "") == "Chargé Client" else expl_param
    exploitants_cache_key = "expl_opts_kyc:" + hashlib.md5(
        f"{client_type}_{periode_param}_{selected_filiale}_{selected_agence}_{_user_scope}".encode("utf-8")).hexdigest()
    exploitants_opts = cache.get(exploitants_cache_key)
    if exploitants_opts is None:
        exploitants_opts = _distinct(qsets_agence, "EXPL")
        cache.set(exploitants_cache_key, exploitants_opts, 300)

    dq = list(qsets_agence)
    if selected_expl:
        dq = _f(dq, EXPL=selected_expl)
    if col_filiale: dq = _f(dq, FILIALE__icontains=col_filiale)
    if col_agence: dq = _f(dq, AGENCE__icontains=col_agence)
    if col_lib_agence: dq = _f(dq, LIB_AGENCE__icontains=col_lib_agence)
    if col_expl: dq = _f(dq, EXPL__icontains=col_expl)
    if col_client: dq = _f(dq, CLIENT__icontains=col_client)
    if col_daterev: dq = _f(dq, DATEREV__icontains=col_daterev)
    if col_ppe: dq = _f(dq, PPE__icontains=col_ppe)
    datouv_start = (request.GET.get('datouv_start') or '').strip()
    datouv_end = (request.GET.get('datouv_end') or '').strip()
    if datouv_start: dq = _f(dq, DATOUV__gte=datouv_start)
    if datouv_end: dq = [q.exclude(DATOUV='').filter(DATOUV__lte=datouv_end) for q in dq]
    if col_risque: dq = _f(dq, RISQUE__icontains=col_risque)

                                                                   
    scorer_counts_key = "scorer_counts_kyc:" + hashlib.md5(
        ("||".join(str(q.query) for q in dq) + f"|{today_str}").encode("utf-8")).hexdigest()
    cached_counts = cache.get(scorer_counts_key)
    if cached_counts is not None:
        scorer_scored_count, scorer_unscored_count, overdue_unscored_count = cached_counts
    else:
        scorer_scored_count = scorer_unscored_count = overdue_unscored_count = 0
        for q in dq:
            a = q.aggregate(
                scored=Count(Case(When(~Q(RISQUE="") & ~Q(RISQUE__isnull=True), then=1), output_field=IntegerField())),
                unscored=Count(Case(When(Q(RISQUE="") | Q(RISQUE__isnull=True), then=1), output_field=IntegerField())),
                overdue_unscored=Count(Case(When((Q(RISQUE="") | Q(RISQUE__isnull=True)) & ~Q(DATEREV="") & Q(DATEREV__lt=today_str), then=1), output_field=IntegerField())),
            )
            scorer_scored_count += a['scored'] or 0
            scorer_unscored_count += a['unscored'] or 0
            overdue_unscored_count += a['overdue_unscored'] or 0
        cache.set(scorer_counts_key, (scorer_scored_count, scorer_unscored_count, overdue_unscored_count), 300)

    total_scorer = scorer_scored_count + scorer_unscored_count
    scoring_rate = (scorer_scored_count / total_scorer * 100) if total_scorer > 0 else 0.0
    if scorer_unscored_count > 0 and scoring_rate >= 99.9:
        scoring_rate = 99.9

    risk_options = cache.get("risk_options_opts_kyc")
    if risk_options is None:
        rv = set(Kyc_pp.objects.exclude(RISQUE="").exclude(RISQUE__isnull=True).values_list('RISQUE', flat=True).distinct())
        rv.update(Kyc_pm.objects.exclude(RISQUE="").exclude(RISQUE__isnull=True).values_list('RISQUE', flat=True).distinct())
        risk_options = sorted(v for v in rv if v)
        cache.set("risk_options_opts_kyc", risk_options, 300)

    if risque_param:
        if risque_param == "sans_classe":
            dq = _f(dq, Q(RISQUE="") | Q(RISQUE__isnull=True))
        else:
            dq = _f(dq, RISQUE=risque_param)

                                            
    _vals = [q.values(*FIELDS) for q in dq]
    if len(_vals) == 1:
        donnees_queryset = _vals[0].order_by("FILIALE", "AGENCE", "EXPL", "CLIENT")
    else:
        donnees_queryset = _vals[0].union(_vals[1], all=True).order_by("FILIALE", "AGENCE", "EXPL", "CLIENT")

    show_non_scored_modal = request.GET.get('show_non_scored_modal') == '1'
    is_overdue = request.GET.get('overdue') == '1'
    non_scored_page = None
    if show_non_scored_modal:
        nsq = list(qsets_agence)
        if selected_expl: nsq = _f(nsq, EXPL=selected_expl)
        if col_filiale: nsq = _f(nsq, FILIALE__icontains=col_filiale)
        if col_agence: nsq = _f(nsq, AGENCE__icontains=col_agence)
        if col_lib_agence: nsq = _f(nsq, LIB_AGENCE__icontains=col_lib_agence)
        if col_expl: nsq = _f(nsq, EXPL__icontains=col_expl)
        if col_client: nsq = _f(nsq, CLIENT__icontains=col_client)
        if col_daterev: nsq = _f(nsq, DATEREV__icontains=col_daterev)
        if col_ppe: nsq = _f(nsq, PPE__icontains=col_ppe)
        nsq = _f(nsq, Q(RISQUE="") | Q(RISQUE__isnull=True))
        if is_overdue:
            nsq = [q.exclude(DATEREV='').filter(DATEREV__lt=today_str) for q in nsq]
        _nsvals = [q.values(*FIELDS) for q in nsq]
        if len(_nsvals) == 1:
            non_scored_qs = _nsvals[0].order_by("CLIENT")
        else:
            non_scored_qs = _nsvals[0].union(_nsvals[1], all=True).order_by("CLIENT")
        ns_paginator = CachedPaginator(non_scored_qs, 50)
        try: non_scored_page = ns_paginator.page(request.GET.get('non_scored_page', 1))
        except (PageNotAnInteger, EmptyPage): non_scored_page = ns_paginator.page(1)

    paginator = CachedPaginator(donnees_queryset, 100)
    try: donnees_page = paginator.page(request.GET.get('page'))
    except PageNotAnInteger: donnees_page = paginator.page(1)
    except EmptyPage: donnees_page = paginator.page(paginator.num_pages)

    current_get = request.GET.copy()
    current_get.pop('page', None)
    get_params = current_get.urlencode()
    
    close_get = current_get.copy()
    close_get.pop('show_non_scored_modal', None)
    close_get.pop('non_scored_page', None)
    close_get.pop('overdue', None)                                                                    
    non_scored_close_params = close_get.urlencode()

    modal_get = close_get.copy()
    modal_get['show_non_scored_modal'] = '1'
    non_scored_modal_params = modal_get.urlencode()

    overdue_modal_get = modal_get.copy()
    overdue_modal_get['overdue'] = '1'
    overdue_non_scored_modal_params = overdue_modal_get.urlencode()

    export_get = close_get.copy()
    export_get['export_unscored'] = '1'
    if is_overdue:                                                     
        export_get['overdue'] = '1'
    non_scored_export_params = export_get.urlencode()
    
    pp_nav = current_get.copy()
    pp_nav['type_client'] = 'pp'
    pp_nav_params = pp_nav.urlencode()
    
    pm_nav = current_get.copy()
    pm_nav['type_client'] = 'pm'
    pm_nav_params = pm_nav.urlencode()
    
    reset_params = f"type_client={client_type}"

    context = {
        "donnees": donnees_page, "filiales": filiales_opts, "agences": agences_opts, "exploitants": exploitants_opts,
        "notation": notation, "periode": periode_param, "filiale_param": selected_filiale, "agence_param": selected_agence,
        "expl_param": selected_expl, "risque_param": risque_param, "client_type": client_type, "risk_options": risk_options,
        "scorer_scored_count": scorer_scored_count, "scorer_unscored_count": scorer_unscored_count, "scoring_rate": scoring_rate,
        "overdue_unscored_count": overdue_unscored_count, "show_non_scored_modal": show_non_scored_modal,
        "non_scored_page": non_scored_page, "users_groupe": users_groupe, "users_filiale": users_filiale,
        "can_pick_filiale": can_pick_filiale, "can_pick_agence": can_pick_agence, "can_pick_expl": can_pick_expl,
        "get_params": get_params, "reset_params": reset_params, "non_scored_modal_params": non_scored_modal_params,
        "overdue_non_scored_modal_params": overdue_non_scored_modal_params, "is_overdue_modal": is_overdue,
        "non_scored_close_params": non_scored_close_params, "non_scored_export_params": non_scored_export_params,
        "pp_nav_params": pp_nav_params, "pm_nav_params": pm_nav_params
    }
    return render(request, "clients_scorer.html", context)

@login_required
def export_csv_scoring_clients(request):
    from .models import Kyc_pm, Kyc_pp
    user = request.user
    periode_param, filiale_param, agence_param, expl_param = request.GET.get("periode", ""), request.GET.get("filiale", ""), request.GET.get("agence", ""), request.GET.get("expl", "")
    risque_param, client_type, export_unscored = request.GET.get("risque", ""), request.GET.get('type_client', 'pp'), request.GET.get('export_unscored') == '1'
    col_agence, col_lib_agence, col_expl, col_client, col_daterev, col_ppe, col_risque = request.GET.get('col_agence', ''), request.GET.get('col_lib_agence', ''), request.GET.get('col_expl', ''), request.GET.get('col_client', ''), request.GET.get('col_daterev', ''), request.GET.get('col_ppe', ''), request.GET.get('col_risque', '')

    today = date.today()
    today_str = today.isoformat()
    if client_type == 'pm':
        qsets = [Kyc_pm.objects.all()]
    elif client_type == 'pp':
        qsets = [Kyc_pp.objects.all()]
    else:
        qsets = [Kyc_pp.objects.all(), Kyc_pm.objects.all()]

    def _f(qs_list, *args, **kw):
        return [q.filter(*args, **kw) for q in qs_list]

    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                    "Contrôle Permanent Groupe", "PASS", "GUEST"]
    is_group_user = (user.organe in users_groupe) or (user.filiale in ["BOA Group", "BOA GROUP"]) or (not user.filiale)
    if not is_group_user:
        if getattr(user, "organe", "") == "Chargé Client": qsets = _f(qsets, FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
        elif user.organe == "Directeur Agence": qsets = _f(qsets, FILIALE=user.filiale, AGENCE=user.agence)
        elif user.organe in ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau", 'Risques', 'DAI', 'Qualité']: qsets = _f(qsets, FILIALE=user.filiale)

    if periode_param == "today": qsets = [q.exclude(DATEREV='').filter(DATEREV__lt=today_str) for q in qsets]
    elif periode_param == "3m": qsets = [q.exclude(DATEREV='').filter(DATEREV__gte=today_str, DATEREV__lte=(today + timedelta(days=90)).isoformat()) for q in qsets]
    elif periode_param == "6m": qsets = [q.exclude(DATEREV='').filter(DATEREV__gte=today_str, DATEREV__lte=(today + timedelta(days=180)).isoformat()) for q in qsets]
    elif periode_param == "1y": qsets = [q.exclude(DATEREV='').filter(DATEREV__gte=today_str, DATEREV__lte=(today + timedelta(days=365)).isoformat()) for q in qsets]
    elif periode_param == "no_date": qsets = _f(qsets, DATEREV='')

    if filiale_param: qsets = _f(qsets, FILIALE=filiale_param)
    if agence_param: qsets = _f(qsets, AGENCE=agence_param)
    if expl_param: qsets = _f(qsets, EXPL=expl_param)

    if export_unscored: qsets = _f(qsets, Q(RISQUE="") | Q(RISQUE__isnull=True))
    elif risque_param:
        if risque_param == "sans_classe": qsets = _f(qsets, Q(RISQUE="") | Q(RISQUE__isnull=True))
        else: qsets = _f(qsets, RISQUE=risque_param)

    if col_agence: qsets = _f(qsets, AGENCE__icontains=col_agence)
    if col_lib_agence: qsets = _f(qsets, LIB_AGENCE__icontains=col_lib_agence)
    if col_expl: qsets = _f(qsets, EXPL__icontains=col_expl)
    if col_client: qsets = _f(qsets, CLIENT__icontains=col_client)
    if col_daterev: qsets = _f(qsets, DATEREV__icontains=col_daterev)
    if col_ppe: qsets = _f(qsets, PPE__icontains=col_ppe)
    if col_risque: qsets = _f(qsets, RISQUE__icontains=col_risque)
    datouv_start = (request.GET.get('datouv_start') or '').strip()
    datouv_end = (request.GET.get('datouv_end') or '').strip()
    if datouv_start: qsets = _f(qsets, DATOUV__gte=datouv_start)
    if datouv_end: qsets = [q.exclude(DATOUV='').filter(DATOUV__lte=datouv_end) for q in qsets]

    cols = ("FILIALE", "AGENCE", "EXPL", "CLIENT", "DATEREV", "PPE", "RISQUE")
    _ev = [q.values(*cols) for q in qsets]
    if len(_ev) == 1:
        donnees = _ev[0].order_by("FILIALE", "AGENCE", "EXPL", "CLIENT")
    else:
        donnees = _ev[0].union(_ev[1], all=True).order_by("FILIALE", "AGENCE", "EXPL", "CLIENT")

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ['FILIALE', 'AGENCE', 'EXPL', 'CLIENT', 'DATEREV', 'PPE', 'RISQUE']
    ws.append(headers)
    for d in donnees:
        ws.append([d["FILIALE"], d["AGENCE"], d["EXPL"], d["CLIENT"], str(d["DATEREV"] or ""), d["PPE"], d["RISQUE"]])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Revue_scoring_{datetime.now().strftime("%Y-%m-%d_%H-%M")}.xlsx"'
    return response

@login_required
def sans_classe(request):
    user = request.user

                             
    roles_exclus = ["Chargé Client"]
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]

    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')
    filiale_txt = request.GET.get('filiale_txt', '').strip()
    agence_txt = request.GET.get('agence_txt', '').strip()
    lib_agence = request.GET.get('lib_agence', '').strip()
    expl_txt = request.GET.get('expl_txt', '').strip()
    client_txt = request.GET.get('client', '').strip()
    risque_txt = request.GET.get('risque', '').strip()

                                                                     
                                                      
    donnees_queryset = DATEREV.objects.filter(Q(RISQUE__isnull=True) | Q(RISQUE=""))

                                                               
    is_group_user = (user.organe in users_groupe) or (user.filiale in ["BOA Group", "BOA GROUP"]) or (not user.filiale)
    if not is_group_user:
        if user.organe == "Chargé Client":
            donnees_queryset = donnees_queryset.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)
        elif user.organe == "Directeur Agence":
            donnees_queryset = donnees_queryset.filter(FILIALE=user.filiale, AGENCE=user.agence)
        elif user.organe in users_filiale:
            donnees_queryset = donnees_queryset.filter(FILIALE=user.filiale)

                                                
    if filiale_param:
        donnees_queryset = donnees_queryset.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees_queryset = donnees_queryset.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees_queryset = donnees_queryset.filter(EXPL__icontains=expl_param)
    if filiale_txt:
        donnees_queryset = donnees_queryset.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        donnees_queryset = donnees_queryset.filter(AGENCE__icontains=agence_txt)
    if lib_agence:
        donnees_queryset = donnees_queryset.filter(LIB_AGENCE__icontains=lib_agence)
    if expl_txt:
        donnees_queryset = donnees_queryset.filter(EXPL__icontains=expl_txt)
    if client_txt:
        donnees_queryset = donnees_queryset.filter(CLIENT__icontains=client_txt)
    if risque_txt:
        donnees_queryset = donnees_queryset.filter(RISQUE__icontains=risque_txt)

                                     
    donnees_queryset = donnees_queryset.order_by("FILIALE", "AGENCE", "EXPL", "CLIENT")

                                                                     
    options_qs = DATEREV.objects.filter(Q(RISQUE__isnull=True) | Q(RISQUE=""))
    if not is_group_user:
        if user.organe == "Chargé Client":
            options_qs = options_qs.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)
        elif user.organe == "Directeur Agence":
            options_qs = options_qs.filter(FILIALE=user.filiale, AGENCE=user.agence)
        elif user.organe in users_filiale:
            options_qs = options_qs.filter(FILIALE=user.filiale)

    filiales = options_qs.values_list('FILIALE', flat=True).distinct().order_by('FILIALE')
    agences = options_qs.values_list('AGENCE', flat=True).distinct().order_by('AGENCE')
    exploitants = options_qs.values_list('EXPL', flat=True).distinct().order_by('EXPL')

                              
    get_params = request.GET.copy()
    if 'page' in get_params:
        del get_params['page']

    paginator = Paginator(donnees_queryset, 100)
    page = request.GET.get('page')

    try:
        donnees_page = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        donnees_page = paginator.page(1)

    context = {
        'donnees': donnees_page,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
        'get_params': get_params.urlencode(),
        'filiale_param': filiale_param,
        'agence_param': agence_param,
        'expl_param': expl_param,
    }

    return render(request, 'sans_classe.html', context)


@login_required
def export_sans_classe(request):
    user = request.user
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

                               
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

                                                              
    donnees = DATEREV.objects.filter(Q(RISQUE__isnull=True) | Q(RISQUE=""))

    is_group_user = (user.organe in users_groupe) or (user.filiale in ["BOA Group", "BOA GROUP"]) or (not user.filiale)
                            
    if is_group_user:
        pass
    elif user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    else:
                                                                           
        donnees = DATEREV.objects.none()

                                           
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)
    if filiale_txt:
        donnees = donnees.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        donnees = donnees.filter(AGENCE__icontains=agence_txt)
    if lib_agence:
        donnees = donnees.filter(LIB_AGENCE__icontains=lib_agence)
    if expl_txt:
        donnees = donnees.filter(EXPL__icontains=expl_txt)
    if client_txt:
        donnees = donnees.filter(CLIENT__icontains=client_txt)
    if risque_txt:
        donnees = donnees.filter(RISQUE__icontains=risque_txt)

                               
    wb = Workbook()
    ws = wb.active
    ws.title = "Sans_Classe_Export"

    headers = ['FILIALE', 'AGENCE', 'EXPL', 'CLIENT', 'DATEREV', 'PPE', 'RISQUE']
    ws.append(headers)

    for d in donnees:
        daterev = d.DATEREV
        if hasattr(daterev, 'tzinfo'):
            daterev = daterev.replace(tzinfo=None)
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, daterev, d.PPE, d.RISQUE
        ])

    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Sans_Classe_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


from django.db.models import Q, Max


@login_required
def sans_classe_s(request):
    user = request.user

                             
    roles_exclus = ["Chargé Client"]
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]

    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

                                                                          
    notes = Notation.objects.filter(flux_stock='Flux')
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])

                                           
    if user.organe == "Chargé Client":
        notation = notation.filter(agent__filiale=user.filiale, agent__code_expl=user.code_expl)
    elif user.organe == "Directeur Agence":
        notation = notation.filter(agent__filiale=user.filiale, agent__agence=user.agence)
    elif user.organe in users_filiale:
        notation = notation.filter(agent__filiale=user.filiale)
                                                                        

                                                                    
                                                  
    donnees_queryset = DATEREV.objects.filter(Q(RISQUE__isnull=True) | Q(RISQUE=""))

                                                   
    if user.organe == "Chargé Client":
        donnees_queryset = donnees_queryset.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees_queryset = donnees_queryset.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees_queryset = donnees_queryset.filter(FILIALE=user.filiale)

                                             
    if filiale_param:
        donnees_queryset = donnees_queryset.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees_queryset = donnees_queryset.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees_queryset = donnees_queryset.filter(EXPL__icontains=expl_param)
    if filiale_txt:
        donnees_queryset = donnees_queryset.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        donnees_queryset = donnees_queryset.filter(AGENCE__icontains=agence_txt)
    if expl_txt:
        donnees_queryset = donnees_queryset.filter(EXPL__icontains=expl_txt)
    if client_txt:
        donnees_queryset = donnees_queryset.filter(CLIENT__icontains=client_txt)
    if risque_txt:
        donnees_queryset = donnees_queryset.filter(RISQUE__icontains=risque_txt)

    donnees_queryset = donnees_queryset.order_by("FILIALE", "AGENCE", "EXPL", "CLIENT")

                                                    
    options_qs = DATEREV.objects.all()
    if user.organe == "Chargé Client":
        options_qs = options_qs.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        options_qs = options_qs.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        options_qs = options_qs.filter(FILIALE=user.filiale)

    filiales = options_qs.values_list('FILIALE', flat=True).distinct().order_by('FILIALE')
    agences = options_qs.values_list('AGENCE', flat=True).distinct().order_by('AGENCE')
    exploitants = options_qs.values_list('EXPL', flat=True).distinct().order_by('EXPL')

                   
    get_params = request.GET.copy()
    if 'page' in get_params:
        del get_params['page']

    paginator = Paginator(donnees_queryset, 100)
    page = request.GET.get('page')

    try:
        donnees_page = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        donnees_page = paginator.page(1)

    context = {
        'donnees': donnees_page,
        'filiales': filiales,
        'notation': notation,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
        'get_params': get_params.urlencode(),
        'filiale_param': filiale_param,
        'agence_param': agence_param,
        'expl_param': expl_param,
    }

    return render(request, 'sans_classe_s.html', context)


@login_required
def export_sans_classe_s(request):
    def strip_tz(value):
        if hasattr(value, 'tzinfo'):
            return value.replace(tzinfo=None)
        return value

    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')
    filiale_txt = request.GET.get('filiale_txt', '').strip()
    agence_txt = request.GET.get('agence_txt', '').strip()
    expl_txt = request.GET.get('expl_txt', '').strip()
    client_txt = request.GET.get('client', '').strip()
    risque_txt = request.GET.get('risque', '').strip()

    donnees = DATEREV.objects.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)

    donnees = donnees.filter(Q(RISQUE__isnull=True) | Q(RISQUE=""))
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)
    if filiale_txt:
        donnees = donnees.filter(FILIALE__icontains=filiale_txt)
    if agence_txt:
        donnees = donnees.filter(AGENCE__icontains=agence_txt)
    if expl_txt:
        donnees = donnees.filter(EXPL__icontains=expl_txt)
    if client_txt:
        donnees = donnees.filter(CLIENT__icontains=client_txt)
    if risque_txt:
        donnees = donnees.filter(RISQUE__icontains=risque_txt)


                                
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients non classés"

              
    headers = ['AGENCE', 'AGENCE', 'EXPL', 'CLIENT', 'DATEREV', 'PPE', 'RISQUE']
    ws.append(headers)

             
    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.DATEREV, d.PPE, d.RISQUE

        ])

                                                 
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

                              
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Clients sans classe de risque {date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response




ITEMS_PER_PAGE = 100                                         

from django.shortcuts import render
from django.db.models import Max
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Kyc_pm, Notation                                                              

                                                                         
from django.db.models import Q
from kyc.models import KycFieldVisibilityConfig

def _kyc_field_empty_q(field_name):
    """Critère « champ vide » partagé entre le filtre d'affichage /non_rens
    et le calcul des taux de complétude par agence (mêmes valeurs de remplissage
    considérées comme vides : '', NULL, XX, RAS, R.A.S(.), longueur 1 ou
    caractères de remplissage pour les champs numériques)."""
    if field_name in ['SALAIRE', 'CAPITAL', 'CA', 'RESULTAT']:
        return (
            Q(**{f"{field_name}__isnull": True}) |
            Q(**{f"{field_name}__exact": ""}) |
            Q(**{f"{field_name}__iexact": "XX"}) |
            Q(**{f"{field_name}__iexact": "RAS"}) |
            Q(**{f"{field_name}__iexact": "R.A.S."}) |
            Q(**{f"{field_name}__iexact": "R.A.S"}) |
            Q(**{f"{field_name}__in": [".", "?", "-", "*"]})
        )
    return (
        Q(**{f"{field_name}__isnull": True}) |
        Q(**{f"{field_name}__exact": ""}) |
        Q(**{f"{field_name}__iexact": "XX"}) |
        Q(**{f"{field_name}__iexact": "RAS"}) |
        Q(**{f"{field_name}__iexact": "R.A.S."}) |
        Q(**{f"{field_name}__iexact": "R.A.S"}) |
        Q(**{f"{field_name}__length": 1})
    )


def _resolve_kyc_config(configs, filiale):
    """Config champs KYC applicable à une filiale : spécifique sinon globale."""
    config = next((c for c in configs if filiale in (c.filiales or [])), None)
    if not config:
        config = next((c for c in configs if not c.filiales), None)
    return config


def apply_kyc_field_config_filter(queryset, client_type):
    """
    Filtre le queryset pour ne garder que les clients dont au moins un des
    champs définis dans KycFieldVisibilityConfig (empty_check_fields) est vide.
    """
                                          
    if not queryset.exists():
        return queryset

    filiales = list(queryset.order_by().values_list("FILIALE", flat=True).distinct())
    if not filiales:
        return queryset.none()

    combined_q = None
    
                                                    
    configs = list(KycFieldVisibilityConfig.objects.filter(client_type=client_type))
    
    for filiale in filiales:
        config = _resolve_kyc_config(configs, filiale)

        if config and config.empty_check_fields:
            missing_q = None
            for field_name in config.empty_check_fields:
                field_q = _kyc_field_empty_q(field_name)
                missing_q = field_q if missing_q is None else missing_q | field_q
                
            if missing_q is not None:
                scoped_q = Q(FILIALE=filiale) & missing_q
                combined_q = scoped_q if combined_q is None else combined_q | scoped_q
        else:
                                                                                                 
                                                                           
                                                                                       
                                                                                                                      
            pass

    if combined_q is None:
        return queryset.none()
        
    return queryset.filter(combined_q)

                                                           
def get_filtered_queryset_pm(request):
    """Garantit que l'utilisateur ne voit que les entreprises (PM) de son périmètre."""
    user = request.user
    queryset = Kyc_pm.objects.all()

    if user.organe == "Chargé Client":
        queryset = queryset.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)

    elif user.organe == "Directeur Agence":
        queryset = queryset.filter(FILIALE=user.filiale, AGENCE=user.agence)

    elif user.organe in ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']:
        queryset = queryset.filter(FILIALE=user.filiale)

    elif user.organe in ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                         "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]:
                                     
        pass

    return queryset.order_by('id')

                                              
def get_filter_lists_pm(user, request):
    """Génère les options des menus déroulants PM selon les droits d'accès."""
    filiale_list, agence_list, expl_list = [], [], []
    base_qs = Kyc_pm.objects.all()

                                                     
    if user.organe == "Chargé Client":
        base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']:
        base_qs = base_qs.filter(FILIALE=user.filiale)

                                                       
    filiale_list = Kyc_pm.objects.values_list("FILIALE", flat=True).distinct().order_by("FILIALE")

                                                     
    f_filiale = request.GET.get("filiale")
    f_agence = request.GET.get("agence")

             
    if f_filiale:
        agence_list = Kyc_pm.objects.filter(FILIALE=f_filiale).values_list("AGENCE", flat=True).distinct()
    elif user.organe in ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']:
        agence_list = base_qs.values_list("AGENCE", flat=True).distinct()

                 
    if f_agence:
        expl_list = Kyc_pm.objects.filter(AGENCE=f_agence).values_list("EXPL", flat=True).distinct()
    elif user.organe == "Directeur Agence":
        expl_list = base_qs.values_list("EXPL", flat=True).distinct()
    elif not f_agence and (f_filiale or user.organe in ["DSI", "Conformité", "Contrôle Permanent"]):
        expl_list = base_qs.values_list("EXPL", flat=True).distinct()

                          
    return filiale_list, agence_list, expl_list

                              
@login_required
def non_rens_pm(request):
    user = request.user

                                                   
    queryset = get_filtered_queryset_pm(request)
    
                                                               
    queryset = apply_kyc_field_config_filter(queryset, "pm")

                                              
    f_filiale = request.GET.get('filiale')
    f_agence = request.GET.get('agence')
    f_expl = request.GET.get('expl')
    f_lib_agence = request.GET.get('col_lib_agence') or request.GET.get('lib_agence')
    f_client = request.GET.get('col_client') or request.GET.get('client')
    f_idm = request.GET.get('col_idm') or request.GET.get('idm')
    f_agec = request.GET.get('col_agec') or request.GET.get('agec')
    f_codape = request.GET.get('col_codape') or request.GET.get('codape')
    f_rcsno = request.GET.get('col_rcsno') or request.GET.get('rcsno')
    f_capital = request.GET.get('col_capital') or request.GET.get('capital')
    f_ca = request.GET.get('col_ca') or request.GET.get('ca')
    f_resultat = request.GET.get('col_resultat') or request.GET.get('resultat')

    col_agence = request.GET.get('col_agence')
    col_expl = request.GET.get('col_expl')
    col_datouv = request.GET.get('col_datouv')

    if f_filiale: queryset = queryset.filter(FILIALE=f_filiale)
    if f_agence: queryset = queryset.filter(AGENCE=f_agence)
    if f_expl: queryset = queryset.filter(EXPL=f_expl)

    if col_agence: queryset = queryset.filter(AGENCE__icontains=col_agence)
    if col_expl: queryset = queryset.filter(EXPL__icontains=col_expl)
    if col_datouv: queryset = queryset.filter(DATOUV__icontains=col_datouv)
    queryset = apply_datouv_period_filter(queryset, request)

    if f_lib_agence: queryset = queryset.filter(LIB_AGENCE__icontains=f_lib_agence)
    if f_client: queryset = queryset.filter(CLIENT__icontains=f_client)
    if f_idm: queryset = queryset.filter(IDM__icontains=f_idm)
    if f_agec: queryset = queryset.filter(AGEC__icontains=f_agec)
    if f_codape: queryset = queryset.filter(CODAPE__icontains=f_codape)
    if f_rcsno: queryset = queryset.filter(RCSNO__icontains=f_rcsno)
    if f_capital: queryset = queryset.filter(CAPITAL__icontains=f_capital)
    if f_ca: queryset = queryset.filter(CA__icontains=f_ca)
    if f_resultat: queryset = queryset.filter(RESULTAT__icontains=f_resultat)

                                                     
    notes = Notation.objects.filter(flux_stock='Flux')
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])

    if user.organe == "Chargé Client":
        notation = notation.filter(agent__filiale=user.filiale, agent__code_expl=user.code_expl)
    elif user.organe == "Directeur Agence":
        notation = notation.filter(agent__filiale=user.filiale, agent__agence=user.agence)
    elif user.organe not in ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "GUEST"]:
        notation = notation.filter(agent__filiale=user.filiale)

                                         
    filiale_list, agence_list, expl_list = get_filter_lists_pm(user, request)

                                                   
    query_params = request.GET.copy()
    if 'page' in query_params: del query_params['page']
    get_params = query_params.urlencode()

    paginator = Paginator(queryset.order_by('id'), 30)
    page_number = request.GET.get('page')
    try:
        objets_page = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        objets_page = paginator.page(1)

    context = {
        "donnees": objets_page,
        "get_params": get_params,
        "filiale_list": filiale_list,
        "agence_list": agence_list,
        "expl_list": expl_list,
        "notation": notation,
        "users_filiale": ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité'],
        "users_groupe": ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                         "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"],
    }

    return render(request, "non_rens_pm.html", context)

from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from .models import Kyc_pm                                    


@login_required
def export_csv_pm(request):
    user = request.user
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

                                                              
    donnees = Kyc_pm.objects.all()
    donnees = apply_kyc_field_config_filter(donnees, "pm")

                                                             
    f_filiale = request.GET.get("filiale")
    f_agence = request.GET.get("agence")
    f_expl = request.GET.get("expl")
    f_lib_agence = request.GET.get("lib_agence")
    f_client = request.GET.get("client")
    f_idm = request.GET.get("idm")
    f_agec = request.GET.get("agec")
    f_codape = request.GET.get("codape")
    f_rcsno = request.GET.get("rcsno")
    f_capital = request.GET.get("capital")
    f_ca = request.GET.get("ca")
    f_resultat = request.GET.get("resultat")

                                                        
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)

                                                                       
    if f_filiale:
        donnees = donnees.filter(FILIALE=f_filiale)
    if f_agence:
        donnees = donnees.filter(AGENCE=f_agence)
    if f_expl:
        donnees = donnees.filter(EXPL=f_expl)
    if f_lib_agence:
        donnees = donnees.filter(LIB_AGENCE__icontains=f_lib_agence)
    if f_client:
        donnees = donnees.filter(CLIENT__icontains=f_client)
    if f_idm:
        donnees = donnees.filter(IDM__icontains=f_idm)
    if f_agec:
        donnees = donnees.filter(AGEC__icontains=f_agec)
    if f_codape:
        donnees = donnees.filter(CODAPE__icontains=f_codape)
    if f_rcsno:
        donnees = donnees.filter(RCSNO__icontains=f_rcsno)
    if f_capital:
        donnees = donnees.filter(CAPITAL__icontains=f_capital)
    if f_ca:
        donnees = donnees.filter(CA__icontains=f_ca)
    if f_resultat:
        donnees = donnees.filter(RESULTAT__icontains=f_resultat)

                                                                           
    donnees = apply_datouv_period_filter(donnees, request)

                                   
    wb = Workbook()
    ws = wb.active
    ws.title = "Export KYC PM"

                                                               
    from kyc.context_processors import kyc_display_fields_processor
    ctx = kyc_display_fields_processor(request)
    display_fields = ctx.get('kyc_pm_display_fields', [])
    headers = [label for field, label in display_fields]
    ws.append(headers)

                            
    for d in donnees:
        row = []
        for field, label in display_fields:
            val = getattr(d, field, "")
            if field == "DATOUV":
                row.append(format_date_for_export(val))
            else:
                row.append(str(val) if val is not None else "")
        ws.append(row)

                                                                       
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 18

                                                              
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Export_KYC_PM_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response

from django.shortcuts import render
from django.db.models import Max
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

                                                   
                                      

                                     
ITEMS_PER_PAGE = 100                                         
from django.shortcuts import render
from django.db.models import Max
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Kyc_pp, Notation                                   


                                                           
def get_filtered_queryset(request):
    """Garantit que l'utilisateur ne voit que son périmètre autorisé."""
    user = request.user
    queryset = Kyc_pp.objects.all()

    if user.organe == "Chargé Client":
        queryset = queryset.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)

    elif user.organe == "Directeur Agence":
        queryset = queryset.filter(FILIALE=user.filiale, AGENCE=user.agence)

    elif user.organe in ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']:
        queryset = queryset.filter(FILIALE=user.filiale)

    elif user.organe in ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                         "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]:
                                                                               
        pass

    return queryset.order_by('id')


                                                              
def get_filter_lists(user, request):
    """Génère les options des menus déroulants selon les droits d'accès."""
    filiale_list, agence_list, expl_list = [], [], []
    base_queryset = Kyc_pp.objects.all()

    if user.organe == "Chargé Client":
        base_queryset = base_queryset.filter(FILIALE=user.filiale, AGENCE= user.agence, EXPL=user.code_expl)
        expl_list = [user.code_expl]

    elif user.organe == "Directeur Agence":
        base_queryset = base_queryset.filter(FILIALE=user.filiale, AGENCE=user.agence)
        expl_list = base_queryset.values_list("EXPL", flat=True).distinct()

    elif user.organe in ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']:
        base_queryset = base_queryset.filter(FILIALE=user.filiale)
        agence_list = base_queryset.values_list("AGENCE", flat=True).distinct()

        agence_filter = request.GET.get("agence")
        if agence_filter:
            expl_list = base_queryset.filter(AGENCE=agence_filter).values_list("EXPL", flat=True).distinct()
        else:
            expl_list = base_queryset.values_list("EXPL", flat=True).distinct()

    elif user.organe in ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                         "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]:
        filiale_list = Kyc_pp.objects.values_list("FILIALE", flat=True).distinct()

        filiale_filter = request.GET.get("filiale")
        agence_filter = request.GET.get("agence")

        if filiale_filter:
            base_queryset = base_queryset.filter(FILIALE=filiale_filter)
            agence_list = base_queryset.values_list("AGENCE", flat=True).distinct()
        if agence_filter:
            expl_list = base_queryset.filter(AGENCE=agence_filter).values_list("EXPL", flat=True).distinct()

    return filiale_list, agence_list, expl_list


                           
@login_required
def non_rens(request):
    user = request.user

                                                      
    queryset = get_filtered_queryset(request)
    
                                                               
    queryset = apply_kyc_field_config_filter(queryset, "pp")

                                                              
    f_filiale = request.GET.get('filiale')
    f_agence = request.GET.get('agence')
    f_expl = request.GET.get('expl')
    f_lib_agence = request.GET.get('col_lib_agence') or request.GET.get('lib_agence')
    f_client = request.GET.get('col_client') or request.GET.get('client')
    f_idp = request.GET.get('col_idp') or request.GET.get('idp')
    f_numid = request.GET.get('col_numid') or request.GET.get('numid')
    f_datnais = request.GET.get('col_datnais') or request.GET.get('datnais')
    f_paynais = request.GET.get('col_paynais') or request.GET.get('paynais')
    f_adresse = request.GET.get('col_adresse') or request.GET.get('adresse')
    f_codape = request.GET.get('col_codape') or request.GET.get('codape')
    f_profession = request.GET.get('col_profession') or request.GET.get('profession')
    f_salaire = request.GET.get('col_salaire') or request.GET.get('salaire')
    f_origine_rev = request.GET.get('col_origine_rev') or request.GET.get('origine_rev')
    f_datvalid = request.GET.get('col_datvalid') or request.GET.get('datvalid')
    f_tel = request.GET.get('col_tel') or request.GET.get('tel')

    col_agence = request.GET.get('col_agence')
    col_expl = request.GET.get('col_expl')
    col_datouv = request.GET.get('col_datouv')

    if f_filiale: queryset = queryset.filter(FILIALE=f_filiale)
    if f_agence: queryset = queryset.filter(AGENCE=f_agence)
    if f_expl: queryset = queryset.filter(EXPL=f_expl)

    if col_agence: queryset = queryset.filter(AGENCE__icontains=col_agence)
    if col_expl: queryset = queryset.filter(EXPL__icontains=col_expl)
    if col_datouv: queryset = queryset.filter(DATOUV__icontains=col_datouv)
    queryset = apply_datouv_period_filter(queryset, request)

    if f_lib_agence: queryset = queryset.filter(LIB_AGENCE__icontains=f_lib_agence)
    if f_client: queryset = queryset.filter(CLIENT__icontains=f_client)
    if f_idp: queryset = queryset.filter(IDP__icontains=f_idp)
    if f_numid: queryset = queryset.filter(NUMID__icontains=f_numid)
    if f_datnais: queryset = queryset.filter(DATNAIS__icontains=f_datnais)
    if f_paynais: queryset = queryset.filter(PAYNAIS__icontains=f_paynais)
    if f_adresse: queryset = queryset.filter(ADRESSE__icontains=f_adresse)
    if f_codape: queryset = queryset.filter(CODAPE__icontains=f_codape)
    if f_profession: queryset = queryset.filter(PROFESSION__icontains=f_profession)
    if f_salaire: queryset = queryset.filter(SALAIRE__icontains=f_salaire)
    if f_origine_rev: queryset = queryset.filter(ORIGINE_REV__icontains=f_origine_rev)
    if f_datvalid: queryset = queryset.filter(DATVALID__icontains=f_datvalid)
    if f_tel: queryset = queryset.filter(TEL__icontains=f_tel)

                                                           
    notes = Notation.objects.filter(flux_stock='Flux')
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])

    if user.organe == "Chargé Client":
        notation = notation.filter(agent__filiale=user.filiale, agent__agence=user.agence,agent__code_expl=user.code_expl)
    elif user.organe == "Directeur Agence":
        notation = notation.filter(agent__filiale=user.filiale, agent__agence=user.agence)
    elif user.organe not in ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "GUEST"]:
        notation = notation.filter(agent__filiale=user.filiale)

                                         
    filiale_list, agence_list, expl_list = get_filter_lists(user, request)

                   
    query_params = request.GET.copy()
    if 'page' in query_params: del query_params['page']
    get_params = query_params.urlencode()

    paginator = Paginator(queryset.order_by('id'), 30)
    page_number = request.GET.get('page')
    try:
        objets_page = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        objets_page = paginator.page(1)

    context = {
        "donnees": objets_page,
        "get_params": get_params,
        "filiale_list": filiale_list,
        "agence_list": agence_list,
        "expl_list": expl_list,
        'users_groupe': ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                         "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"],
        'users_filiale': ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité'],
        'notation': notation,
    }

    return render(request, "non_rens.html", context)


@login_required
def export_taux_completude_agence(request):
    return _export_taux_completude(request, "AGENCE", "Agence", "agence")


@login_required
def export_taux_completude_agent(request):
    return _export_taux_completude(request, "EXPL", "Agent", "agent")


def _export_taux_completude(request, group_field, group_header, slug):
    """Export Excel des taux de complétude par agence ou par agent (Flux /
    Stock, PP ou PM) selon group_field ("AGENCE" ou "EXPL").

    Calcul « maison » sur Kyc_pp / Kyc_pm : pour chaque groupe, taux de
    complétude de chaque champ configuré dans /champs_kyc (empty_check_fields,
    critère « vide » identique au filtre d'affichage de /non_rens), et taux de
    fiabilisation final = minimum des taux par champ. Flux = clients dont la
    DATOUV (ISO) tombe dans la fenêtre QualityFluxConfig ; Stock = toute la base.
    Réservé à tous les profils sauf Chargé Client et Directeur Agence.
    """
    import math

    user = request.user
    if getattr(user, "organe", None) in ("Chargé Client", "Directeur Agence"):
        return HttpResponse("Accès non autorisé.", status=403)

                                                                              
    type_param = (request.GET.get("type") or "").lower()
    client_types = [type_param] if type_param in ("pp", "pm") else ["pp", "pm"]

    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]
    is_group_user = user.organe in users_groupe or not getattr(user, "filiale", None)

    flux_start, flux_end = flux_datouv_window()

    wb = Workbook()
    wb.remove(wb.active)

                                                                        
                                                         
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_fill = PatternFill(start_color="09982E", end_color="09982E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    low_fill = PatternFill(start_color="D90A0A", end_color="D90A0A", fill_type="solid")
    low_font = Font(color="FFFFFF", bold=True)
    ok_fill = PatternFill(start_color="E8F5EC", end_color="E8F5EC", fill_type="solid")
    ok_font = Font(color="09982E", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for client_type in client_types:
        model = Kyc_pp if client_type == "pp" else Kyc_pm
        label = client_type.upper()

                                                                                 
                                                                            
        if is_group_user:
            filiale_param = (request.GET.get("filiale") or "").strip()
            if filiale_param:
                filiales = [filiale_param]
            else:
                filiales = sorted(f for f in model.objects.order_by()
                                  .values_list("FILIALE", flat=True).distinct() if f)
        else:
            filiales = [user.filiale]

        configs = list(KycFieldVisibilityConfig.objects.filter(client_type=client_type))

                                                                                
                                                                              
        fields_by_filiale = {}
        all_fields = []
        for filiale in filiales:
            config = _resolve_kyc_config(configs, filiale)
            fields = [f for f in (config.empty_check_fields if config else []) or []]
            fields_by_filiale[filiale] = fields
            for f in fields:
                if f not in all_fields:
                    all_fields.append(f)

        field_labels = {}
        for c in configs:
            for f, lbl in (c.field_labels or {}).items():
                if lbl:
                    field_labels.setdefault(f, lbl)

        show_filiale_col = len(filiales) > 1
        _write_taux_sheets(wb, model, label, filiales, fields_by_filiale, all_fields,
                           field_labels, show_filiale_col, group_field, group_header,
                           flux_start, flux_end, header_fill, header_font, low_fill,
                           low_font, ok_fill, ok_font, cell_border, center, math)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    types_label = "_".join(t.upper() for t in client_types)
    filename = f"Taux_completude_{slug}_{types_label}_{date_str}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _write_taux_sheets(wb, model, label, filiales, fields_by_filiale, all_fields,
                       field_labels, show_filiale_col, group_field, group_header,
                       flux_start, flux_end, header_fill, header_font, low_fill,
                       low_font, ok_fill, ok_font, cell_border, center, math):
    for mode, sheet_name in (("flux", f"Flux {label}"), ("stock", f"Stock {label}")):
        ws = wb.create_sheet(sheet_name)
        headers = (["Filiale"] if show_filiale_col else []) + [group_header, "Nbre clients concernés"]\
                  + [field_labels.get(f, f) for f in all_fields] + ["Taux de complétude"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = cell_border
                                                                       
        ws.freeze_panes = "B2"

        for filiale in filiales:
            fields = fields_by_filiale.get(filiale) or []
            if not fields:
                continue

            qs = model.objects.filter(FILIALE=filiale)
            if mode == "flux":
                qs = (qs.exclude(DATOUV="").exclude(DATOUV__isnull=True)
                        .filter(DATOUV__gte=flux_start, DATOUV__lte=flux_end))

            concerned_q = None
            for f in fields:
                fq = _kyc_field_empty_q(f)
                concerned_q = fq if concerned_q is None else concerned_q | fq

            annotations = {"total": Count("id"),
                           "concernes": Count("id", filter=concerned_q)}
            for f in fields:
                annotations[f"vide_{f}"] = Count("id", filter=_kyc_field_empty_q(f))

            rows = (qs.values(group_field).annotate(**annotations).order_by(group_field))

            for row in rows:
                total = row["total"] or 0
                if total == 0:
                    continue
                taux_champs = {}
                for f in fields:
                    taux_champs[f] = int(math.floor(100 - (row[f"vide_{f}"] / total) * 100))
                taux_final = min(taux_champs.values()) if taux_champs else None

                line = ([filiale] if show_filiale_col else []) + [row[group_field], row["concernes"]]
                for f in all_fields:
                    line.append(f"{taux_champs[f]}%" if f in taux_champs else "")
                line.append(f"{taux_final}%" if taux_final is not None else "N/A")
                ws.append(line)

                for cell in ws[ws.max_row]:
                    cell.border = cell_border
                taux_cell = ws.cell(row=ws.max_row, column=len(headers))
                taux_cell.alignment = center
                if taux_final is not None and taux_final < 90:
                    taux_cell.fill = low_fill
                    taux_cell.font = low_font
                elif taux_final is not None:
                    taux_cell.fill = ok_fill
                    taux_cell.font = ok_font

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18


@login_required
def export_csv_pp(request):
    user = request.user
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

                                                                       
    donnees = Kyc_pp.objects.all()
    donnees = apply_kyc_field_config_filter(donnees, "pp")

                                                          
                                                                          
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence , EXPL=user.code_expl)

    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
        expl_filter = request.GET.get("expl")
        if expl_filter:
            donnees = donnees.filter(EXPL=expl_filter)

    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
        agence_filter = request.GET.get("agence")
        expl_filter = request.GET.get("expl")
        if agence_filter:
            donnees = donnees.filter(AGENCE=agence_filter)
        if expl_filter:
            donnees = donnees.filter(EXPL=expl_filter)

    elif user.organe in users_groupe:
        filiale_filter = request.GET.get("filiale")
        agence_filter = request.GET.get("agence")
        expl_filter = request.GET.get("expl")

        if filiale_filter:
            donnees = donnees.filter(FILIALE=filiale_filter)
        if agence_filter:
            donnees = donnees.filter(AGENCE=agence_filter)
        if expl_filter:
            donnees = donnees.filter(EXPL=expl_filter)

    donnees = apply_datouv_period_filter(donnees, request)

                                                                 
    wb = Workbook()
    ws = wb.active
    ws.title = "Export KYC"

    from kyc.context_processors import kyc_display_fields_processor
    ctx = kyc_display_fields_processor(request)
    display_fields = ctx.get('kyc_pp_display_fields', [])
    headers = [label for field, label in display_fields]
    ws.append(headers)

    for d in donnees:
        row = []
        for field, label in display_fields:
            val = getattr(d, field, "")
            if field == "DATOUV" or field == "DATNAIS" or field == "DATVALID":
                row.append(format_date_for_export(val))
            else:
                row.append(str(val) if val is not None else "")
        ws.append(row)

    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Champs_non_renseignés_PP_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response




import json
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import TauxEvolution, TauxEvolution_filiale, Notation

def _dashboard_data_cache_version():
    """
    Versionne le cache avec les dates max des tables de taux.
    Quand une injection matinale met Ã  jour les données, la version change.
    """
    latest_filiale = TauxEvolution_filiale.objects.aggregate(last_date=Max('date')).get('last_date')
    latest_expl = TauxEvolution.objects.aggregate(last_date=Max('date')).get('last_date')
    rules_version = cache.get('quality_control_rules_version', 1)
    return f"{latest_filiale or 'none'}|{latest_expl or 'none'}|rules:{rules_version}"


_DASHBOARD_GROUP_ORGANES = frozenset([
    "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST",
])

                                                                             
                                                                          
_DASHBOARD_QS_PARAMS = ("mode", "periode", "filiale", "expl", "utilisateur")


def _dashboard_effective_scope(user):
    """Scope effectif (classe de rôle, filiale, agence, expl) d'un utilisateur.

    Deux utilisateurs de même classe et même périmètre voient exactement le même
    contenu dashboard : Conformité/Risques/DSI/... d'une même filiale partagent
    la même entrée de cache (l'organe exact ne change rien au rendu).
    """
    organe = str(getattr(user, "organe", "") or "")
    filiale = str(getattr(user, "filiale", "") or "")
    agence = str(getattr(user, "agence", "") or "")
    expl = str(getattr(user, "code_expl", "") or "")
    if organe in _DASHBOARD_GROUP_ORGANES:
        return "groupe", "", "", ""
    if organe == "Chargé Client":
        return "expl", filiale, agence, expl
    if organe == "Directeur Agence":
        return "agence", filiale, agence, ""
    return "filiale", filiale, "", ""


def _quality_rate_snapshot(quality_scope, applicability, flux_stock='stock'):
    """Dernier taux qualité précalculé (table TauxQualite) pour ce scope.

    On prend le snapshot le plus récent, même s'il date de la veille : un taux
    d'hier affiché instantanément vaut mieux qu'un rescan live de Kyc_pp
    (~80 s) quand le batch du matin n'est pas encore passé. Renvoie None si
    aucun snapshot n'existe pour ce scope (le repli live s'applique alors).
    flux_stock='flux' renvoie le taux du flux (fenêtre DATOUV configurée) ;
    pas de repli live pour le flux : None -> non affiché.
    """
    snapshot = (
        TauxQualite.objects.filter(
            filiale=quality_scope.get('filiale'),
            agence=quality_scope.get('agence'),
            expl=quality_scope.get('expl'),
            applicability=applicability,
            flux_stock=flux_stock,
        )
        .order_by('-date')
        .first()
    )
    if snapshot is None:
        return None
                                                                               
                                                                          
    if flux_stock == 'flux' and not snapshot.total:
        return None
    return snapshot.rate


def _flux_window_label():
    """Libellé court de la fenêtre flux configurée, pour affichage dashboard."""
    from kyc.models import QualityFluxConfig
    config = QualityFluxConfig.objects.filter(active=True).order_by('-updated_at').first()
    return 'mois précédent' if (config and config.flux_window == 'mois') else 'veille'


def _build_dashboard_cache_key(prefix, user, request, extra=""):
                                                                              
                                                                               
                                                                              
                                                                            
                                                              
    role, filiale, agence, expl = _dashboard_effective_scope(user)
    normalized_qs = "&".join(
        f"{param}={request.GET[param]}"
        for param in _DASHBOARD_QS_PARAMS
        if request.GET.get(param)
    )
    scope = "|".join([
        role,
        filiale,
        agence,
        expl,
        normalized_qs,
        extra,
        _dashboard_data_cache_version(),
    ])
    scope_hash = hashlib.md5(scope.encode("utf-8")).hexdigest()
                                                                             
                                                                              
    return f"dashboard:{prefix}:v2:{scope_hash}"


@login_required
def statistiques(request):
    context_cache_key = _build_dashboard_cache_key("statistiques", request.user, request)
    cached_context = cache.get(context_cache_key)
    if cached_context is not None:
        return render(request, 'statistiques.html', cached_context)

    user = request.user
                                                    
    user_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]

                           
    mode = request.GET.get('mode', 'Flux')
    is_stock = (mode == 'Stock')
    code_flux_stock = "S" if is_stock else "F"
    periode = request.GET.get('periode', 'journalier')
    periodes_valides = {'journalier', 'hebdomadaire', 'mensuel', 'annuel'}
    if periode not in periodes_valides:
        periode = 'mensuel'

                                   
    target_filiale = getattr(user, 'filiale', None)
    if user.organe in user_groupe:
        f_get = request.GET.get('filiale')
        if f_get:
            target_filiale = f_get
    else:
        target_filiale = user.filiale

                                 
    selected_user_filter = request.GET.get('utilisateur', '').strip()
    if user.organe == "Chargé Client":
        selected_expl = user.code_expl
    elif user.organe == "Directeur Agence":
        agents_de_lagence = ProfileV.objects.filter(filiale=user.filiale, agence=user.agence).values_list('code_expl', flat=True)
        req_expl = request.GET.get('expl') or selected_user_filter
        if req_expl in agents_de_lagence:
            selected_expl = req_expl
        else:
            selected_expl = None
    else:
        selected_expl = request.GET.get('expl') or selected_user_filter

                                 
    latest_filiale_data = TauxEvolution_filiale.objects.filter(filiale=target_filiale).order_by('-date').first()
    if is_stock:
        last_pp_fil = latest_filiale_data.stock_PP if latest_filiale_data else 0
        last_pm_fil = latest_filiale_data.stock_PM if latest_filiale_data else 0
    else:
        last_pp_fil = latest_filiale_data.flux_PP if latest_filiale_data else 0
        last_pm_fil = latest_filiale_data.flux_PM if latest_filiale_data else 0

                                                         
    base_qs_expl = TauxEvolution.objects.filter(
        flux_stock=code_flux_stock,
        filiale=target_filiale,
        expl=selected_expl
    ).order_by('date')

    latest_taux_date = (
        TauxEvolution.objects
        .filter(flux_stock=code_flux_stock, filiale=target_filiale)
        .aggregate(last_date=Max('date'))
        .get('last_date')
    )

    def build_period_key(d):
        if periode == 'journalier':
            return d
        if periode == 'hebdomadaire':
            year, week, _ = d.isocalendar()
            return (year, week)
        if periode == 'annuel':
            return d.year
        return (d.year, d.month)

    def format_period_label(key):
        if periode == 'journalier':
            return key.strftime('%d/%m/%Y')
        if periode == 'hebdomadaire':
            return f"S{key[1]}-{key[0]}"
        if periode == 'annuel':
            return str(key)
        return f"{key[1]:02d}/{key[0]}"

    def aggregate_by_period(queryset):
        # Valeurs 'N/A' / non numeriques ignorees => saut sur la courbe (pas de 0)
        grouped = {}
        for obj in queryset:
            if obj.taux is None:
                continue
            try:
                val = float(obj.taux)
            except (TypeError, ValueError):
                continue
            key = build_period_key(obj.date)
            bucket = grouped.setdefault(key, {"sum": 0.0, "count": 0})
            bucket["sum"] += val
            bucket["count"] += 1
        return {k: round(v["sum"] / v["count"], 2) for k, v in grouped.items() if v["count"] > 0}

    dict_expl_pp = aggregate_by_period(base_qs_expl.filter(pp_pm="P"))
    dict_expl_pm = aggregate_by_period(base_qs_expl.filter(pp_pm="M"))

    period_keys = sorted(set(dict_expl_pp.keys()) | set(dict_expl_pm.keys()))
    labels_chart = [format_period_label(k) for k in period_keys]
    labels_table = labels_chart[:]
    data_expl_pp = [dict_expl_pp.get(k) for k in period_keys]
    data_expl_pm = [dict_expl_pm.get(k) for k in period_keys]

    def _delta_last(series):
        vals = [v for v in series if v is not None]
        return round(vals[-1] - vals[-2], 2) if len(vals) > 1 else 0

    def _last_value(series):
        vals = [v for v in series if v is not None]
        return vals[-1] if vals else 0

    var_pp = _delta_last(data_expl_pp)
    var_pm = _delta_last(data_expl_pm)

                                        
    expl_queryset = TauxEvolution.objects.filter(filiale=target_filiale, flux_stock=code_flux_stock)
    if user.organe == "Directeur Agence":
        agents_de_lagence = ProfileV.objects.filter(filiale=user.filiale, agence=user.agence).values_list('code_expl', flat=True)
        expl_queryset = expl_queryset.filter(expl__in=agents_de_lagence)
    elif user.organe == "Chargé Client":
        expl_queryset = expl_queryset.filter(expl=user.code_expl)

    liste_expl = list(expl_queryset.values_list('expl', flat=True).distinct().order_by('expl'))
    profiles_by_expl = {
        p.code_expl: p
                                                                                
                                                                        
                                                                             
        for p in ProfileV.objects.filter(code_expl__in=liste_expl, filiale=target_filiale)
    }
    liste_expl_display = []
    for code in liste_expl:
        profile = profiles_by_expl.get(code)
        full_name = f"{getattr(profile, 'first_name', '')} {getattr(profile, 'last_name', '')}".strip() if profile else ""
        label = f"{code} - {full_name}" if full_name else code
        liste_expl_display.append({'code': code, 'label': label})

                             
                                                                                
                                                                               
    agent_info = (
        ProfileV.objects.filter(filiale=target_filiale, code_expl=selected_expl).first()
        if selected_expl else None
    )
    notation_obj = Notation.objects.filter(agent__code_expl=selected_expl, flux_stock=mode).order_by('-date_notation').first()

                           
    if user.organe in user_groupe:
        liste_filiales = TauxEvolution_filiale.objects.values_list('filiale', flat=True).distinct().order_by('filiale')
    else:
        liste_filiales = [user.filiale] if user.filiale else []

    quality_scope = evaluate_data_quality_scope(user)
    if selected_expl:
        quality_scope = {
            'filiale': target_filiale,
            'agence': getattr(agent_info, 'agence', None) if agent_info else None,
            'expl': selected_expl,
            'label': f"Agent {selected_expl}",
        }
    rules_version = cache.get('quality_control_rules_version', 1)

    def compute_quality_rate_by_typology(applicability):
                                                                            
                                                                               
                                                                               
                                           
        snapshot_rate = _quality_rate_snapshot(quality_scope, applicability)
        if snapshot_rate is not None:
            return snapshot_rate

        scope_signature = (
            f"{quality_scope.get('filiale')}|{quality_scope.get('agence')}|"
            f"{quality_scope.get('expl')}|{applicability}"
        )
        scope_hash = hashlib.md5(scope_signature.encode('utf-8')).hexdigest()
        cache_key = f"quality_control:dashboard_rate:v{rules_version}:{scope_hash}"
        cached_rate = cache.get(cache_key)
        if cached_rate is not None:
            return cached_rate

        rules = list(DataQualityRule.objects.filter(active=True, applicability=applicability))
        total_ok = 0
        total_evaluated = 0
        for rule in rules:
            stat = evaluate_data_quality_rule(
                rule,
                filiale=quality_scope.get('filiale'),
                agence=quality_scope.get('agence'),
                expl=quality_scope.get('expl'),
            )
            total_ok += stat.get('ok_count', 0)
            total_evaluated += stat.get('total', 0)

        rate = round(total_ok / total_evaluated * 100, 1) if total_evaluated else 0
        cache.set(cache_key, rate, timeout=3600)
        return rate

    quality_rate_pp = compute_quality_rate_by_typology('PP')
    quality_rate_pm = compute_quality_rate_by_typology('PM')

    context = {
        'mode': mode,
        'is_stock': is_stock,
        'periode': periode,
        'selected_filiale': target_filiale,
        'selected_expl': selected_expl,
        'selected_user_filter': selected_expl,
        'agent_nom': (
            (f"{agent_info.first_name} {agent_info.last_name}".strip() or agent_info.code_expl or agent_info.username)
            if agent_info else selected_expl
        ),
        'agent_note': notation_obj.note if notation_obj else "N/A",
        'labels_json': json.dumps(labels_chart),
        'data_expl_pp': json.dumps(data_expl_pp),
        'data_expl_pm': json.dumps(data_expl_pm),
        'last_pp_expl': _last_value(data_expl_pp),
        'last_pm_expl': _last_value(data_expl_pm),
        'last_pp_fil': last_pp_fil,
        'last_pm_fil': last_pm_fil,
        'var_pp': var_pp,
        'var_pm': var_pm,
        'historique': list(reversed(list(zip(labels_table, data_expl_pp, data_expl_pm)))),
        'liste_expl': liste_expl,
        'liste_expl_display': liste_expl_display,
        'user_groupe': user_groupe,
        'liste_filiales': liste_filiales,
        'quality_rate_pp': quality_rate_pp,
        'quality_rate_pm': quality_rate_pm,
        'quality_rate_pp_flux': _quality_rate_snapshot(quality_scope, 'PP', flux_stock='flux'),
        'quality_rate_pm_flux': _quality_rate_snapshot(quality_scope, 'PM', flux_stock='flux'),
        'flux_window_label': _flux_window_label(),
        'quality_scope_label': quality_scope.get('label'),
        'latest_taux_date': latest_taux_date,
    }
    cache.set(context_cache_key, context, timeout=60 * 60 * 12)
    return render(request, 'statistiques.html', context)
def export_stats_pp(request):
                                                          
    user = request.user
    organe = user.organe
    filiale = user.filiale
    expl_user = user.expl

    if organe in ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                  "Contrôle Permanent Groupe", "PASS", "GUEST"]:
        qs = TauxEvolution.objects.all()
    elif organe in ["Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']:
        qs = TauxEvolution.objects.filter(filiale=filiale)
    elif organe == "Directeur Agence":
        qs = TauxEvolution.objects.filter(agence=user.agence)
    elif organe == "Chargé Client":
        qs = TauxEvolution.objects.filter(expl=expl_user)
    else:
        qs = TauxEvolution.objects.none()

                                          
    selected_expl = request.GET.get('expl')
    if selected_expl:
        qs = qs.filter(expl=selected_expl)

    data = list(qs.values('filiale', 'agence', 'expl', 'date', 'taux'))
    df = pd.DataFrame(data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="rapport_taux.xlsx"'
    df.to_excel(response, index=False)
    return response


@login_required
def daterev_ppe(request):
           
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

    user = request.user

                
    periode_param = request.GET.get("periode", "")
    filiale_param = request.GET.get("filiale", "")
    agence_param = request.GET.get("agence", "")
    expl_param = request.GET.get("expl", "")

    base_qs = DATEREV.objects.all().filter(DATEREV__isnull=False, PPE='O')

    is_group_user = (user.organe in users_groupe) or (user.filiale in ["BOA Group", "BOA GROUP"]) or (not user.filiale)
    if not is_group_user:
        if getattr(user, "organe", "") == "Chargé Client":
            base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
        elif user.organe == "Directeur Agence":
            base_qs = base_qs.filter(FILIALE=user.filiale, AGENCE=user.agence)
        elif user.organe in users_filiale:
            base_qs = base_qs.filter(FILIALE=user.filiale)

    today = date.today()
    qs_period = base_qs
    if periode_param == "today":
        qs_period = qs_period.filter(DATEREV__lte=today)
    elif periode_param == "3m":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=90))
    elif periode_param == "6m":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=180))
    elif periode_param == "1y":
        qs_period = qs_period.filter(DATEREV__gte=today, DATEREV__lte=today + timedelta(days=365))

    can_pick_filiale = is_group_user

    selected_filiale = filiale_param if can_pick_filiale else getattr(user, "filiale", "")

    filiales_opts = qs_period.values_list("FILIALE", flat=True).distinct().order_by("FILIALE")

    qs_filiale = qs_period
    if selected_filiale:
        qs_filiale = qs_filiale.filter(FILIALE=selected_filiale)

    can_pick_agence = (user.organe in users_groupe) or (user.organe in users_filiale) or (
            user.organe == "Directeur Agence")

    if user.organe == "Directeur Agence":
        selected_agence = getattr(user, "agence", "")
    else:
        selected_agence = agence_param

    agences_opts = qs_filiale.values_list("AGENCE", flat=True).distinct().order_by("AGENCE")

    qs_agence = qs_filiale
    if selected_agence:
        qs_agence = qs_agence.filter(AGENCE=selected_agence)

    can_pick_expl = (user.organe in users_groupe) or (user.organe in users_filiale) or (
            user.organe == "Directeur Agence")

    if getattr(user, "organe", "") == "Chargé Client":
        selected_expl = getattr(user, "code_expl", "")
    else:
        selected_expl = expl_param

    exploitants_opts = qs_agence.values_list("EXPL", flat=True).distinct().order_by("EXPL")

    donnees = qs_agence
    if selected_expl:
        donnees = donnees.filter(EXPL=selected_expl)
    count_risque_non_eleve = donnees.exclude(RISQUE="Risque eleve").count()

    context = {
        "donnees": donnees.order_by("FILIALE", "AGENCE", "EXPL", "CLIENT"),
        "total_count": donnees.count(),                               
        "count_risque_non_eleve": count_risque_non_eleve,                      
                 
        "filiales": filiales_opts,
        "agences": agences_opts,
        "exploitants": exploitants_opts,

                              
        "periode": periode_param,
        "filiale_param": selected_filiale,
        "agence_param": selected_agence,
        "expl_param": selected_expl,

                                                     
        "users_groupe": users_groupe,
        "users_filiale": users_filiale,

                                      
        "can_pick_filiale": can_pick_filiale,
        "can_pick_agence": can_pick_agence,
        "can_pick_expl": can_pick_expl,
    }
    return render(request, 'daterev_ppe.html', context)



from django.shortcuts import render
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Max

@login_required
def non_anom(request):
    user = request.user
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau", 'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
                    "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"]

                                                                     
                              
                                                                     
    notes = Notation.objects.filter(flux_stock='Flux')
    latest_notes = notes.values('agent').annotate(latest_date=Max('date_notation'))
    notation = notes.filter(date_notation__in=[n['latest_date'] for n in latest_notes])

    if hasattr(user, 'filiale') and hasattr(user, 'code_expl'):
        notation = notation.filter(agent__filiale=user.filiale, agent__code_expl=user.code_expl)

                                                                     
                    
                                                                     
    filiale_filter = request.GET.get("filiale")
    agence_filter = request.GET.get("agence")
    expl_filter = request.GET.get("expl")

    filiale_list = []
    agence_list = []
    expl_list = []

                                                                     
                             
                                                                     
    is_group_user = (user.organe in users_groupe) or (user.filiale in ["BOA Group", "BOA GROUP"]) or (not user.filiale)

                                                                                
                                                                                    
    def _kyc_distinct(field, filiale=None, agence=None):
        import hashlib as _h
        from django.core.cache import cache as _c
        from django.utils import timezone as _tz
                                                                            
                                                                             
        _sig = f"{field}:{filiale or ''}:{agence or ''}:{_tz.localdate().isoformat()}"
        ck = f"non_anom_dd:{_h.md5(_sig.encode('utf-8')).hexdigest()}"
        cached = _c.get(ck)
        if cached is not None:
            return cached
        vals = set()
        for model in (Kyc_pp, Kyc_pm):
            qs = model.objects.all()
            if filiale:
                qs = qs.filter(FILIALE=filiale)
            if agence:
                qs = qs.filter(AGENCE=agence)
            vals.update(qs.exclude(**{field: ""}).exclude(**{f"{field}__isnull": True})
                          .values_list(field, flat=True).distinct())
        result = sorted(v for v in vals if v)
        _c.set(ck, result, 3600)
        return result

    if is_group_user:
        from kyc.models import DataQualityRule
        from kyc.forms import DataQualityRuleForm
        rule_filiales = set()
        for rule in DataQualityRule.objects.filter(active=True):
            parsed = DataQualityRuleForm._parse_filiales(rule.filiale)
            for f in parsed:
                if f:
                    rule_filiales.add(f)
        filiale_list = sorted(rule_filiales | set(_kyc_distinct("FILIALE")))
        if filiale_filter:
            agence_list = _kyc_distinct("AGENCE", filiale=filiale_filter)
            expl_list = _kyc_distinct("EXPL", filiale=filiale_filter, agence=agence_filter or None)

    elif user.organe == "Chargé Client":
                                                                                  
        agence_list = [user.agence] if user.agence else []
        expl_list = [user.code_expl] if user.code_expl else []

    elif user.organe == "Directeur Agence":
                                                    
        agence_list = [user.agence] if user.agence else []
        expl_list = _kyc_distinct("EXPL", filiale=user.filiale, agence=user.agence)

    elif user.organe in users_filiale:
                                             
        filiale_list = [user.filiale]
        agence_list = _kyc_distinct("AGENCE", filiale=user.filiale)
        expl_list = _kyc_distinct("EXPL", filiale=user.filiale, agence=agence_filter or None)

    import urllib.parse
    import hashlib
    from datetime import datetime
    from django.utils import timezone
    from django.core.cache import cache
    from kyc.models import DataQualityRule

                                                        
    def get_rule_failures(rule, queryset_eval):
        client_fields = ['CLIENT', 'EXPL', 'FILIALE', 'AGENCE']
        failures = []
        today_date = timezone.localdate()

        def safe_parse_date(value):
            if not value: return None
            if hasattr(value, 'date'): return value.date()
            if isinstance(value, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y%m%d"):
                    try:
                        return datetime.strptime(value.strip(), fmt).date()
                    except ValueError:
                        continue
            return None

        def calculate_age(birth_date_str):
            parsed = safe_parse_date(birth_date_str)
            if not parsed: return None
            return today_date.year - parsed.year - ((today_date.month, today_date.day) < (parsed.month, parsed.day))

        if rule.control_type == 'simple':
            param = (rule.parameter or '').strip().lower()
            field_name = rule.field_name
            rows = queryset_eval.values(*client_fields, field_name).iterator(chunk_size=2000)
            
            if not param or param == 'existence':
                for row in rows:
                    val = row.get(field_name)
                    if val is None or str(val).strip() == "":
                        failures.append({
                            'client': row.get('CLIENT', ''),
                            'filiale': row.get('FILIALE', ''),
                            'agence': row.get('AGENCE', ''),
                            'expl': row.get('EXPL', ''),
                            'values': [str(val or '')]
                        })
            elif param.isdigit() or (param.startswith('len') or param.startswith('long')):
                import re
                match = re.search(r'\d+', param)
                target_len = int(match.group()) if match else int(param)
                for row in rows:
                    val = str(row.get(field_name) or '')
                    if len(val) != target_len:
                        failures.append({
                            'client': row.get('CLIENT', ''),
                            'filiale': row.get('FILIALE', ''),
                            'agence': row.get('AGENCE', ''),
                            'expl': row.get('EXPL', ''),
                            'values': [val]
                        })
            else:
                target_val = rule.parameter
                for row in rows:
                    val = row.get(field_name)
                    if str(val or '').strip() != str(target_val or '').strip():
                        failures.append({
                            'client': row.get('CLIENT', ''),
                            'filiale': row.get('FILIALE', ''),
                            'agence': row.get('AGENCE', ''),
                            'expl': row.get('EXPL', ''),
                            'values': [str(val or '')]
                        })
                        
        elif rule.control_type == 'composite':
            conditions = rule.conditions.all()
            cond_fields = [c.field_name for c in conditions]
            unique_cond_fields = list(dict.fromkeys(cond_fields))
            fields_to_fetch = list(client_fields) + unique_cond_fields
            rows = queryset_eval.values(*fields_to_fetch).iterator(chunk_size=2000)
            
            for row in rows:
                if _dq_conditions_flag(conditions, lambda f: row.get(f, ''), today_date, safe_parse_date, calculate_age):
                    failures.append({
                        'client': row.get('CLIENT', ''),
                        'filiale': row.get('FILIALE', ''),
                        'agence': row.get('AGENCE', ''),
                        'expl': row.get('EXPL', ''),
                        'values': [str(row.get(f, '') or '') for f in unique_cond_fields]
                    })
        return failures

                           
    rules_qs = DataQualityRule.objects.filter(active=True).prefetch_related('conditions')
    
                                 
    if is_group_user:
        target_filiale = filiale_filter
    else:
        target_filiale = user.filiale

    if target_filiale:
        rules_qs = rules_qs.filter(Q(filiale__icontains=f"|{target_filiale}|") | Q(filiale=""))

                                 
    q = request.GET.get('q', '').strip()
    if q:
        rules_qs = rules_qs.filter(
            Q(name__icontains=q) |
            Q(field_name__icontains=q) |
            Q(parameter__icontains=q)
        )
    rules_qs = rules_qs.order_by('id')

                         
    if is_group_user:
        eval_filiale = filiale_filter
        eval_agence = agence_filter
        eval_expl = expl_filter
    elif user.organe == "Chargé Client":
        eval_filiale = user.filiale
        eval_agence = user.agence
        eval_expl = user.code_expl
    elif user.organe == "Directeur Agence":
        eval_filiale = user.filiale
        eval_agence = user.agence
        eval_expl = expl_filter
    elif user.organe in users_filiale:
        eval_filiale = user.filiale
        eval_agence = agence_filter
        eval_expl = expl_filter
    else:
        eval_filiale = filiale_filter
        eval_agence = agence_filter
        eval_expl = expl_filter

    rules_version = cache.get('quality_control_rules_version', 1)
    data_refresh_bucket = timezone.localdate().isoformat()
    cache_ttl_seconds = 86400

    from kyc.models import get_rule_number_map
    rule_number_map = get_rule_number_map()
    rules_with_stats = []
    for rule in rules_qs:
        rule_num = rule_number_map.get((rule.name or '').strip(), rule.id)
        rule._rule_number = rule_num
        rule_eval_filiale = _rule_eval_filiale(rule, eval_filiale)
        non_anom_signature = (
            f"{rule.id}|{rule.name}|{rule.applicability}|{rule.field_name}|"
            f"{rule.control_type}|{rule.parameter}|{rule.filiale}|"
            f"{rule_eval_filiale}|{eval_agence}|{eval_expl}"
        )
        non_anom_key = f"quality_control:non_anom:v{rules_version}:d{data_refresh_bucket}:{hashlib.md5(non_anom_signature.encode('utf-8')).hexdigest()}"
        
        stat = cache.get(non_anom_key)
        if stat is None:
            stat = _evaluate_data_quality_rule_scoped(rule, filiale=rule_eval_filiale, agence=eval_agence, expl=eval_expl)
            cache.set(non_anom_key, stat, timeout=cache_ttl_seconds)

        total_eval = stat.get('total', 0)
        stat['compliance_rate'] = compliance_rate_floor(stat.get('ok_count', 0), total_eval, stat.get('fail_count', 0))

        from kyc.forms import DataQualityRuleForm
        parsed_filiales = DataQualityRuleForm._parse_filiales(rule.filiale)
        if user.organe not in users_groupe and user.filiale:
            if not parsed_filiales or user.filiale in parsed_filiales:
                parsed_filiales = [user.filiale]
            else:
                parsed_filiales = []

        if parsed_filiales:
            visible_filiales = parsed_filiales[:3]
            hidden_count = max(0, len(parsed_filiales) - 3)
            display_str = ", ".join(parsed_filiales)
        else:
            visible_filiales = ["Toutes les filiales"]
            hidden_count = 0
            display_str = "Toutes les filiales"

        filiales_summary = {
            'display': display_str,
            'visible': visible_filiales,
            'hidden_count': hidden_count,
        }

        rules_with_stats.append({
            'rule': rule,
            'stat': stat,
            'rule_number': rule_num,
            'filiales_summary': filiales_summary,
        })

    selected_rule_id = request.GET.get("rule")
    selected_rule = None
    selected_rule_number = None
    show_rule_modal = False
    failures_page = None
    selected_rule_conditions = []
    selected_rule_filiales_display = ""
    failure_columns = []
    
    failure_client_filter = request.GET.get("failure_client", "")
    failure_filiale_filter = request.GET.get("failure_filiale", "")
    failure_agence_filter = request.GET.get("failure_agence", "")
    failure_expl_filter = request.GET.get("failure_expl", "")
    failure_message = ""

    if selected_rule_id:
        try:
            selected_rule = DataQualityRule.objects.get(pk=selected_rule_id, active=True)
            selected_rule_number = rule_number_map.get((selected_rule.name or '').strip(), selected_rule.pk)
            selected_rule._rule_number = selected_rule_number
            show_rule_modal = True
            selected_rule_conditions = selected_rule.conditions.all()
            
            from kyc.forms import DataQualityRuleForm
            parsed_filiales = DataQualityRuleForm._parse_filiales(selected_rule.filiale)
            if user.organe not in users_groupe and user.filiale:
                if not parsed_filiales or user.filiale in parsed_filiales:
                    parsed_filiales = [user.filiale]
                else:
                    parsed_filiales = []
            selected_rule_filiales_display = ", ".join(parsed_filiales) if parsed_filiales else "Toutes les filiales"
            
            if selected_rule.control_type == 'simple':
                failure_columns = [{
                    'name': selected_rule.field_name.upper(),
                    'param': f'failure_{selected_rule.field_name}',
                    'filter_value': request.GET.get(f'failure_{selected_rule.field_name}', '')
                }]
            else:
                cond_fields = [c.field_name for c in selected_rule_conditions]
                unique_cond_fields = list(dict.fromkeys(cond_fields))
                failure_columns = [{
                    'name': f.upper(),
                    'param': f'failure_{f}',
                    'filter_value': request.GET.get(f'failure_{f}', '')
                } for f in unique_cond_fields]
                
            model = Kyc_pp if selected_rule.applicability == 'PP' else Kyc_pm
            queryset_eval = model.objects.all()
            
            rule_eval_filiale = _rule_eval_filiale(selected_rule, eval_filiale)
            if rule_eval_filiale and rule_eval_filiale != 'GROUPE':
                queryset_eval = queryset_eval.filter(FILIALE=rule_eval_filiale)
            elif parsed_filiales:
                queryset_eval = queryset_eval.filter(FILIALE__in=parsed_filiales)
                
            if eval_agence:
                queryset_eval = queryset_eval.filter(AGENCE=eval_agence)
            if eval_expl:
                queryset_eval = queryset_eval.filter(EXPL=eval_expl)
                
            failures = get_rule_failures(selected_rule, queryset_eval)
            
            if failure_client_filter:
                failures = [f for f in failures if failure_client_filter.lower() in str(f['client']).lower()]
            if failure_filiale_filter:
                failures = [f for f in failures if failure_filiale_filter.lower() in str(f['filiale']).lower()]
            if failure_agence_filter:
                failures = [f for f in failures if failure_agence_filter.lower() in str(f['agence']).lower()]
            if failure_expl_filter:
                failures = [f for f in failures if failure_expl_filter.lower() in str(f['expl']).lower()]

            for i, col in enumerate(failure_columns):
                val_filter = col['filter_value'].strip()
                if val_filter:
                    failures = [f for f in failures if val_filter.lower() in str(f['values'][i]).lower()]
            
            paginator = Paginator(failures, 15)
            page_number = request.GET.get('page')
            try:
                failures_page = paginator.page(page_number)
            except PageNotAnInteger:
                failures_page = paginator.page(1)
            except EmptyPage:
                failures_page = paginator.page(paginator.num_pages)
                
        except DataQualityRule.DoesNotExist:
            failure_message = "La règle spécifiée n'existe pas ou est inactive."

                            
    base_params = {}
    if q: base_params['q'] = q
    if filiale_filter: base_params['filiale'] = filiale_filter
    if agence_filter: base_params['agence'] = agence_filter
    if expl_filter: base_params['expl'] = expl_filter
    base_querystring = urllib.parse.urlencode(base_params)

    pagination_params = dict(request.GET.items())
    pagination_params.pop('page', None)
    pagination_querystring = urllib.parse.urlencode(pagination_params)

    export_querystring = pagination_querystring

    context = {
        "filiale_list": filiale_list,
        "agence_list": agence_list,
        "expl_list": expl_list,
        "users_groupe": users_groupe,
        "users_filiale": users_filiale,
        "notation": notation,

        "rules": rules_with_stats,
        "total_rules": len(rules_with_stats),
        "pp_rules_count": sum(1 for item in rules_with_stats if item['rule'].applicability == 'PP'),
        "pm_rules_count": sum(1 for item in rules_with_stats if item['rule'].applicability == 'PM'),
        "total_failures": sum(item['stat'].get('fail_count', 0) for item in rules_with_stats),
        "rule_search": q,
        "can_pick_filiale": is_group_user,
        "can_pick_agence": user.organe in users_filiale or is_group_user,
        "is_charge_client": user.organe == "Chargé Client",
        "is_group_user": is_group_user,
        "show_rule_modal": show_rule_modal,
        "selected_rule": selected_rule,
        "selected_rule_number": selected_rule_number,
        "selected_rule_conditions": selected_rule_conditions,
        "selected_rule_filiales_display": selected_rule_filiales_display,
        "failure_columns": failure_columns,
        "failures_page": failures_page,
        "failure_client_filter": failure_client_filter,
        "failure_filiale_filter": failure_filiale_filter,
        "failure_agence_filter": failure_agence_filter,
        "failure_expl_filter": failure_expl_filter,
        "failure_message": failure_message,
        "base_querystring": base_querystring,
        "pagination_querystring": pagination_querystring,
        "export_querystring": export_querystring,
    }

    return render(request, "non_anom.html", context)

@login_required
def devise(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau", 'Risques', 'DAI', 'Qualité']
    users_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]

    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

                                                                          
                                                                                  
    devise_obj = Devise.objects.filter(filiale=user.filiale).first()
    devise_valeur = devise_obj.devise if devise_obj else None                                                         

                                                                              
    donnees = Kyc_pp.objects.exclude(DEVISE=devise_valeur).exclude(DEVISE__isnull=True).exclude(DEVISE="").exclude(DEVISE="NA")

                                                
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
                                                                              

                                     
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)
    donnees = _apply_pp_header_filters(donnees, request, include_devise=True)
    donnees = apply_datouv_period_filter(donnees, request)

                                                             
    filiales = donnees.values_list('FILIALE', flat=True).distinct()
    agences = donnees.values_list('AGENCE', flat=True).distinct()
    exploitants = donnees.values_list('EXPL', flat=True).distinct()
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']

                       
    ITEMS_PER_PAGE = 25
    paginator = Paginator(donnees.order_by('id'), ITEMS_PER_PAGE)
    page_number = request.GET.get('page')

    try:
        objets_page = paginator.page(page_number)
    except PageNotAnInteger:
        objets_page = paginator.page(1)
    except EmptyPage:
        objets_page = paginator.page(paginator.num_pages)

                 
    total_devise = donnees.count()
    missing_devise = get_incomplete_clients_queryset(donnees, 'pp').count()
    complete_devise = max(0, total_devise - missing_devise)
    compliance_rate = round((complete_devise / total_devise) * 100, 1) if total_devise > 0 else 100.0

                            
    devise_counts = list(
        donnees.values('DEVISE')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    devise_items = []
    for dc in devise_counts:
        dev_code = dc['DEVISE'] or "Non renseignée"
        pct = round((dc['count'] / total_devise) * 100, 1) if total_devise > 0 else 0.0
        devise_items.append({
            'label': dev_code,
            'value': dc['count'],
            'suffix': f'({pct}%)'
        })

    export_params = request.GET.copy()
    export_params['incompletes'] = '1'
    export_incomplets_url = f"{reverse('export_devise_pp')}?{export_params.urlencode()}"

    kpi_cards = [
        {
            'tone': 'emerald',
            'label': 'Répartition par Devise',
            'value': total_devise,
            'subtitle': 'Comptes en devise étrangère',
            'show_modal': True,
            'items': devise_items
        },
        {
            'tone': 'red',
            'label': 'Comptes Incomplets',
            'value': missing_devise,
            'subtitle': 'Dossiers avec données manquantes',
            'export_url': export_incomplets_url,
        },
        {
            'tone': 'blue',
            'label': 'Taux de complétude',
            'value': compliance_rate,
            'suffix': '%',
            'subtitle': 'Dossiers complets / total',
        }
    ]

    context = {
        "donnees": objets_page,
        "devise_filiale": devise_valeur,                                           
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
        'get_params': query_params.urlencode(),
        'kpi_cards': kpi_cards,
    }

    return render(request, 'devise.html', context)


@login_required
def export_devise_pp(request):
    user = request.user

    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                    "Contrôle Permanent Groupe", "PASS", "GUEST"]

                                                          
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

    devise_obj = Devise.objects.filter(filiale=user.filiale).first()
    devise_valeur = devise_obj.devise if devise_obj else None

                                                                            
    donnees = Kyc_pp.objects.exclude(DEVISE=devise_valeur).exclude(DEVISE__isnull=True).exclude(DEVISE="").exclude(DEVISE="NA")

                                                                      
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    elif user.organe in users_groupe:
        pass                              

                                                       
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)
    donnees = _apply_pp_header_filters(donnees, request, include_devise=True)
    donnees = apply_datouv_period_filter(donnees, request)

    if request.GET.get('incompletes') == '1':
        donnees = get_incomplete_clients_queryset(donnees, 'pp')

                            

                                
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comptes Devise PP"                         

              
    headers = ["FILIALE", "AGENCE", "EXPL", "CLIENT", "DEVISE", "CODAPE", "IDP", "PAYNAIS", "PROFESSION", "ADRESSE", "PAYS_RESID",
               "NUMID", "SALAIRE", "ORIGINE_REV", "DATVALID", "TEL"]
    ws.append(headers)

             
    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.DEVISE, d.CODAPE, d.IDP,
            d.PAYNAIS, d.PROFESSION, d.ADRESSE, d.PAYS_RESID, d.NUMID, d.SALAIRE, d.ORIGINE_REV, d.DATVALID, d.TEL
        ])

                                                 
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

                              
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    if request.GET.get('incompletes') == '1':
        filename = f"Comptes_en_devise_PP_incomplets_{date_str}.xlsx"
    else:
        filename = f"Comptes_en_devise_PP_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def devise_pm(request):
    roles_exclus = ["Chargé Client"]
    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]
    user = request.user
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')

                                          
    devise_obj = Devise.objects.filter(filiale=user.filiale).first()
    devise_valeur = devise_obj.devise if devise_obj else None

                                                                              
    donnees = Kyc_pm.objects.exclude(DEVISE=devise_valeur).exclude(DEVISE__isnull=True).exclude(DEVISE="").exclude(DEVISE="NA")

                                                
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    elif user.organe in users_groupe:
        pass                              

                                     
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)
    donnees = apply_datouv_period_filter(donnees, request)

                                                 
                                                                                                  

                                         
    filiales = donnees.values_list('FILIALE', flat=True).distinct()
    agences = donnees.values_list('AGENCE', flat=True).distinct()
    exploitants = donnees.values_list('EXPL', flat=True).distinct()

                                   
                                                                     
    queryset = donnees.order_by('id')

                               
                                                   
    ITEMS_PER_PAGE = 25
    paginator = Paginator(queryset, ITEMS_PER_PAGE)

    page_number = request.GET.get('page')
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']

    try:
        objets_page = paginator.page(page_number)
    except PageNotAnInteger:
                                                                  
        objets_page = paginator.page(1)
    except EmptyPage:
                                                                
        objets_page = paginator.page(paginator.num_pages)

                    
    total_devise = donnees.count()
    missing_devise = get_incomplete_clients_queryset(donnees, 'pm').count()
    complete_devise = max(0, total_devise - missing_devise)
    compliance_rate = round((complete_devise / total_devise) * 100, 1) if total_devise > 0 else 100.0

                            
    devise_counts = list(
        donnees.values('DEVISE')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    devise_items = []
    for dc in devise_counts:
        dev_code = dc['DEVISE'] or "Non renseignée"
        pct = round((dc['count'] / total_devise) * 100, 1) if total_devise > 0 else 0.0
        devise_items.append({
            'label': dev_code,
            'value': dc['count'],
            'suffix': f'({pct}%)'
        })

    export_params = request.GET.copy()
    export_params['incompletes'] = '1'
    export_incomplets_url = f"{reverse('export_devise_pm')}?{export_params.urlencode()}"

    kpi_cards = [
        {
            'tone': 'emerald',
            'label': 'Répartition par Devise',
            'value': total_devise,
            'subtitle': 'Comptes PM en devise',
            'show_modal': True,
            'items': devise_items
        },
        {
            'tone': 'red',
            'label': 'Comptes PM Incomplets',
            'value': missing_devise,
            'subtitle': 'Dossiers avec données manquantes',
            'export_url': export_incomplets_url,
        },
        {
            'tone': 'blue',
            'label': 'Taux de complétude',
            'value': compliance_rate,
            'suffix': '%',
            'subtitle': 'Dossiers complets / total',
        }
    ]

    context = {
                                                      
        "donnees": objets_page,
        "devise_filiale": devise_valeur,
        'filiales': filiales,
        'agences': agences,
        'exploitants': exploitants,
        'roles_exclus': roles_exclus,
        'users_groupe': users_groupe,
        'users_filiale': users_filiale,
        'get_params': query_params.urlencode(),
        'kpi_cards': kpi_cards,
    }

    return render(request, 'devise_pm.html', context)


@login_required
def export_devise_pm(request):
    user = request.user

    users_filiale = ["DSI", "Conformité", "Contrôle Permanent", "Directeur Réseau",'Risques', 'DAI', 'Qualité']
    users_groupe = ["Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone", "Conformité Groupe",
                    "Contrôle Permanent Groupe", "PASS", "GUEST"]

                                                          
    filiale_param = request.GET.get('filiale', '')
    agence_param = request.GET.get('agence', '')
    expl_param = request.GET.get('expl', '')


    devise_obj = Devise.objects.filter(filiale=user.filiale).first()
    devise_valeur = devise_obj.devise if devise_obj else None                                                         

                                                                              
    donnees = Kyc_pm.objects.exclude(DEVISE=devise_valeur).exclude(DEVISE__isnull=True).exclude(DEVISE="").exclude(DEVISE="NA")

                                                                      
    if user.organe == "Chargé Client":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence, EXPL=user.code_expl)
    elif user.organe == "Directeur Agence":
        donnees = donnees.filter(FILIALE=user.filiale, AGENCE=user.agence)
    elif user.organe in users_filiale:
        donnees = donnees.filter(FILIALE=user.filiale)
    elif user.organe in users_groupe:
        pass                              

                                                       
    if filiale_param:
        donnees = donnees.filter(FILIALE__icontains=filiale_param)
    if agence_param:
        donnees = donnees.filter(AGENCE__icontains=agence_param)
    if expl_param:
        donnees = donnees.filter(EXPL__icontains=expl_param)
    donnees = apply_datouv_period_filter(donnees, request)

    if request.GET.get('incompletes') == '1':
        donnees = get_incomplete_clients_queryset(donnees, 'pm')

                            

                                
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comptes Devise PP"                         

              
    headers = ["FILIALE", "AGENCE", "EXPL", "CLIENT", "DEVISE", "AGEC", "CODAPE", "IDM", "RCSNO", "CAPITAL", "CA",
               "RESULTAT", "TEL"]
    ws.append(headers)

             
    for d in donnees:
        ws.append([
            d.FILIALE, d.AGENCE, d.EXPL, d.CLIENT, d.DEVISE, d.AGEC, d.CODAPE, d.IDM,
            d.RCSNO, d.CAPITAL, d.CA, d.RESULTAT, d.TEL
        ])

                                                 
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

                              
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    if request.GET.get('incompletes') == '1':
        filename = f"Comptes_en_devise_PM_incomplets_{date_str}.xlsx"
    else:
        filename = f"Comptes_en_devise_PM_{date_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def evolution_taux(request):
    user = request.user

    context = {}

    if user.organe == "Conformité Groupe":
                                                              
        filiales = TauxEvolution_filiale.objects.values_list("filiale", flat=True).distinct()
        data_filiales = {}

        for filiale in filiales:
            qs = TauxEvolution_filiale.objects.filter(filiale=filiale).order_by("id")

            dates = [str(i.id) for i in qs]                                                             
            taux_pp = [round((t.flux_PP / t.stock_PP) * 100, 2) if t.stock_PP else 0 for t in qs]
            taux_pm = [round((t.flux_PM / t.stock_PM) * 100, 2) if t.stock_PM else 0 for t in qs]

            data_filiales[filiale] = {
                "dates": dates,
                "taux_pp": taux_pp,
                "taux_pm": taux_pm,
            }

        context["data_filiales"] = data_filiales

    elif user.organe == "Conformité":
                                        
        qs = TauxEvolution_filiale.objects.filter(filiale=user.filiale).order_by("id")

        dates = [str(i.id) for i in qs]                                              
        taux_pp = [round((t.flux_PP / t.stock_PP) * 100, 2) if t.stock_PP else 0 for t in qs]
        taux_pm = [round((t.flux_PM / t.stock_PM) * 100, 2) if t.stock_PM else 0 for t in qs]

        context.update({
            "filiale": user.filiale,
            "dates_pp": dates,
            "taux_pp": taux_pp,
            "dates_pm": dates,
            "taux_pm": taux_pm,
        })

    return render(request, "statistiques.html", context)



@login_required
def taux_evolution_view(request):
    user = request.user
    if user.organe in ["Chargé Client", "Directeur Agence"]:
        from django.shortcuts import redirect
        return redirect('statistiques')
    context_cache_key = _build_dashboard_cache_key("evolution_filiale", user, request)
    cached_context = cache.get(context_cache_key)
    if cached_context is not None:
        return render(request, 'evolution_par_filiale.html', cached_context)

    user_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]

    filiale_sel = request.GET.get('filiale')

    if user.organe in user_groupe:
        filiales = TauxEvolution_filiale.objects.values_list('filiale', flat=True).distinct().order_by('filiale')
    else:
        filiales = TauxEvolution_filiale.objects.filter(filiale=user.filiale).values_list('filiale', flat=True).distinct().order_by('filiale')
        if not filiale_sel or filiale_sel != user.filiale:
            filiale_sel = user.filiale

    if not filiale_sel and filiales:
        filiale_sel = filiales[0]

    periode = request.GET.get('periode', 'journalier')
    periodes_valides = {'journalier', 'hebdomadaire', 'mensuel', 'annuel'}
    if periode not in periodes_valides:
        periode = 'journalier'

    rows = list(
        TauxEvolution_filiale.objects
        .filter(filiale=filiale_sel)
        .order_by('date')
        .values_list('date', 'flux_PM', 'flux_PP')
    )
    latest_taux_date = rows[-1][0] if rows else None

    def build_period_key(d):
        if periode == 'journalier':
            return d
        if periode == 'hebdomadaire':
            year, week, _ = d.isocalendar()
            return (year, week)
        if periode == 'annuel':
            return d.year
        return (d.year, d.month)

    def format_period_label(key):
        if periode == 'journalier':
            return key.strftime('%d/%m/%Y')
        if periode == 'hebdomadaire':
            return f"S{key[1]}-{key[0]}"
        if periode == 'annuel':
            return str(key)
        return f"{key[1]:02d}/{key[0]}"

    def _num(v):
        # Valeurs absentes ou 'N/A' : ignorees pour faire un saut sur la courbe
        if v is None:
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    grouped = {}
    for d, pm, pp in rows:
        key = build_period_key(d)
        bucket = grouped.setdefault(
            key,
            {"sum_pm": 0.0, "cnt_pm": 0, "sum_pp": 0.0, "cnt_pp": 0, "latest_date": d},
        )
        vpm = _num(pm)
        if vpm is not None:
            bucket["sum_pm"] += vpm
            bucket["cnt_pm"] += 1
        vpp = _num(pp)
        if vpp is not None:
            bucket["sum_pp"] += vpp
            bucket["cnt_pp"] += 1
        if d > bucket["latest_date"]:
            bucket["latest_date"] = d

    def _avg(bucket, sum_key, cnt_key):
        cnt = bucket[cnt_key]
        return round(bucket[sum_key] / cnt, 2) if cnt else None

    period_keys = sorted(grouped.keys())
    labels = [format_period_label(k) for k in period_keys]
    data_pm = [_avg(grouped[k], "sum_pm", "cnt_pm") for k in period_keys]
    data_pp = [_avg(grouped[k], "sum_pp", "cnt_pp") for k in period_keys]
    history_rows = [
        (
            grouped[k]["latest_date"],
            data_pm[idx],
            data_pp[idx],
            labels[idx],
        )
        for idx, k in enumerate(period_keys)
    ]

    kpi_pm = {'last': 0, 'diff': 0, 'status': 'up'}
    kpi_pp = {'last': 0, 'diff': 0, 'status': 'up'}

    def _fill_kpi(kpi, series):
        vals = [v for v in series if v is not None]
        if not vals:
            return
        kpi['last'] = round(vals[-1], 2)
        if len(vals) >= 2:
            kpi['diff'] = round(vals[-1] - vals[-2], 2)
            kpi['status'] = 'up' if kpi['diff'] >= 0 else 'down'

    _fill_kpi(kpi_pm, data_pm)
    _fill_kpi(kpi_pp, data_pp)

    quality_scope = evaluate_data_quality_scope(user)
                                                                               
                                                                               
    if filiale_sel and quality_scope.get('filiale') != filiale_sel:
        quality_scope = dict(quality_scope, filiale=filiale_sel, label=filiale_sel)
    rules_version = cache.get('quality_control_rules_version', 1)

    def compute_quality_rate_by_typology(applicability):
        scope_signature = (
            f"{quality_scope.get('filiale')}|{quality_scope.get('agence')}|"
            f"{quality_scope.get('expl')}|{applicability}|evolution_filiale"
        )
        scope_hash = hashlib.md5(scope_signature.encode('utf-8')).hexdigest()
        cache_key = f"quality_control:evolution_rate:v{rules_version}:{scope_hash}"
        cached_rate = cache.get(cache_key)
        if cached_rate is not None:
            return cached_rate

        snapshot_rate = _quality_rate_snapshot(quality_scope, applicability)
        if snapshot_rate is not None:
            return snapshot_rate

        rules = list(DataQualityRule.objects.filter(active=True, applicability=applicability))
        total_ok = 0
        total_evaluated = 0
        for rule in rules:
            stat = evaluate_data_quality_rule(
                rule,
                filiale=quality_scope.get('filiale'),
                agence=quality_scope.get('agence'),
                expl=quality_scope.get('expl'),
            )
            total_ok += stat.get('ok_count', 0)
            total_evaluated += stat.get('total', 0)

        rate = round(total_ok / total_evaluated * 100, 1) if total_evaluated else 0
        cache.set(cache_key, rate, timeout=3600)
        return rate

                                                                                 
                                                                                
    flux_rate_pp = _quality_rate_snapshot(quality_scope, 'PP', flux_stock='flux')
    flux_rate_pm = _quality_rate_snapshot(quality_scope, 'PM', flux_stock='flux')
    if flux_rate_pp is None:
        flux_rate_pp = compute_quality_rate_by_typology('PP')
    if flux_rate_pm is None:
        flux_rate_pm = compute_quality_rate_by_typology('PM')

    context = {
        'labels_json': json.dumps(labels),
        'data_pm_json': json.dumps(data_pm),
        'data_pp_json': json.dumps(data_pp),
        'filiales': list(filiales),
        'filiale_sel': filiale_sel,
        'periode': periode,
        'kpi_pm': kpi_pm,
        'kpi_pp': kpi_pp,
        'quality_rate_pp': flux_rate_pp,
        'quality_rate_pm': flux_rate_pm,
        'flux_window_label': _flux_window_label(),
        'quality_scope_label': quality_scope.get('label'),
        'history_rows': list(reversed(history_rows[-10:])),
        'latest_taux_date': latest_taux_date,
        'is_filiale_user': user.organe not in user_groupe,
    }
    cache.set(context_cache_key, context, timeout=60 * 60 * 12)
    return render(request, 'evolution_par_filiale.html', context)


@login_required
def taux_evolution_view_stock(request):
    user = request.user
    if user.organe in ["Chargé Client", "Directeur Agence"]:
        from django.shortcuts import redirect
        return redirect('statistiques')
    context_cache_key = _build_dashboard_cache_key("evolution_filiale_stock", user, request)
    cached_context = cache.get(context_cache_key)
    if cached_context is not None:
        return render(request, 'evolution_par_filiale_stock.html', cached_context)

    user_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]

    filiale_sel = request.GET.get('filiale')

    if user.organe in user_groupe:
        filiales = TauxEvolution_filiale.objects.values_list('filiale', flat=True).distinct().order_by('filiale')
    else:
        filiales = TauxEvolution_filiale.objects.filter(filiale=user.filiale).values_list('filiale', flat=True).distinct().order_by('filiale')
        if not filiale_sel or filiale_sel != user.filiale:
            filiale_sel = user.filiale

    if not filiale_sel and filiales:
        filiale_sel = filiales[0]

    periode = request.GET.get('periode', 'journalier')
    periodes_valides = {'journalier', 'hebdomadaire', 'mensuel', 'annuel'}
    if periode not in periodes_valides:
        periode = 'journalier'

    rows = list(
        TauxEvolution_filiale.objects
        .filter(filiale=filiale_sel)
        .order_by('date')
        .values_list('date', 'stock_PM', 'stock_PP')
    )
    latest_taux_date = rows[-1][0] if rows else None

    def build_period_key(d):
        if periode == 'journalier':
            return d
        if periode == 'hebdomadaire':
            year, week, _ = d.isocalendar()
            return (year, week)
        if periode == 'annuel':
            return d.year
        return (d.year, d.month)

    def format_period_label(key):
        if periode == 'journalier':
            return key.strftime('%d/%m/%Y')
        if periode == 'hebdomadaire':
            return f"S{key[1]}-{key[0]}"
        if periode == 'annuel':
            return str(key)
        return f"{key[1]:02d}/{key[0]}"

    def _num(v):
        # Valeurs absentes ou 'N/A' : ignorees pour faire un saut sur la courbe
        if v is None:
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    grouped = {}
    for d, pm, pp in rows:
        key = build_period_key(d)
        bucket = grouped.setdefault(
            key,
            {"sum_pm": 0.0, "cnt_pm": 0, "sum_pp": 0.0, "cnt_pp": 0, "latest_date": d},
        )
        vpm = _num(pm)
        if vpm is not None:
            bucket["sum_pm"] += vpm
            bucket["cnt_pm"] += 1
        vpp = _num(pp)
        if vpp is not None:
            bucket["sum_pp"] += vpp
            bucket["cnt_pp"] += 1
        if d > bucket["latest_date"]:
            bucket["latest_date"] = d

    def _avg(bucket, sum_key, cnt_key):
        cnt = bucket[cnt_key]
        return round(bucket[sum_key] / cnt, 2) if cnt else None

    period_keys = sorted(grouped.keys())
    labels = [format_period_label(k) for k in period_keys]
    data_pm = [_avg(grouped[k], "sum_pm", "cnt_pm") for k in period_keys]
    data_pp = [_avg(grouped[k], "sum_pp", "cnt_pp") for k in period_keys]
    history_rows = [
        (
            grouped[k]["latest_date"],
            data_pm[idx],
            data_pp[idx],
            labels[idx],
        )
        for idx, k in enumerate(period_keys)
    ]

    kpi_pm = {'last': 0, 'diff': 0, 'status': 'up'}
    kpi_pp = {'last': 0, 'diff': 0, 'status': 'up'}

    def _fill_kpi(kpi, series):
        vals = [v for v in series if v is not None]
        if not vals:
            return
        kpi['last'] = round(vals[-1], 2)
        if len(vals) >= 2:
            kpi['diff'] = round(vals[-1] - vals[-2], 2)
            kpi['status'] = 'up' if kpi['diff'] >= 0 else 'down'

    _fill_kpi(kpi_pm, data_pm)
    _fill_kpi(kpi_pp, data_pp)

    quality_scope = evaluate_data_quality_scope(user)
                                                                               
                                                                               
    if filiale_sel and quality_scope.get('filiale') != filiale_sel:
        quality_scope = dict(quality_scope, filiale=filiale_sel, label=filiale_sel)
    rules_version = cache.get('quality_control_rules_version', 1)

    def compute_quality_rate_by_typology(applicability):
        scope_signature = (
            f"{quality_scope.get('filiale')}|{quality_scope.get('agence')}|"
            f"{quality_scope.get('expl')}|{applicability}|evolution_filiale_stock"
        )
        scope_hash = hashlib.md5(scope_signature.encode('utf-8')).hexdigest()
        cache_key = f"quality_control:evolution_stock_rate:v{rules_version}:{scope_hash}"
        cached_rate = cache.get(cache_key)
        if cached_rate is not None:
            return cached_rate

        snapshot_rate = _quality_rate_snapshot(quality_scope, applicability)
        if snapshot_rate is not None:
            return snapshot_rate

        rules = list(DataQualityRule.objects.filter(active=True, applicability=applicability))
        total_ok = 0
        total_evaluated = 0
        for rule in rules:
            stat = evaluate_data_quality_rule(
                rule,
                filiale=quality_scope.get('filiale'),
                agence=quality_scope.get('agence'),
                expl=quality_scope.get('expl'),
            )
            total_ok += stat.get('ok_count', 0)
            total_evaluated += stat.get('total', 0)

        rate = round(total_ok / total_evaluated * 100, 1) if total_evaluated else 0
        cache.set(cache_key, rate, timeout=3600)
        return rate

    context = {
        'labels_json': json.dumps(labels),
        'data_pm_json': json.dumps(data_pm),
        'data_pp_json': json.dumps(data_pp),
        'filiales': list(filiales),
        'filiale_sel': filiale_sel,
        'periode': periode,
        'kpi_pm': kpi_pm,
        'kpi_pp': kpi_pp,
                                                                    
        'quality_rate_pp': compute_quality_rate_by_typology('PP'),
        'quality_rate_pm': compute_quality_rate_by_typology('PM'),
        'quality_scope_label': quality_scope.get('label'),
        'history_rows': list(reversed(history_rows[-10:])),
        'latest_taux_date': latest_taux_date,
        'is_filiale_user': user.organe not in user_groupe,
    }
    cache.set(context_cache_key, context, timeout=60 * 60 * 12)
    return render(request, 'evolution_par_filiale_stock.html', context)


import csv
import io
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required

User = get_user_model()


@login_required
def bulk_user_upload(request):
    current_user = request.user

                                                                         
                                                                             
    if current_user.organe not in ("PASS", "DSI"):
        messages.error(request, "Vous n'avez pas la permission de créer des comptes utilisateurs.")
        return redirect('accueil')

    if request.method == "POST":
        csv_file = request.FILES.get('file')

        if not csv_file or not csv_file.name.endswith('.csv'):
            messages.error(request, "Veuillez sélectionner un fichier CSV valide.")
            return redirect('bulk_user_upload')

                                                                         
        organes_autorises = {value for value, _ in Organe}

        try:
                                
            data_set = csv_file.read().decode('UTF-8')
            io_string = io.StringIO(data_set)
            next(io_string)                     

            users_created = 0
            errors = 0
            rejected = []

            for row in csv.reader(io_string, delimiter=',', quotechar='"'):
                                                                            
                                                                     
                try:
                    organe = (row[3] or "").strip()
                    if organe not in organes_autorises:
                        rejected.append(f"{row[0]} (organe '{organe}' invalide)")
                        errors += 1
                        continue
                                                                              
                                                             
                    if current_user.organe == "DSI" and organe == "PASS":
                        rejected.append(f"{row[0]} (organe PASS reserve a PASS)")
                        errors += 1
                        continue

                    user, created = User.objects.get_or_create(
                        username=row[0],
                        defaults={
                            'first_name': row[1],
                            'last_name': row[2],
                            'organe': organe,
                            'téléphone': row[4],
                            'agence': row[6],
                            'code_expl': row[7],
                        }
                    )
                    if created:
                        if current_user.organe == "DSI":
                            user.filiale = current_user.filiale
                        user.set_password(row[5])             
                        user.save()
                        users_created += 1
                        log_audit(
                            request,
                            category=AuditEvent.CAT_SECURITE,
                            action="Creation de compte (import CSV)",
                            target=user.username,
                            details=(f"Compte {user.username} cree par import de masse | "
                                     f"organe={user.organe} | filiale={user.filiale}"),
                        )
                except Exception:
                    errors += 1
                    continue

            messages.success(request, f"{users_created} utilisateurs créés avec succès. ({errors} erreurs)")
            if rejected:
                messages.warning(request, "Lignes refusées : " + " ; ".join(rejected[:10]))

        except Exception as e:
            messages.error(request, f"Erreur lors du traitement : {e}")

    return render(request, 'bulk_upload.html')


from openpyxl import Workbook
from django.http import HttpResponse


@login_required
def download_excel_template(request):
                                          
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Utilisateurs"

                                         
    headers = ['username', 'first_name', 'last_name', 'organe', 'téléphone', 'password', 'agence', 'expl']
    ws.append(headers)

                        
    ws.append(['m.diop', 'Moussa', 'Diop', 'Conformité', '771234567', 'Boa2026!', 'Agence Dakar', 'EXPL001'])

                                    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="template_kyc_bulk.xlsx"'

    wb.save(response)
    return response


@login_required
def kyc_field_config(request):
                                                                                        
    KYC_PP_FIELD_LABELS = [(f, f) for f in KYC_PP_CONFIG_FIELDS]
    KYC_PM_FIELD_LABELS = [(f, f) for f in KYC_PM_CONFIG_FIELDS]

                                  
    filiale_choices = sorted(list(set(
        list(Kyc_pp.objects.exclude(FILIALE="").values_list('FILIALE', flat=True).distinct()) +
        list(Kyc_pm.objects.exclude(FILIALE="").values_list('FILIALE', flat=True).distinct())
    )))

    selected_filiale = request.GET.get('filiale_modal', '').strip()

    if request.method == "POST":
        action = request.POST.get('action', '')
        if action == "save_filiale_modal":
                                                                    
            sel_filiale = request.POST.get('selected_filiale', '').strip()
            if sel_filiale:
                for ct in ['pp', 'pm']:
                    config_id = request.POST.get(f'{ct}_config_id', '')
                    empty_fields = request.POST.getlist(f'{ct}_empty_fields')
                    display_fields = request.POST.getlist(f'{ct}_display_fields')
                    master_fields = KYC_PP_CONFIG_FIELDS if ct == 'pp' else KYC_PM_CONFIG_FIELDS
                    field_labels = {}
                    for f in master_fields:
                        val = (request.POST.get(f'{ct}_label_{f}') or '').strip()
                        if val:
                            field_labels[f] = val

                    if config_id:
                        config = KycFieldVisibilityConfig.objects.get(id=config_id)
                        config.empty_check_fields = empty_fields
                        config.display_fields = display_fields
                        config.field_labels = field_labels
                        config.save()
                    else:
                                                                 
                        KycFieldVisibilityConfig.objects.create(
                            client_type=ct,
                            filiales=[sel_filiale],
                            empty_check_fields=empty_fields,
                            display_fields=display_fields,
                            field_labels=field_labels
                        )
                messages.success(request, f"Configurations spécifiques pour la filiale {sel_filiale} enregistrées.")
            return redirect('kyc_field_config')
            
        else:
                                                            
            client_type = request.POST.get('client_type', 'pp')
            config_id = request.POST.get('config_id', '')
            scope = request.POST.get('scope', 'global')
            filiales = request.POST.getlist('filiales') if scope == 'filiales' else []
            empty_fields = request.POST.getlist('empty_fields')
            display_fields = request.POST.getlist('display_fields')
            master_fields = KYC_PP_CONFIG_FIELDS if client_type == 'pp' else KYC_PM_CONFIG_FIELDS
            field_labels = {}
            for f in master_fields:
                val = (request.POST.get(f'label_{f}') or '').strip()
                if val:
                    field_labels[f] = val

            if action == "delete":
                if config_id:
                    KycFieldVisibilityConfig.objects.filter(id=config_id).delete()
                    messages.success(request, "Configuration supprimée.")
                return redirect('kyc_field_config')

            if config_id:
                config = KycFieldVisibilityConfig.objects.get(id=config_id)
                config.empty_check_fields = empty_fields
                config.display_fields = display_fields
                config.field_labels = field_labels
                if not config.filiales or scope == 'filiales':
                    config.filiales = filiales
                config.save()
            else:
                KycFieldVisibilityConfig.objects.create(
                    client_type=client_type,
                    filiales=filiales,
                    empty_check_fields=empty_fields,
                    display_fields=display_fields,
                    field_labels=field_labels
                )
            messages.success(request, "Configuration enregistrée.")
            return redirect('kyc_field_config')

                               
    configs_qs = KycFieldVisibilityConfig.objects.all()
    
                                         
    for ct in ['pp', 'pm']:
        global_exists = any(not c.filiales for c in configs_qs if c.client_type == ct)
        if not global_exists:
            default_fields = [f[0] for f in (KYC_PP_FIELD_LABELS if ct == 'pp' else KYC_PM_FIELD_LABELS)]
            KycFieldVisibilityConfig.objects.create(
                client_type=ct,
                filiales=[],
                empty_check_fields=[],
                display_fields=default_fields
            )
            configs_qs = KycFieldVisibilityConfig.objects.all()

    configs_list = list(configs_qs)
    for c in configs_list:
        c.is_global = not c.filiales or len(c.filiales) == 0
        if c.is_global:
            c.scope_label = "Toutes les filiales"
        else:
            c.scope_label = f"Filiales : {', '.join(c.filiales)}"
        c.empty_fields = c.empty_check_fields
        c.display_field_names = c.display_fields
        c.custom_labels = c.field_labels or {}

    sections = [
        {
            'client_type': 'pp',
            'title': 'Particuliers (PP)',
            'fields': KYC_PP_FIELD_LABELS,
            'configs': [c for c in configs_list if c.client_type == 'pp']
        },
        {
            'client_type': 'pm',
            'title': 'Entreprises (PM)',
            'fields': KYC_PM_FIELD_LABELS,
            'configs': [c for c in configs_list if c.client_type == 'pm']
        }
    ]

    selected_filiale_configs = []
    if selected_filiale:
        for ct in ['pp', 'pm']:
            spec_config = None
            for c in configs_list:
                if c.client_type == ct and not c.is_global and selected_filiale in (c.filiales or []):
                    spec_config = c
                    break
            
            if spec_config:
                is_specific = True
                config_id = spec_config.id
                empty_fields = spec_config.empty_check_fields
                display_fields = spec_config.display_fields
                custom_labels = spec_config.field_labels or {}
                scope_label = f"Règle spécifique pour {selected_filiale}"
            else:
                is_specific = False
                global_c = next((c for c in configs_list if c.client_type == ct and c.is_global), None)
                config_id = None
                empty_fields = global_c.empty_check_fields if global_c else []
                display_fields = global_c.display_fields if global_c else [f[0] for f in (KYC_PP_FIELD_LABELS if ct == 'pp' else KYC_PM_FIELD_LABELS)]
                custom_labels = (global_c.field_labels or {}) if global_c else {}
                scope_label = "Hérité du global (Toutes les filiales)"

            selected_filiale_configs.append({
                'client_type': ct,
                'title': 'Particuliers (PP)' if ct == 'pp' else 'Entreprises (PM)',
                'config_id': config_id,
                'is_specific': is_specific,
                'scope_label': scope_label,
                'empty_fields': empty_fields,
                'display_field_names': display_fields,
                'custom_labels': custom_labels,
                'fields': KYC_PP_FIELD_LABELS if ct == 'pp' else KYC_PM_FIELD_LABELS,
                'filiales': [selected_filiale]
            })

    context = {
        'filia': getattr(request.user, 'filiale', ''),
        'filiale_choices': filiale_choices,
        'sections': sections,
        'selected_filiale': selected_filiale,
        'selected_filiale_configs': selected_filiale_configs,
    }
    return render(request, 'kyc_field_config.html', context)


def get_rate_color(rate, threshold):
    return "#ef4444" if rate < threshold else "#10b981"


def export_pilotage_excel(scope_data, summary, completeness_rows, quality_rows, notations_list,
                          notation_kpis, filiale_rates=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    BOA_GREEN = "0a3d2e"
    header_fill = PatternFill("solid", fgColor=BOA_GREEN)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    banner_fill = PatternFill("solid", fgColor=BOA_GREEN)
    title_font = Font(bold=True, size=16, color="FFFFFF")
    subtitle_font = Font(size=10, italic=True, color="D7E8DF")
    ok_fill = PatternFill("solid", fgColor="E8F5EC")
    ok_font = Font(color="0a7d34", bold=True)
    low_fill = PatternFill("solid", fgColor="FEE2E2")
    low_font = Font(color="B91C1C", bold=True)
    stripe_fill = PatternFill("solid", fgColor="F7FAF8")
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    thin_side = Side(style='thin', color='D0D7DE')
    thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def _style_header_row(ws, row_num=1):
        for cell in ws[row_num]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin
        ws.freeze_panes = ws.cell(row=row_num + 1, column=1).coordinate

    def _style_data_rows(ws, first_row, last_col, numeric_cols=()):
        for row_num in range(first_row, ws.max_row + 1):
            striped = (row_num - first_row) % 2 == 1
            for col_num in range(1, last_col + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin
                cell.alignment = center if col_num in numeric_cols else left
                if striped:
                    cell.fill = stripe_fill

    def _style_status_col(ws, col_idx, first_row=2):
        for row_num in range(first_row, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            if cell.value == "Sous seuil":
                cell.fill = low_fill
                cell.font = low_font
            elif cell.value == "Conforme":
                cell.fill = ok_fill
                cell.font = ok_font
            cell.alignment = center
            cell.border = thin


    ws1 = wb.active
    ws1.title = "Synthèse"
    ws1.sheet_properties.tabColor = BOA_GREEN

    ws1.append(["RAPPORT DE PILOTAGE KYC - BOA GROUP"])
    ws1.append([f"Périmètre : {scope_data.get('selected_filiale') or 'GROUPE'}    |    "
                f"Date de génération : {timezone.localtime().strftime('%d/%m/%Y %H:%M')}    |    "
                f"Seuil d'analyse : {summary.get('threshold', 90.0)}%"])
    ws1.append([])
    ws1.merge_cells("A1:C1")
    ws1.merge_cells("A2:C2")
    for row_num in (1, 2, 3):
        for col_num in (1, 2, 3):
            ws1.cell(row=row_num, column=col_num).fill = banner_fill
    ws1["A1"].font = title_font
    ws1["A1"].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws1["A2"].font = subtitle_font
    ws1["A2"].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws1.row_dimensions[1].height = 28
    ws1.row_dimensions[2].height = 20
    ws1.row_dimensions[3].height = 8

    def _fmt_valeur(value, unite):
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return f"{value}{unite}" if unite == "%" else value

    ws1.append(["Indicateur", "Valeur", "Statut"])
    header_row_1 = ws1.max_row
    ws1.append(["Taux de complétude global", _fmt_valeur(summary.get("completeness_rate"), "%"), "Sous seuil" if (summary.get("completeness_rate") or 0) < summary.get("threshold", 90.0) else "Conforme"])
    ws1.append(["Taux de complétude PP", _fmt_valeur(summary.get("completeness_rate_pp"), "%"), "Sous seuil" if (summary.get("completeness_rate_pp") or 0) < summary.get("threshold", 90.0) else "Conforme"])
    ws1.append(["Taux de complétude PM", _fmt_valeur(summary.get("completeness_rate_pm"), "%"), "Sous seuil" if (summary.get("completeness_rate_pm") or 0) < summary.get("threshold", 90.0) else "Conforme"])
    ws1.append(["Taux de conformité qualité global", _fmt_valeur(summary.get("quality_rate"), "%"), "Sous seuil" if (summary.get("quality_rate") or 0) < summary.get("threshold", 90.0) else "Conforme"])
    ws1.append(["Taux de conformité qualité PP", _fmt_valeur(summary.get("quality_rate_pp"), "%"), "Sous seuil" if (summary.get("quality_rate_pp") or 0) < summary.get("threshold", 90.0) else "Conforme"])
    ws1.append(["Taux de conformité qualité PM", _fmt_valeur(summary.get("quality_rate_pm"), "%"), "Sous seuil" if (summary.get("quality_rate_pm") or 0) < summary.get("threshold", 90.0) else "Conforme"])
    ws1.append(["Nombre de champs sous seuil", summary.get("low_completeness_count"), ""])
    ws1.append(["Nombre de règles sous seuil", summary.get("low_quality_count"), ""])
    _style_header_row(ws1, header_row_1)
    _style_data_rows(ws1, header_row_1 + 1, 3, numeric_cols=(2,))
    _style_status_col(ws1, 3, header_row_1 + 1)

    ws_fil = None
    if scope_data.get("scope") == "groupe" and filiale_rates:
        threshold_x = summary.get("threshold", 90.0)
        ws_fil = wb.create_sheet(title="Synthèse par filiale")
        ws_fil.sheet_properties.tabColor = "2563EB"
        ws_fil.append(["Filiale", "Clients PP", "Clients PM",
                       "Complétude globale", "Complétude PP", "Complétude PM",
                       "Qualité globale", "Qualité PP", "Qualité PM",
                       "Statut complétude", "Statut qualité"])
        for fr in filiale_rates:
            cg, qg = fr.get("comp_global"), fr.get("qual_global")
            ws_fil.append([
                fr.get("filiale"),
                fr.get("total_pp"), fr.get("total_pm"),
                cg, fr.get("comp_pp"), fr.get("comp_pm"),
                qg, fr.get("qual_pp"), fr.get("qual_pm"),
                "" if cg is None else ("Sous seuil" if cg < threshold_x else "Conforme"),
                "" if qg is None else ("Sous seuil" if qg < threshold_x else "Conforme"),
            ])
        _style_header_row(ws_fil)
        _style_data_rows(ws_fil, 2, 11, numeric_cols=(2, 3, 4, 5, 6, 7, 8, 9))
        _style_status_col(ws_fil, 10)
        _style_status_col(ws_fil, 11)

    ws2 = wb.create_sheet(title="Complétude")
    ws2.sheet_properties.tabColor = "0a7d34"
    ws2.append(["Type", "Filiale", "Champ (Code)", "Champ (Libellé)", "Total Clients", "Incomplets", "Taux", "Conformité"])
    for row in completeness_rows:
        status = "Sous seuil" if row.get("is_below_threshold") else "Conforme"
        ws2.append([
            row.get("type"),
            row.get("filiale"),
            row.get("field_name"),
            row.get("field_label"),
            row.get("total_clients"),
            row.get("missing_count"),
            row.get("rate"),
            status
        ])
    _style_header_row(ws2)
    _style_data_rows(ws2, 2, 8, numeric_cols=(5, 6, 7))
    _style_status_col(ws2, 8)


    ws3 = wb.create_sheet(title="Qualité")
    ws3.sheet_properties.tabColor = "F5A623"
    ws3.append(["Type", "Filiale", "Règle", "Champ", "Total Clients", "Anomalies", "Taux", "Conformité"])
    for row in quality_rows:
        status = "Sous seuil" if row.get("is_below_threshold") else "Conforme"
        ws3.append([
            row.get("type"),
            row.get("scope_label"),
            row.get("rule_name"),
            row.get("field_label"),
            row.get("total_clients"),
            row.get("fail_count"),
            row.get("rate"),
            status
        ])
    _style_header_row(ws3)
    _style_data_rows(ws3, 2, 8, numeric_cols=(5, 6, 7))
    _style_status_col(ws3, 8)


    ws4 = wb.create_sheet(title="Notation")
    ws4.sheet_properties.tabColor = "6366F1"
    ws4.append(["Agent Evalué", "Code Exploitant", "Filiale", "Note", "Flux / Stock", "Recommandations", "Evalué par", "Date évaluation"])
    for n in notations_list:
        ws4.append([
            n.agent.username,
            getattr(n.agent, "code_expl", "N/A"),
            n.agent.filiale,
            n.note,
            n.flux_stock,
            n.recommandation or "",
            n.note_par.username,
            n.date_notation.strftime("%d/%m/%Y %H:%M") if n.date_notation else ""
        ])
    _style_header_row(ws4)
    _style_data_rows(ws4, 2, 8, numeric_cols=(8,))

    for ws in [w for w in (ws1, ws_fil, ws2, ws3, ws4) if w is not None]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        ws.sheet_view.showGridLines = False

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    scope_label = "GROUPE" if scope_data.get("scope") == "groupe" else scope_data.get("selected_filiale", "FILIALE")
    scope_safe = re.sub(r"[^A-Za-z0-9_-]+", "_", scope_label)
    date_file = timezone.localtime().strftime("%Y%m%d")
    response['Content-Disposition'] = f'attachment; filename="rapport_pilotage_kyc_{scope_safe}_{date_file}.xlsx"'
    wb.save(response)
    return response


def _empty_field_missing_counts(queryset, field_names):
    """Compte en UNE SEULE requête les cellules « vides » (NULL / vide / espaces,
    au sens du script R) pour chaque champ, via une agrégation conditionnelle.
    Remplace les N COUNT(...) séparés qui rescannaient la table à chaque champ."""
    if not field_names:
        return {}
    aggs = {f"m_{i}": Count("pk", filter=empty_field_q(fn)) for i, fn in enumerate(field_names)}
    row = queryset.aggregate(**aggs)
    return {fn: (row.get(f"m_{i}") or 0) for i, fn in enumerate(field_names)}


def _pilotage_base_payload(scope, selected_filiale, allowed_filiales,
                           pp_fields_to_analyze, pm_fields_to_analyze,
                           rules_to_evaluate, use_cache=True):
    """Partie INDÉPENDANTE DU SEUIL du pilotage (complétude + qualité) sur les
    grosses tables Kyc_pp / Kyc_pm. Le seuil n'entre dans aucune requête : il ne
    sert qu'à l'affichage (couleurs, marquage « sous seuil »), recalculé côté vue.

    Le résultat est donc mis en cache journalier et préchauffé par
    warm_ui_caches (_warm_pilotage), pour un chargement quasi instantané au seuil
    par défaut. La version qualité (_quality_cache_version) invalide le cache dès
    qu'une règle change ; la date le renouvelle chaque jour après import."""
    rules_to_evaluate = list(rules_to_evaluate)
    scope_key = selected_filiale if scope == 'filiale' else 'GROUPE:' + ','.join(allowed_filiales)
    fields_sig = (','.join(fn for fn, _ in pp_fields_to_analyze) + '|'
                  + ','.join(fn for fn, _ in pm_fields_to_analyze))
    rules_sig = ','.join(str(r.id) for r in rules_to_evaluate)
    raw = f"{scope}|{scope_key}|{fields_sig}|{rules_sig}"
    key = (f"pilotage:base:v{_quality_cache_version()}:"
           f"d{timezone.localdate().isoformat()}:"
           f"{hashlib.md5(raw.encode('utf-8')).hexdigest()}")
    if use_cache:
        cached = cache.get(key)
        if cached is not None:
            return cached

    if scope == 'filiale':
        pp_queryset = Kyc_pp.objects.filter(FILIALE=selected_filiale)
        pm_queryset = Kyc_pm.objects.filter(FILIALE=selected_filiale)
    else:
        pp_queryset = Kyc_pp.objects.filter(FILIALE__in=allowed_filiales)
        pm_queryset = Kyc_pm.objects.filter(FILIALE__in=allowed_filiales)

    total_pp = pp_queryset.count()
    total_pm = pm_queryset.count()
    scope_label = selected_filiale if scope == 'filiale' else 'GROUPE'

                                                                          
    pp_missing = _empty_field_missing_counts(pp_queryset, [fn for fn, _ in pp_fields_to_analyze])
    pm_missing = _empty_field_missing_counts(pm_queryset, [fn for fn, _ in pm_fields_to_analyze])

    completeness_rows_pp = []
    total_missing_pp = 0
    for f_name, f_label in pp_fields_to_analyze:
        missing_count = pp_missing.get(f_name, 0)
        rate = completeness_rate_r(missing_count, total_pp)
        if rate is None:
            rate = 100.0
        completeness_rows_pp.append({
            'type': 'PP', 'filiale': scope_label, 'field_name': f_name,
            'field_label': f_label, 'total_clients': total_pp,
            'missing_count': missing_count, 'rate': rate,
        })
        total_missing_pp += missing_count

    completeness_rows_pm = []
    total_missing_pm = 0
    for f_name, f_label in pm_fields_to_analyze:
        missing_count = pm_missing.get(f_name, 0)
        rate = completeness_rate_r(missing_count, total_pm)
        if rate is None:
            rate = 100.0
        completeness_rows_pm.append({
            'type': 'PM', 'filiale': scope_label, 'field_name': f_name,
            'field_label': f_label, 'total_clients': total_pm,
            'missing_count': missing_count, 'rate': rate,
        })
        total_missing_pm += missing_count

             
    quality_rows_pp = []
    quality_rows_pm = []
    total_ok_pp = total_eval_rules_pp = 0
    total_ok_pm = total_eval_rules_pm = 0
    eval_fil = selected_filiale if scope == 'filiale' else None
    for rule in rules_to_evaluate:
        stat = evaluate_data_quality_rule(rule, filiale=eval_fil)
        total = stat.get('total', 0)
        fail_count = stat.get('fail_count', 0)
        ok_count = stat.get('ok_count', 0)
        rate = compliance_rate_floor(ok_count, total, fail_count)
        if rate is None:
            rate = 100.0
        row = {
            'id': rule.id, 'rule_number': rule.rule_number, 'type': rule.applicability, 'scope_label': scope_label,
            'rule_name': rule.name,
            'field_label': dict(DATA_QUALITY_FIELD_CHOICES).get(rule.field_name, rule.field_name),
            'total_clients': total, 'fail_count': fail_count, 'rate': rate,
            'export_url': reverse('kyc:export_rule_failures', kwargs={'rule_id': rule.id}),
        }
        if rule.applicability == 'PP':
            quality_rows_pp.append(row)
            total_ok_pp += ok_count
            total_eval_rules_pp += total
        else:
            quality_rows_pm.append(row)
            total_ok_pm += ok_count
            total_eval_rules_pm += total

    payload = {
        'total_pp': total_pp, 'total_pm': total_pm,
        'completeness_rows_pp': completeness_rows_pp,
        'completeness_rows_pm': completeness_rows_pm,
        'total_missing_pp': total_missing_pp, 'total_missing_pm': total_missing_pm,
        'quality_rows_pp': quality_rows_pp, 'quality_rows_pm': quality_rows_pm,
        'total_ok_pp': total_ok_pp, 'total_eval_rules_pp': total_eval_rules_pp,
        'total_ok_pm': total_ok_pm, 'total_eval_rules_pm': total_eval_rules_pm,
    }
    cache.set(key, payload, timeout=86400)
    return payload


def _pilotage_page_cache_key(can_group, scope, selected_filiale, allowed_filiales,
                             threshold, pp_fields_to_analyze, pm_fields_to_analyze,
                             rules_to_evaluate, selected_report_fields, selected_report_rules):
    """Clé de cache du CONTEXTE COMPLET de /pilotage-kyc.

    Comme les autres dashboards : hash de (scope effectif + paramètres GET
    significatifs + version des données). Renouvelée chaque jour (localdate) et
    dès qu'une règle qualité change (_quality_cache_version). Le seuil et la
    sélection de champs/règles du rapport en font partie."""
    fields_sig = (','.join(fn for fn, _ in pp_fields_to_analyze) + '|'
                  + ','.join(fn for fn, _ in pm_fields_to_analyze))
    rules_sig = ','.join(str(r.id) for r in rules_to_evaluate)
    raw = '|'.join([
        'G' if can_group else 'F',
        scope,
        selected_filiale or '',
        ','.join(allowed_filiales or []),
        f'{threshold:g}',
        fields_sig,
        rules_sig,
        ','.join(sorted(selected_report_fields or [])),
        ','.join(sorted(selected_report_rules or [])),
        f'v{_quality_cache_version()}',
        timezone.localdate().isoformat(),
    ])
    return 'pilotage:page:v1:' + hashlib.md5(raw.encode('utf-8')).hexdigest()


def _pilotage_filiale_rates(allowed_filiales, pp_fields_to_analyze, pm_fields_to_analyze,
                            rules_to_evaluate, use_cache=True):
    """Synthèse complétude + qualité filiale par filiale, pour le rapport GROUPE.

    Ne retient que les filiales qui ont effectivement des données en base
    (au moins un client PP ou PM). Résultat mis en cache à la journée, comme
    _pilotage_base_payload : le calcul est lourd (N filiales x N champs)."""
    rules_to_evaluate = list(rules_to_evaluate)
    raw = ('|'.join(allowed_filiales) + '#'
           + ','.join(fn for fn, _ in pp_fields_to_analyze) + '#'
           + ','.join(fn for fn, _ in pm_fields_to_analyze) + '#'
           + ','.join(str(r.id) for r in rules_to_evaluate))
    key = (f"pilotage:filrates:v{_quality_cache_version()}:"
           f"d{timezone.localdate().isoformat()}:"
           f"{hashlib.md5(raw.encode('utf-8')).hexdigest()}")
    if use_cache:
        cached = cache.get(key)
        if cached is not None:
            return cached

    filiale_rates = []
    for fil in allowed_filiales:
        fpp = Kyc_pp.objects.filter(FILIALE=fil)
        fpm = Kyc_pm.objects.filter(FILIALE=fil)
        tpp_f, tpm_f = fpp.count(), fpm.count()
        if not (tpp_f or tpm_f):
            continue
        miss_pp = sum(fpp.filter(empty_field_q(fn)).count()
                      for fn, _ in pp_fields_to_analyze)
        miss_pm = sum(fpm.filter(empty_field_q(fn)).count()
                      for fn, _ in pm_fields_to_analyze)
        ev_pp = tpp_f * len(pp_fields_to_analyze)
        ev_pm = tpm_f * len(pm_fields_to_analyze)
        c_pp = completeness_rate_r(miss_pp, ev_pp)
        c_pm = completeness_rate_r(miss_pm, ev_pm)
        c_g = completeness_rate_r(miss_pp + miss_pm, ev_pp + ev_pm)
        ok_pp = ev_q_pp = ok_pm = ev_q_pm = 0
        for rule in rules_to_evaluate:
            st = evaluate_data_quality_rule(rule, filiale=fil)
            if rule.applicability == 'PP':
                ok_pp += st.get('ok_count', 0); ev_q_pp += st.get('total', 0)
            else:
                ok_pm += st.get('ok_count', 0); ev_q_pm += st.get('total', 0)
        q_pp = round(ok_pp / ev_q_pp * 100, 1) if ev_q_pp else None
        q_pm = round(ok_pm / ev_q_pm * 100, 1) if ev_q_pm else None
        q_g = round((ok_pp + ok_pm) / (ev_q_pp + ev_q_pm) * 100, 1) if (ev_q_pp + ev_q_pm) else None
        filiale_rates.append({
            'filiale': fil,
            'total_pp': tpp_f, 'total_pm': tpm_f,
            'comp_global': c_g, 'comp_pp': c_pp, 'comp_pm': c_pm,
            'qual_global': q_g, 'qual_pp': q_pp, 'qual_pm': q_pm,
        })

    cache.set(key, filiale_rates, timeout=86400)
    return filiale_rates


@login_required
def pilotage_kyc(request):
    user = request.user

                                                          
    user_groupe = [
        "Directeur Zone UEMOA", "Directeur Zone Centre", "Directeur Zone Anglophone",
        "Conformité Groupe", "Contrôle Permanent Groupe", "PASS", "GUEST"
    ]
    can_group = user.is_superuser or user.organe in user_groupe

                                                               
    filiale_organes = ["Conformité", "DSI", "Qualité", "Contrôle Permanent"]
    if not can_group and user.organe not in filiale_organes:
        messages.error(request, "Accès refusé : le pilotage est réservé aux organes Conformité, DSI, Qualité et Contrôle Permanent.")
        return redirect("accueil")
    
                                                                            
                                                                             
    _fil_key = 'pilotage:allowed_filiales:' + timezone.localdate().isoformat()
    allowed_filiales = None if getattr(request, '_force_daily_cache_refresh', False) else cache.get(_fil_key)
    if allowed_filiales is None:
        allowed_filiales = sorted(list(set(
            list(Kyc_pp.objects.exclude(FILIALE="").values_list('FILIALE', flat=True).distinct()) +
            list(Kyc_pm.objects.exclude(FILIALE="").values_list('FILIALE', flat=True).distinct())
        )))
        cache.set(_fil_key, allowed_filiales, timeout=86400)
    if not can_group:
        user_filiale = getattr(user, 'filiale', '')
        allowed_filiales = [user_filiale] if user_filiale else []
        scope = "filiale"
        selected_filiale = user_filiale
    else:
        scope = request.GET.get('scope', 'filiale')
        if scope not in ['filiale', 'groupe']:
            scope = 'filiale'
        
        if scope == 'filiale':
            selected_filiale = request.GET.get('filiale', '')
            if not selected_filiale and allowed_filiales:
                selected_filiale = allowed_filiales[0]
        else:
            selected_filiale = ""

                                                                               
                                                                                  
                                                                                 
                                                      
    user_filiale = (getattr(user, 'filiale', '') or '').strip()
    if user_filiale and not can_group:
        allowed_filiales = [user_filiale]
        scope = 'filiale'
        selected_filiale = user_filiale

               
    try:
        threshold = float(request.GET.get('threshold', '90.0').replace(',', '.'))
    except ValueError:
        threshold = 90.0

    KYC_PP_FIELD_LABELS = [
        ("CLIENT", "CLIENT"),
        ("EXPL", "EXPL"),
        ("FILIALE", "FILIALE"),
        ("AGENCE", "AGENCE"),
        ("LIB_AGENCE", "LIB_AGENCE"),
        ("IDP", "IDP"),
        ("PAYNAIS", "PAYNAIS"),
        ("PROFESSION", "PROFESSION"),
        ("SALAIRE", "SALAIRE"),
        ("NUMID", "NUMID"),
        ("CODAPE", "CODAPE"),
        ("TEL", "TEL"),
        ("DATNAIS", "DATNAIS"),
        ("ADRESSE", "ADRESSE"),
        ("DATVALID", "DATVALID"),
        ("ORIGINE_REV", "ORIGINE_REV"),
        ("INTITULE_COMPTE", "INTITULE_COMPTE"),
        ("EMPLOYEUR", "EMPLOYEUR"),
        ("PAYS_RESID", "PAYS_RESID"),
        ("LIEU_DELIVRANCE_CIN", "LIEU_DELIVRANCE_CIN"),
        ("BOITE_POSTALE", "BOITE_POSTALE"),
        ("CONSENT_BIC", "CONSENT_BIC"),
        ("DATOUV", "DATOUV"),
        ("PPE", "PPE"),
        ("DEVISE", "DEVISE"),
        ("RESID", "RESID"),
        ("DATEREV", "DATEREV"),
        ("RISQUE", "RISQUE"),
    ]

    KYC_PM_FIELD_LABELS = [
        ("CLIENT", "CLIENT"),
        ("EXPL", "EXPL"),
        ("FILIALE", "FILIALE"),
        ("AGENCE", "AGENCE"),
        ("LIB_AGENCE", "LIB_AGENCE"),
        ("IDM", "IDM"),
        ("CODAPE", "CODAPE"),
        ("AGEC", "AGEC"),
        ("CAPITAL", "CAPITAL"),
        ("CA", "CA"),
        ("RESULTAT", "RESULTAT"),
        ("RCSNO", "RCSNO"),
        ("ORIGINE_REV", "ORIGINE_REV"),
        ("TEL", "TEL"),
        ("INTITULE_COMPTE", "INTITULE_COMPTE"),
        ("ADRESSE_SOCIALE", "ADRESSE_SOCIALE"),
        ("NUMERO_FISCAL", "NUMERO_FISCAL"),
        ("PAYS_JUR", "PAYS_JUR"),
        ("ACTIONNAIRE", "ACTIONNAIRE"),
        ("MANDATAIRE", "MANDATAIRE"),
        ("BOITE_POSTALE", "BOITE_POSTALE"),
        ("CONSENT_BIC", "CONSENT_BIC"),
        ("DATOUV", "DATOUV"),
        ("DEVISE", "DEVISE"),
        ("RESID", "RESID"),
        ("DATEREV", "DATEREV"),
        ("PPE", "PPE"),
        ("RISQUE", "RISQUE"),
    ]

                      
    pp_config = None
    if scope == 'filiale' and selected_filiale:
        pp_config = next((c for c in KycFieldVisibilityConfig.objects.filter(client_type='pp') if selected_filiale in (c.filiales or [])), None)
    if not pp_config:
        pp_config = next((c for c in KycFieldVisibilityConfig.objects.filter(client_type='pp') if not c.filiales), None)
        
    if pp_config and pp_config.empty_check_fields:
        pp_fields_list = pp_config.empty_check_fields
    elif pp_config and pp_config.display_fields:
        pp_fields_list = pp_config.display_fields
    else:
        pp_fields_list = [f[0] for f in KYC_PP_FIELD_LABELS]
        
    pp_active_fields = [(f_name, dict(KYC_PP_FIELD_LABELS).get(f_name, f_name)) for f_name in pp_fields_list]

                      
    pm_config = None
    if scope == 'filiale' and selected_filiale:
        pm_config = next((c for c in KycFieldVisibilityConfig.objects.filter(client_type='pm') if selected_filiale in (c.filiales or [])), None)
    if not pm_config:
        pm_config = next((c for c in KycFieldVisibilityConfig.objects.filter(client_type='pm') if not c.filiales), None)
        
    if pm_config and pm_config.empty_check_fields:
        pm_fields_list = pm_config.empty_check_fields
    elif pm_config and pm_config.display_fields:
        pm_fields_list = pm_config.display_fields
    else:
        pm_fields_list = [f[0] for f in KYC_PM_FIELD_LABELS]
        
    pm_active_fields = [(f_name, dict(KYC_PM_FIELD_LABELS).get(f_name, f_name)) for f_name in pm_fields_list]

                                         
    selected_report_fields = request.GET.getlist('report_fields')
    if selected_report_fields:
        pp_fields_to_analyze = [f for f in pp_active_fields if f[0] in selected_report_fields]
        pm_fields_to_analyze = [f for f in pm_active_fields if f[0] in selected_report_fields]
    else:
        pp_fields_to_analyze = pp_active_fields
        pm_fields_to_analyze = pm_active_fields

                                                        
    all_quality_rules = DataQualityRule.objects.filter(active=True)
    selected_report_rules = request.GET.getlist('report_rules')
    if selected_report_rules:
        rules_to_evaluate = all_quality_rules.filter(id__in=selected_report_rules)
    else:
        rules_to_evaluate = all_quality_rules

    # Cache plein contexte (comme les autres dashboards) : les données KYC sont
    # injectées une fois par jour, la page peut donc être servie depuis le cache.
    # On saute le cache pour les exports (réponse fichier) et pour la préchauffe.
    _force_refresh = getattr(request, '_force_daily_cache_refresh', False)
    _page_cache_key = None
    if not request.GET.get('export'):
        _page_cache_key = _pilotage_page_cache_key(
            can_group, scope, selected_filiale, allowed_filiales, threshold,
            pp_fields_to_analyze, pm_fields_to_analyze, rules_to_evaluate,
            selected_report_fields, selected_report_rules,
        )
        if not _force_refresh:
            _cached_context = cache.get(_page_cache_key)
            if _cached_context is not None:
                return render(request, 'pilotage_kyc.html', _cached_context)




    base = _pilotage_base_payload(
        scope, selected_filiale, allowed_filiales,
        pp_fields_to_analyze, pm_fields_to_analyze, rules_to_evaluate,
        use_cache=not getattr(request, '_force_daily_cache_refresh', False),
    )

    def _mark(rows):
        marked = []
        for r in rows:
            r = dict(r)
            r['is_below_threshold'] = r['rate'] < threshold
            marked.append(r)
        return marked

    total_pp = base['total_pp']
    total_pm = base['total_pm']

                  
    completeness_rows_pp = _mark(base['completeness_rows_pp'])
    completeness_rows_pm = _mark(base['completeness_rows_pm'])
    total_missing_pp = base['total_missing_pp']
    total_missing_pm = base['total_missing_pm']
    total_evaluated_pp = total_pp * len(pp_fields_to_analyze)
    total_evaluated_pm = total_pm * len(pm_fields_to_analyze)

    completeness_rows = completeness_rows_pp + completeness_rows_pm
    low_completeness_rows = [r for r in completeness_rows if r['is_below_threshold']]
    low_completeness_rows.sort(key=lambda r: r['rate'])

                                                                                     
    # KPI Global / PP / PM : même règle que la qualité — taux tronqué à 0,1 %,
    # 100 % seulement si 0 cellule vide, sinon plafonné à 99,9 %.
    def _completeness_rate_1dec(empty_cells, total_cells):
        r = _rate_floor_1dec(total_cells - empty_cells, total_cells)
        return 100.0 if r is None else r

    completeness_rate_pp = _completeness_rate_1dec(total_missing_pp, total_evaluated_pp)
    completeness_rate_pm = _completeness_rate_1dec(total_missing_pm, total_evaluated_pm)
    completeness_rate = _completeness_rate_1dec(
        total_missing_pp + total_missing_pm,
        total_evaluated_pp + total_evaluated_pm)

    low_completeness_count_pp = sum(1 for r in completeness_rows_pp if r['is_below_threshold'])
    low_completeness_count_pm = sum(1 for r in completeness_rows_pm if r['is_below_threshold'])
    low_completeness_count = low_completeness_count_pp + low_completeness_count_pm

             
    quality_rows_pp = _mark(base['quality_rows_pp'])
    quality_rows_pm = _mark(base['quality_rows_pm'])
    total_ok_pp = base['total_ok_pp']
    total_eval_rules_pp = base['total_eval_rules_pp']
    total_ok_pm = base['total_ok_pm']
    total_eval_rules_pm = base['total_eval_rules_pm']

    quality_rows = quality_rows_pp + quality_rows_pm
    low_quality_rows = [r for r in quality_rows if r['is_below_threshold']]
    low_quality_rows.sort(key=lambda r: r['rate'])

                            
    # Même logique que la complétude : taux tronqué (floor), calculé sur les totaux
    # globaux et cohérent avec les lignes de détail (compliance_rate_floor).
    quality_rate_pp = compliance_rate_floor(
        total_ok_pp, total_eval_rules_pp, total_eval_rules_pp - total_ok_pp)
    if quality_rate_pp is None:
        quality_rate_pp = 100.0

    quality_rate_pm = compliance_rate_floor(
        total_ok_pm, total_eval_rules_pm, total_eval_rules_pm - total_ok_pm)
    if quality_rate_pm is None:
        quality_rate_pm = 100.0

    total_ok_global = total_ok_pp + total_ok_pm
    total_eval_global = total_eval_rules_pp + total_eval_rules_pm
    quality_rate = compliance_rate_floor(
        total_ok_global, total_eval_global, total_eval_global - total_ok_global)
    if quality_rate is None:
        quality_rate = 100.0

    low_quality_count_pp = sum(1 for r in quality_rows_pp if r['is_below_threshold'])
    low_quality_count_pm = sum(1 for r in quality_rows_pm if r['is_below_threshold'])
    low_quality_count = low_quality_count_pp + low_quality_count_pm

    derniere_maj_kyc = TauxQualite.objects.aggregate(Max('date'))['date__max']

    summary_dict = {
        'threshold': threshold,
        'completeness_rate': completeness_rate,
        'completeness_rate_pp': completeness_rate_pp,
        'completeness_rate_pm': completeness_rate_pm,
        'low_completeness_count': low_completeness_count,
        'low_completeness_count_pp': low_completeness_count_pp,
        'low_completeness_count_pm': low_completeness_count_pm,
        'completeness_total': total_pp + total_pm,
        'completeness_total_pp': total_pp,
        'completeness_total_pm': total_pm,
        'quality_rate': quality_rate,
        'quality_rate_pp': quality_rate_pp,
        'quality_rate_pm': quality_rate_pm,
        'low_quality_count': low_quality_count,
        'low_quality_count_pp': low_quality_count_pp,
        'low_quality_count_pm': low_quality_count_pm,
        'quality_total': total_pp + total_pm,
        'derniere_maj_kyc': derniere_maj_kyc,
    }

                          
    if scope == 'filiale':
        notations = Notation.objects.filter(agent__filiale=selected_filiale).select_related('agent', 'note_par')
    else:
        notations = Notation.objects.filter(agent__filiale__in=allowed_filiales).select_related('agent', 'note_par')

    total_notations = notations.count()
    notations_list = list(notations.order_by('-date_notation'))



    seen_agents = set()
    latest_notes = []
    latest_notations = []
    for n in notations_list:
        agent_id = getattr(n.agent, 'pk', None)
        if agent_id in seen_agents:
            continue
        seen_agents.add(agent_id)
        latest_notes.append(n.note)
        latest_notations.append(n)

    total_agents = len(latest_notes)
    excellence_count = sum(1 for note in latest_notes if note in ('Très Bien', 'Bien'))
    excellence_rate = round((excellence_count / total_agents) * 100, 1) if total_agents > 0 else 0.0

    notation_kpis = {
        'total_agents': total_agents,
        'total_notations': total_notations,
        'excellence_rate': excellence_rate
    }

             
    export_format = request.GET.get('export')
    if export_format:
        scope_data = {
            'scope': scope,
            'selected_filiale': selected_filiale
        }


        filiale_rates = []
        if scope == 'groupe':
            filiale_rates = _pilotage_filiale_rates(
                allowed_filiales, pp_fields_to_analyze, pm_fields_to_analyze,
                rules_to_evaluate,
                use_cache=not getattr(request, '_force_daily_cache_refresh', False),
            )
        if export_format == 'pdf':
            from kyc.pilotage_exports import export_pilotage_pdf
            return export_pilotage_pdf(scope_data, summary_dict, completeness_rows, quality_rows,
                                       filiale_rates=filiale_rates,
                                       notations_list=notations_list, notation_kpis=notation_kpis)
        elif export_format == 'pptx':
            from kyc.pilotage_exports import export_pilotage_pptx
            return export_pilotage_pptx(scope_data, summary_dict, completeness_rows, quality_rows,
                                        notations_list=notations_list, notation_kpis=notation_kpis,
                                        filiale_rates=filiale_rates)
        elif export_format == 'excel':
            return export_pilotage_excel(scope_data, summary_dict, completeness_rows, quality_rows,
                                         notations_list, notation_kpis,
                                         filiale_rates=filiale_rates)

                          
    def _rate_desc(rows):
        return sorted(rows, key=lambda r: r['rate'] if r.get('rate') is not None else -1, reverse=True)

    comp_rows_pp_sorted = _rate_desc(completeness_rows_pp)
    comp_rows_pm_sorted = _rate_desc(completeness_rows_pm)
    qual_rows_pp_sorted = _rate_desc(quality_rows_pp)
    qual_rows_pm_sorted = _rate_desc(quality_rows_pm)

    chart_comp_pp_labels = [r['field_label'] for r in comp_rows_pp_sorted]
    chart_comp_pp_values = [r['rate'] for r in comp_rows_pp_sorted]
    chart_comp_pp_colors = [get_rate_color(r['rate'], threshold) for r in comp_rows_pp_sorted]

    chart_comp_pm_labels = [r['field_label'] for r in comp_rows_pm_sorted]
    chart_comp_pm_values = [r['rate'] for r in comp_rows_pm_sorted]
    chart_comp_pm_colors = [get_rate_color(r['rate'], threshold) for r in comp_rows_pm_sorted]

    chart_qual_pp_labels = [r['rule_name'] for r in qual_rows_pp_sorted]
    chart_qual_pp_values = [r['rate'] for r in qual_rows_pp_sorted]
    chart_qual_pp_colors = [get_rate_color(r['rate'], threshold) for r in qual_rows_pp_sorted]

    chart_qual_pm_labels = [r['rule_name'] for r in qual_rows_pm_sorted]
    chart_qual_pm_values = [r['rate'] for r in qual_rows_pm_sorted]
    chart_qual_pm_colors = [get_rate_color(r['rate'], threshold) for r in qual_rows_pm_sorted]

    # Graphiques basés sur l'évaluation actuelle (la plus récente) de chaque agent,
    # pas sur l'historique complet des notations.
    notes_order = ['Très Bien', 'Bien', 'Passable', 'Insuffisant']
    colors_map = {
        'Très Bien': '#10b981',
        'Bien': '#3b82f6',
        'Passable': '#f59e0b',
        'Insuffisant': '#ef4444'
    }

    chart_notation_overall_labels = notes_order
    chart_notation_overall_values = [
        sum(1 for note in latest_notes if note == lbl) for lbl in notes_order
    ]
    chart_notation_overall_colors = ['#10b981', '#3b82f6', '#f59e0b', '#ef4444']

    chart_notation_filiales = sorted({
        (n.agent.filiale or '') for n in latest_notations if (n.agent.filiale or '')
    })

    chart_notation_by_filiale_datasets = []
    for note in notes_order:
        data = []
        for fil in chart_notation_filiales:
            count = sum(
                1 for n in latest_notations
                if (n.agent.filiale or '') == fil and n.note == note
            )
            data.append(count)
        chart_notation_by_filiale_datasets.append({
            'label': note,
            'data': data,
            'backgroundColor': colors_map[note]
        })

    context = {
        'scope': scope,
        'can_group': can_group,
        'allowed_filiales': allowed_filiales,
        'selected_filiale': selected_filiale,
        'threshold': threshold,
        'summary': summary_dict,
        
        'pp_active_fields': pp_active_fields,
        'pm_active_fields': pm_active_fields,
        'selected_report_fields': selected_report_fields,
        'selected_report_rules': [int(rid) for rid in selected_report_rules if rid.isdigit()],
        'all_quality_rules': list(all_quality_rules),
        
        'low_completeness_rows': low_completeness_rows,
        'low_quality_rows': low_quality_rows,
        
        'notations_list': notations_list,
        'notation_kpis': notation_kpis,
        
                
        'chart_comp_pp_labels': chart_comp_pp_labels,
        'chart_comp_pp_values': chart_comp_pp_values,
        'chart_comp_pp_colors': chart_comp_pp_colors,
        'chart_comp_pm_labels': chart_comp_pm_labels,
        'chart_comp_pm_values': chart_comp_pm_values,
        'chart_comp_pm_colors': chart_comp_pm_colors,
        
        'chart_qual_pp_labels': chart_qual_pp_labels,
        'chart_qual_pp_values': chart_qual_pp_values,
        'chart_qual_pp_colors': chart_qual_pp_colors,
        'chart_qual_pm_labels': chart_qual_pm_labels,
        'chart_qual_pm_values': chart_qual_pm_values,
        'chart_qual_pm_colors': chart_qual_pm_colors,
        
        'chart_notation_overall_labels': chart_notation_overall_labels,
        'chart_notation_overall_values': chart_notation_overall_values,
        'chart_notation_overall_colors': chart_notation_overall_colors,
        
        'chart_notation_filiales': chart_notation_filiales,
        'chart_notation_by_filiale_datasets': chart_notation_by_filiale_datasets,
        
        'pp_fields_json': json.dumps(pp_active_fields),
        'pm_fields_json': json.dumps(pm_active_fields),
    }

    if _page_cache_key is not None:
        cache.set(_page_cache_key, context, timeout=86400)

    return render(request, 'pilotage_kyc.html', context)


                                                                               
                  
                                                                               

def _parse_daterev(value):
    """Parse une chaîne DATEREV en objet date, retourne None si invalide."""
    if not value:
        return None
    value = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y%m%d"):
        try:
            from datetime import datetime as _dt
            return _dt.strptime(value, fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _paid_daterev_filiales():
    """Filiales dont le module Rappels DATEREV est payé/actif."""
    from kyc.models import FilialeModuleConfig
    return set(FilialeModuleConfig.objects
               .filter(daterev_reminder_paye_active=True)
               .values_list('filiale', flat=True))


def _get_exploitants_daterev_expired(filiale_filter=None, days_before=30, only_paid=False):
    """
    Retourne un dict {filiale: {expl: {clients_pp, clients_pm}}}
    pour les exploitants ayant des clients avec DATEREV dépassée ou dans `days_before` jours.
    Optimisé : seuls les champs utiles sont récupérés, le parsing date se fait en Python
    uniquement sur les lignes ayant une DATEREV non vide.
    Si `only_paid=True`, restreint aux filiales dont le module Rappels DATEREV est payé.
    """
    from datetime import date, timedelta
    today = date.today()
    limit_date = today + timedelta(days=days_before)
    today_str = today.isoformat()                                             

    paid_filiales = _paid_daterev_filiales() if only_paid else None
    if only_paid and not paid_filiales:
        return {}                                         

    result = {}

    for model, label in [(Kyc_pp, 'PP'), (Kyc_pm, 'PM')]:
        id_field = 'IDP' if label == 'PP' else 'IDM'
        fields = ['FILIALE', 'EXPL', 'CLIENT', 'AGENCE', 'LIB_AGENCE', 'DATEREV', id_field, 'RISQUE']
        qs = (model.objects
              .exclude(DATEREV='').exclude(DATEREV__isnull=True)
              .exclude(EXPL='').exclude(FILIALE='')
              .only(*fields))
        if filiale_filter:
            qs = qs.filter(FILIALE=filiale_filter)
        if paid_filiales is not None:
            qs = qs.filter(FILIALE__in=paid_filiales)

        key_label = f'clients_{label.lower()}'
        for obj in qs.values(*fields):
            dr = _parse_daterev(obj['DATEREV'])
            if dr is None or dr > limit_date:
                continue

            filiale = obj['FILIALE']
            expl = obj['EXPL']

            if dr < today:
                echeance = f'Échue depuis {(today - dr).days} j.'
            elif dr == today:
                echeance = "Échue aujourd'hui"
            else:
                echeance = f'Dans {(dr - today).days} j.'

            fil_dict = result.setdefault(filiale, {})
            expl_dict = fil_dict.setdefault(expl, {'clients_pp': [], 'clients_pm': []})

            expl_dict[key_label].append({
                'client': obj['CLIENT'],
                'agence': obj['AGENCE'],
                'lib_agence': obj['LIB_AGENCE'],
                'idpm': obj[id_field],
                'daterev': dr,
                'echeance': echeance,
                'risque': obj['RISQUE'],
                'statut': 'Dépassée' if dr < today else f'Dans {(dr - today).days} j.',
            })

    return result


@login_required
def daterev_reminder(request):
    from accounts.models import ProfileV
    if request.user.organe not in ("PASS", "DSI"):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    config = EmailReminderConfig.objects.filter(active=True).order_by('-updated_at').first()
    days_before = config.days_before if config else 30

    filiale_filter = request.GET.get('filiale', '').strip()

    from django.core.cache import cache
    from django.utils import timezone
    import hashlib
    _fil_slug = hashlib.md5((filiale_filter or 'all').encode()).hexdigest()[:12]
    _cache_key = f"drevpaid:{_fil_slug}:{timezone.localdate().isoformat()}:{days_before}"
    raw = cache.get(_cache_key)
    if raw is None:
        raw = _get_exploitants_daterev_expired(filiale_filter or None, days_before, only_paid=True)
        cache.set(_cache_key, raw, timeout=3600)

                                                                                
    all_profiles = {}
    for p in ProfileV.objects.exclude(code_expl='').exclude(code_expl__isnull=True):
        key = (p.filiale.strip().upper(), p.code_expl.strip().upper())
        all_profiles[key] = p

                                                   
    data = []
    total_exploitants = 0
    total_clients = 0

    for filiale in sorted(raw.keys()):
        expls_data = raw[filiale]
        exploitants_list = []

        for expl in sorted(expls_data.keys()):
            clients = expls_data[expl]
            user_obj = all_profiles.get((filiale.strip().upper(), expl.strip().upper()))

            count_pp = len(clients['clients_pp'])
            count_pm = len(clients['clients_pm'])
            total = count_pp + count_pm

            exploitants_list.append({
                'expl': expl,
                'user': user_obj,
                'count_pp': count_pp,
                'count_pm': count_pm,
                'total': total,
                'sent': False,
                'clients_pp': sorted(clients['clients_pp'], key=lambda x: x['daterev']),
                'clients_pm': sorted(clients['clients_pm'], key=lambda x: x['daterev']),
            })
            total_exploitants += 1
            total_clients += total

        if exploitants_list:
            data.append({'filiale': filiale, 'exploitants': exploitants_list})

                                                                                  
    paid_filiales = _paid_daterev_filiales()
    filiales_list = sorted(filter(None, set(
        list(Kyc_pp.objects.values_list('FILIALE', flat=True).distinct()) +
        list(Kyc_pm.objects.values_list('FILIALE', flat=True).distinct())
    ) & paid_filiales))

    return render(request, 'daterev_reminder.html', {
        'data': data,
        'config': config,
        'filiale_filter': filiale_filter,
        'filiales_list': filiales_list,
        'days_before': days_before,
        'total_exploitants': total_exploitants,
        'total_clients': total_clients,
    })


@login_required
def send_daterev_reminders(request):
    from kyc.daterev_mailer import send_daterev_reminders_core

    if request.method != 'POST':
        return redirect('daterev_reminder')
    if request.user.organe not in ("PASS", "DSI"):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    config = EmailReminderConfig.objects.filter(active=True).order_by('-updated_at').first()
    if not config:
        messages.error(request, "Aucune configuration SMTP active trouvée. Configurez-la dans l'admin.")
        return redirect('daterev_reminder')

    filiale_filter = request.POST.get('filiale', '')
    expl_filter = request.POST.get('expl', '')
                                                                       
    data = _get_exploitants_daterev_expired(filiale_filter or None, config.days_before, only_paid=True)
    if not data:
        messages.warning(request, "Aucun envoi : aucune filiale payée concernée (vérifiez le statut « Rappels DATEREV PAYE » dans l'admin).")
        return redirect('daterev_reminder')

    try:
        sent, skipped = send_daterev_reminders_core(config, filiale=filiale_filter or None,
                                                    expl=expl_filter or None, only_paid=True)
    except Exception as e:
        messages.error(request, f"Erreur SMTP : {e}")
        return redirect('daterev_reminder')

    if sent:
        messages.success(request, f"{sent} mail(s) envoyé(s) avec succès.")
    if skipped:
        messages.warning(request, f"{skipped} exploitant(s) ignoré(s) — email non trouvé.")
    if not sent and not skipped:
        messages.info(request, "Aucun rappel à envoyer pour ce périmètre.")

    return redirect('daterev_reminder')


@login_required
def test_smtp_config(request):
    import smtplib
    from email.mime.text import MIMEText

    if request.user.organe not in ("PASS", "DSI"):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    config = EmailReminderConfig.objects.filter(active=True).order_by('-updated_at').first()
    if not config:
        messages.error(request, "Aucune configuration SMTP active. Ajoutez-en une dans l'admin Django.")
        return redirect('daterev_reminder')

    test_recipient = request.POST.get('test_email') or request.user.email

                                     
    if config.smtp_use_ssl and config.smtp_use_tls:
        messages.error(request, "Configuration invalide : SSL et TLS ne peuvent pas être activés simultanément. Port 465 → SSL uniquement. Port 587 → TLS uniquement.")
        return redirect('daterev_reminder')
    if config.smtp_port == 465 and config.smtp_use_tls and not config.smtp_use_ssl:
        messages.error(request, "Configuration suspecte : port 465 nécessite SSL (pas TLS/STARTTLS). Activez 'Utiliser SSL' et désactivez 'Utiliser TLS' dans l'admin.")
        return redirect('daterev_reminder')
    if config.smtp_port == 587 and config.smtp_use_ssl and not config.smtp_use_tls:
        messages.error(request, "Configuration suspecte : port 587 nécessite TLS/STARTTLS (pas SSL direct). Activez 'Utiliser TLS' et désactivez 'Utiliser SSL' dans l'admin.")
        return redirect('daterev_reminder')

    def _build_server(cfg):
        if cfg.smtp_use_ssl:
            srv = smtplib.SMTP_SSL(cfg.smtp_host, cfg.smtp_port, timeout=15)
        else:
            srv = smtplib.SMTP(cfg.smtp_host, cfg.smtp_port, timeout=15)
            if cfg.smtp_use_tls:
                srv.ehlo()
                srv.starttls()
                srv.ehlo()
        return srv

    auth_used = False
    try:
        server = _build_server(config)
        if config.smtp_user and config.smtp_password:
            try:
                server.login(config.smtp_user, config.smtp_password)
                auth_used = True
            except smtplib.SMTPAuthenticationError:
                                                                         
                server.quit()
                server = _build_server(config)

        msg = MIMEText(
            "<h3>Test SMTP — KYC Portal BOA</h3><p>La configuration SMTP fonctionne correctement.</p>"
            f"<p style='color:#666;font-size:12px'>Serveur : {config.smtp_host}:{config.smtp_port} · "
            f"Mode : {'SSL' if config.smtp_use_ssl else 'STARTTLS' if config.smtp_use_tls else 'Non chiffré'} · "
            f"Auth : {'oui' if auth_used else 'relais sans auth'}</p>",
            'html', 'utf-8'
        )
        msg['Subject'] = " Test de configuration SMTP"
        msg['From'] = f"{config.from_name} <{config.from_email}>"
        msg['To'] = test_recipient
        server.sendmail(config.from_email, test_recipient, msg.as_string())
        server.quit()
        auth_note = "avec authentification" if auth_used else "sans authentification (relais interne)"
        messages.success(request, f"✓ Email envoyé à {test_recipient} via {config.smtp_host}:{config.smtp_port} — {auth_note}.")
    except smtplib.SMTPConnectError as e:
        messages.error(request, f"Échec SMTP : impossible de se connecter à {config.smtp_host}:{config.smtp_port} — {e}")
    except Exception as e:
        err = str(e)
        hint = ""
        if "WRONG_VERSION_NUMBER" in err:
            hint = " → Mauvaise combinaison port/SSL : port 587 = TLS, port 465 = SSL."
        elif "Connection refused" in err:
            hint = f" → Le serveur {config.smtp_host}:{config.smtp_port} refuse la connexion."
        elif "timed out" in err.lower():
            hint = " → Délai dépassé : vérifiez que le port SMTP n'est pas bloqué par un pare-feu."
        messages.error(request, f"Échec SMTP : {e}{hint}")

    return redirect('daterev_reminder')


@login_required
def export_daterev_excel(request):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    if request.user.organe not in ("PASS", "DSI"):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    config = EmailReminderConfig.objects.filter(active=True).order_by('-updated_at').first()
    days_before = config.days_before if config else 30
    filiale_filter = request.GET.get('filiale', '').strip()
    expl_filter = request.GET.get('expl', '').strip()

    from django.core.cache import cache
    from django.utils import timezone
    import hashlib
    _fil_slug = hashlib.md5((filiale_filter or 'all').encode()).hexdigest()[:12]
    _cache_key = f"drev:{_fil_slug}:{timezone.localdate().isoformat()}:{days_before}"
    raw = cache.get(_cache_key) or _get_exploitants_daterev_expired(filiale_filter or None, days_before)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rappels DATEREV"

    green_fill = PatternFill("solid", fgColor="0a3d2e")
    orange_fill = PatternFill("solid", fgColor="FFF3CD")
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    normal = Font(size=9)
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    headers = ['Filiale', 'Exploitant', 'Client', 'Agence', 'Lib. Agence', 'DATEREV', 'Type', 'Statut']
    ws.append(headers)
    for i, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i)
        cell.font = header_font
        cell.fill = green_fill
        cell.alignment = center
        cell.border = thin

    row_num = 2
    for filiale in sorted(raw.keys()):
        for expl in sorted(raw[filiale].keys()):
            if expl_filter and expl.strip().upper() != expl_filter.upper():
                continue
            clients = raw[filiale][expl]
            for c in sorted(clients['clients_pp'], key=lambda x: x['daterev']):
                ws.append([filiale, expl, c['client'], c['agence'], c.get('lib_agence', ''), str(c['daterev']), 'PP', c['statut']])
                fill = red_fill if c['statut'] == 'Dépassée' else orange_fill
                for col in range(1, 9):
                    cell = ws.cell(row=row_num, column=col)
                    cell.font = normal
                    cell.border = thin
                    if col == 8:
                        cell.fill = fill
                row_num += 1
            for c in sorted(clients['clients_pm'], key=lambda x: x['daterev']):
                ws.append([filiale, expl, c['client'], c['agence'], c.get('lib_agence', ''), str(c['daterev']), 'PM', c['statut']])
                fill = red_fill if c['statut'] == 'Dépassée' else orange_fill
                for col in range(1, 9):
                    cell = ws.cell(row=row_num, column=col)
                    cell.font = normal
                    cell.border = thin
                    if col == 8:
                        cell.fill = fill
                row_num += 1

    for col, width in zip('ABCDEFGH', [14, 12, 14, 10, 22, 12, 6, 14]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname_parts = [p for p in [filiale_filter, expl_filter] if p]
    fname = f"daterev_{'_'.join(fname_parts) if fname_parts else 'global'}_{timezone.localdate()}.xlsx"
    fname = fname.replace(' ', '_')
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


